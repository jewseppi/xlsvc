"""
Produce an Apple Numbers-compatible copy of a LibreOffice-exported .xlsx.

LibreOffice's XLSX export trips Apple Numbers' (strict) importer in several
independent ways — its style / number-format table and its worksheet XML — so
Numbers refuses the file with "couldn't read the file", even though Excel,
Google Sheets and LibreOffice all open it. (Numbers opens workbooks written by
openpyxl without issue.)

This module rebuilds the workbook through openpyxl (clean worksheet and style
XML that Numbers accepts) and then grafts LibreOffice's original media and
drawings back in by raw byte copy. Because the images are copied verbatim rather
than re-encoded, formats Pillow can't handle (wdp / wmf) survive, and the
drawing anchors still reference the same cells, so the product images land in
the right place.

Cell formatting is preserved as far as Numbers allows: fills (background
colours), borders, alignment and font weight/size/colour are carried over. The
two things Numbers rejects are sanitised rather than dropped wholesale — the
font *name* is normalised to Arial (LibreOffice's East Asian font entries are
the offender, not the bold/size/colour) and bracketed number-format codes such
as [RED] become General. The untouched original file remains the full-fidelity
copy for Excel / LibreOffice.
"""
import copy
import io
import re
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font

from processing_helpers import load_workbook_resilient

_NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
_R_ID = "{%s}id" % _NS["r"]
_DRAWING_REL_ID = "rIdNumbersDrawing"

_CONTENT_TYPES = {
    "png": "image/png",
    "jpeg": "image/jpeg",
    "jpg": "image/jpeg",
    "gif": "image/gif",
    "bmp": "image/bmp",
    "tif": "image/tiff",
    "tiff": "image/tiff",
    "emf": "image/x-emf",
    "wmf": "image/x-wmf",
    "wdp": "image/vnd.ms-photo",
}


def _safe_number_format(fmt):
    """Drop number-format codes Numbers can't parse (colour/conditional codes
    in brackets, e.g. [RED] or [$-409]); keep simple ones like 0.00 or 0%."""
    if fmt and "[" not in fmt:
        return fmt
    return "General"


def _rebuild_clean(src_path):
    """Rebuild a Numbers-clean workbook (clean worksheet/style XML) carrying
    cached values plus the safe parts of cell formatting, and return its bytes.

    Carried: fills (background colours), borders, alignment, and fonts
    (weight/size/colour) with the font *name* normalised to Arial — LibreOffice's
    original font entries (e.g. East Asian faces with non-Latin charsets) are
    what Numbers rejects. Number formats are sanitised by _safe_number_format.
    """
    src = load_workbook_resilient(src_path, data_only=True)
    out = Workbook()
    out.remove(out.active)
    for name in src.sheetnames:
        ws = out.create_sheet(title=name[:31])
        s = src[name]
        for row in s.iter_rows():
            for c in row:
                nc = ws.cell(row=c.row, column=c.column, value=c.value)
                if c.has_style:
                    f = c.font
                    nc.font = Font(
                        name="Arial", size=f.size, bold=f.bold, italic=f.italic,
                        underline=f.underline, strike=f.strike, color=f.color,
                    )
                    nc.fill = copy.copy(c.fill)
                    nc.border = copy.copy(c.border)
                    nc.alignment = copy.copy(c.alignment)
                    nc.number_format = _safe_number_format(c.number_format)
    if not out.sheetnames:  # pragma: no cover - never leave a zero-sheet book
        out.create_sheet(title="Sheet1")
    src.close()
    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)
    return buf.read()


def _sheet_name_to_drawing(zf):
    """Map sheet name -> drawing part filename (e.g. 'drawing3.xml')."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {r.get("Id"): r.get("Target") for r in rels}
    mapping = {}
    for sheet in wb.find("m:sheets", _NS):
        target = rid_to_target.get(sheet.get(_R_ID), "")
        base = target.split("/")[-1]
        rel_path = "xl/worksheets/_rels/%s.rels" % base
        if rel_path not in zf.namelist():
            continue
        for rel in ET.fromstring(zf.read(rel_path)):
            if "drawing" in rel.get("Target", ""):
                mapping[sheet.get("name")] = rel.get("Target").split("/")[-1]
                break
    return mapping


def _sheet_name_to_part(zf):
    """Map sheet name -> worksheet part filename (e.g. 'sheet1.xml')."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {r.get("Id"): r.get("Target") for r in rels}
    return {
        sheet.get("name"): rid_to_target[sheet.get(_R_ID)].split("/")[-1]
        for sheet in wb.find("m:sheets", _NS)
    }


def to_numbers_compatible(src_path, dst_path):
    """Write a Numbers-compatible copy of ``src_path`` to ``dst_path``.

    Returns ``dst_path`` on success. Raises on failure; callers should treat
    the conversion as best-effort and keep the original file regardless.
    """
    clean_bytes = _rebuild_clean(src_path)

    with zipfile.ZipFile(src_path) as zlo, \
            zipfile.ZipFile(io.BytesIO(clean_bytes)) as zclean:
        name_to_drawing = _sheet_name_to_drawing(zlo)
        clean_name_to_part = _sheet_name_to_part(zclean)

        media = [n for n in zlo.namelist() if n.startswith("xl/media/")]
        drawings = [n for n in zlo.namelist() if n.startswith("xl/drawings/")]

        # map worksheet part filename -> drawing filename (truncated-name safe)
        part_to_drawing = {}
        for name, drawing in name_to_drawing.items():
            part = clean_name_to_part.get(name[:31])
            if part:
                part_to_drawing[part] = drawing

        with zipfile.ZipFile(dst_path, "w", zipfile.ZIP_DEFLATED) as out:
            for item in zclean.infolist():
                n = item.filename
                data = zclean.read(n)
                if n == "[Content_Types].xml" and (media or part_to_drawing):
                    data = _augment_content_types(data, media, drawings)
                elif (re.match(r"xl/worksheets/sheet\d+\.xml$", n)
                        and n.split("/")[-1] in part_to_drawing):
                    data = _attach_drawing_ref(data)
                out.writestr(item, data)

            # per-worksheet relationship pointing at the grafted drawing
            for part, drawing in part_to_drawing.items():
                out.writestr(
                    "xl/worksheets/_rels/%s.rels" % part,
                    _worksheet_drawing_rels(drawing),
                )

            # raw-copy LibreOffice media + drawings (preserves wdp/wmf bytes)
            for n in media + drawings:
                out.writestr(n, zlo.read(n))

    return dst_path


def _attach_drawing_ref(worksheet_xml):
    """Insert a <drawing> element (and the r: namespace) into a worksheet."""
    head = worksheet_xml.split(b">", 1)[0]
    if b"xmlns:r=" not in head:
        worksheet_xml = worksheet_xml.replace(
            b"<worksheet ",
            b'<worksheet xmlns:r="%s" ' % _NS["r"].encode(),
            1,
        )
    drawing_el = b'<drawing r:id="%s"/>' % _DRAWING_REL_ID.encode()
    return worksheet_xml.replace(b"</worksheet>", drawing_el + b"</worksheet>", 1)


def _worksheet_drawing_rels(drawing_filename):
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="%s" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        'Target="../drawings/%s"/></Relationships>' % (_DRAWING_REL_ID, drawing_filename)
    )


def _augment_content_types(content_types_xml, media, drawings):
    """Add image Default entries and drawing Override entries.

    Image extensions never collide with the rels/xml Defaults openpyxl already
    writes, so the new Default entries can be appended directly.
    """
    text = content_types_xml.decode("utf-8")
    additions = []
    for ext in sorted({n.rsplit(".", 1)[-1].lower() for n in media if "." in n}):
        ctype = _CONTENT_TYPES.get(ext, "application/octet-stream")
        additions.append('<Default Extension="%s" ContentType="%s"/>' % (ext, ctype))
    for drawing in drawings:
        if re.match(r"xl/drawings/drawing\d+\.xml$", drawing):
            additions.append(
                '<Override PartName="/%s" '
                'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                % drawing
            )
    return text.replace("</Types>", "".join(additions) + "</Types>").encode("utf-8")
