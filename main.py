from deletion_report import generate_deletion_report, capture_row_data
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS, cross_origin
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from db import get_db, init_db
from processing_helpers import evaluate_cell_value, is_empty_or_zero, column_to_index
from file_utils import calculate_file_hash, ensure_directories, get_file_path, allowed_file, validate_excel_file
from auth_helpers import request_counts, rate_limit, generate_download_token, verify_download_token, validate_password_strength, is_admin_user, validate_invitation_token
from cleanup import cleanup_old_files
from macro_generator import generate_libreoffice_macro, generate_instructions
from github_app import GitHubAppAuth
import sqlite3
import os
from datetime import datetime, timedelta
import uuid
from openpyxl import load_workbook
import hashlib
import requests
import secrets
import jwt as jwt_lib
import time
import json
import re

app = Flask(__name__)

# Configuration
app.config['JWT_SECRET_KEY'] = os.environ.get('SECRET_KEY')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['PROCESSED_FOLDER'] = 'processed'
app.config['MACROS_FOLDER'] = 'macros'
app.config['REPORTS_FOLDER'] = 'reports'
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

# Initialize extensions
CORS(app, origins=['http://localhost:5173', 'https://xlsvc.jsilverman.ca'])
jwt = JWTManager(app)

# Security headers middleware
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# Initialize database on startup
init_db()
ensure_directories(app)


# Newsletter subscription endpoint (used by landing.html form)
@app.route('/api/subscribe', methods=['POST'])
@cross_origin(origins=['https://xlsvc.jsilverman.ca', 'http://localhost:5173'])
def subscribe():
    """Subscribe an email address to notifications/newsletter."""
    try:
        data = request.get_json() or {}
        email = str(data.get('email', '')).strip().lower()

        # Basic email validation
        if not email or not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
            return jsonify({'error': 'Invalid email'}), 400

        conn = get_db()

        # Check if already subscribed
        existing = conn.execute(
            'SELECT id FROM subscribers WHERE email = ?',
            (email,)
        ).fetchone()

        if existing:
            conn.close()
            return jsonify({'error': 'Already subscribed'}), 409

        # Insert new subscriber
        conn.execute(
            'INSERT INTO subscribers (email, created_at) VALUES (?, ?)',
            (email, datetime.utcnow().isoformat())
        )
        conn.commit()
        conn.close()

        return jsonify({'success': True, 'message': 'Subscribed successfully'}), 200

    except Exception as e:
        print(f"Subscribe error: {e}")
        return jsonify({'error': 'Server error'}), 500

# Authentication endpoints
@app.route('/api/register', methods=['POST'])
@rate_limit(max_requests=5, window_seconds=300)  # 5 attempts per 5 minutes
def register():
    try:
        data = request.get_json()
        invitation_token = data.get('invitation_token')
        password = data.get('password')
        
        if not invitation_token or not password:
            return jsonify({
                'error': 'Invitation token and password are required. Registration is by invitation only.'
            }), 400
        
        # Validate invitation token
        is_valid, email, error_msg = validate_invitation_token(invitation_token)
        if not is_valid:
            return jsonify({'error': error_msg or 'Invalid invitation token'}), 400
        
        # Validate password strength
        is_strong, password_error = validate_password_strength(password)
        if not is_strong:
            return jsonify({'error': password_error}), 400
        
        # Hash password
        password_hash = generate_password_hash(password)
        
        conn = get_db()
        try:
            # Create user account
            conn.execute(
                'INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, 0)',
                (email, password_hash)
            )
            
            # Mark invitation token as used
            conn.execute(
                'UPDATE invitation_tokens SET used_at = ? WHERE token = ?',
                (datetime.utcnow().isoformat(), invitation_token)
            )
            
            conn.commit()
            
            # Create access token
            access_token = create_access_token(identity=email)
            
            return jsonify({
                'message': 'User registered successfully',
                'access_token': access_token,
                'email': email
            }), 201
            
        except sqlite3.IntegrityError:
            return jsonify({'error': 'Email already registered'}), 409
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Registration error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/login', methods=['POST'])
@rate_limit(max_requests=5, window_seconds=300)  # 5 attempts per 5 minutes
def login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': 'Email and password required'}), 400
        
        conn = get_db()
        user = conn.execute(
            'SELECT * FROM users WHERE email = ?', (email,)
        ).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            access_token = create_access_token(identity=email)
            return jsonify({
                'message': 'Login successful',
                'access_token': access_token
            }), 200
        else:
            return jsonify({'error': 'Invalid credentials'}), 401
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# File upload endpoint
@app.route('/api/upload', methods=['POST'])
@jwt_required()
def upload_file():
    try:
        current_user_email = get_jwt_identity()
        
        # Get user ID
        conn = get_db()
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        user_id = user['id']
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Only .xls and .xlsx files allowed'}), 400
        
        # Validate file content using magic bytes
        try:
            validate_excel_file(file)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400

        # Generate unique filename for storage
        original_filename = secure_filename(file.filename)
        file_extension = original_filename.rsplit('.', 1)[1].lower()
        stored_filename = f"{uuid.uuid4()}.{file_extension}"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], stored_filename)
        
        # Save file temporarily to calculate hash
        file.save(file_path)
        file_size = os.path.getsize(file_path)
        file_hash = calculate_file_hash(file_path)
        
        # Check if this exact file already exists for this user
        existing_file = conn.execute(
            '''SELECT id, original_filename FROM files 
               WHERE user_id = ? AND file_hash = ? AND original_filename = ?''',
            (user_id, file_hash, original_filename)
        ).fetchone()
        
        if existing_file:
            # Delete the newly uploaded duplicate
            os.remove(file_path)
            conn.close()
            
            return jsonify({
                'message': 'File already exists',
                'file_id': existing_file['id'],
                'filename': existing_file['original_filename'],
                'duplicate': True
            }), 200
        
        # Save to database with hash
        file_id = conn.execute(
            '''INSERT INTO files (user_id, original_filename, stored_filename, file_size, file_hash, file_type) 
            VALUES (?, ?, ?, ?, ?, ?)''',
            (user_id, original_filename, stored_filename, file_size, file_hash, 'original')
        ).lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'File uploaded successfully',
            'file_id': file_id,
            'filename': original_filename,
            'size': file_size,
            'duplicate': False
        }), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Get user's files
@app.route('/api/files', methods=['GET'])
@jwt_required()
def get_files():
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        
        # Get only ORIGINAL files
        files = conn.execute(
            '''SELECT f.* FROM files f
            JOIN users u ON f.user_id = u.id
            WHERE u.email = ? AND (f.file_type = 'original' OR f.file_type IS NULL)
            ORDER BY f.upload_date DESC''',
            (current_user_email,)
        ).fetchall()
    
        # Convert to list of dictionaries and check if files exist on disk
        valid_files = []
        for file in files:
            file_dict = dict(file)
            stored_filename = file_dict.get('stored_filename', '')
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], stored_filename)
            
            # Only include files that actually exist on disk
            if stored_filename and os.path.exists(file_path):
                valid_files.append(file_dict)
            else:
                print(f"DEBUG: File not found on disk: {file_path}")
        
        conn.close()
        
        return jsonify({
            'files': valid_files
        }), 200
        
    except Exception as e:
        print(f"DEBUG: get_files error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/process/<int:file_id>', methods=['POST'])
@jwt_required()
def process_file(file_id):
    try:
        current_user_email = get_jwt_identity()
        
        # Get filter rules from request body
        data = request.get_json() or {}
        filter_rules = data.get('filter_rules', [])
        
        if not filter_rules or len(filter_rules) == 0:
            return jsonify({
                'error': 'filter_rules required and must be a non-empty array'
            }), 400
        
        print(f"DEBUG: Manual processing with {len(filter_rules)} filter rules")
        
        # Get file info and verify ownership
        conn = get_db()
        file_info = conn.execute(
            '''SELECT f.* FROM files f
               JOIN users u ON f.user_id = u.id
               WHERE f.id = ? AND u.email = ?''',
            (file_id, current_user_email)
        ).fetchone()
        
        if not file_info:
            return jsonify({'error': 'File not found'}), 404
        
        file_dict = dict(file_info)
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], file_dict['stored_filename'])
        
        if not os.path.exists(input_path):
            return jsonify({'error': 'File not found on disk'}), 404
        
        # Analyze the Excel file
        processing_log = []
        rows_to_delete_by_sheet = {}
        deleted_rows_data = {}  # ← NEW: For report generation
        total_rows_to_delete = 0
        
        try:
            wb_calc = load_workbook(input_path, data_only=True)
            
            for sheet_name in wb_calc.sheetnames:
                sheet = wb_calc[sheet_name]
                processing_log.append(f"Analyzing sheet: {sheet_name}")
                
                rows_to_delete = []
                max_row = sheet.max_row
                
                for row_num in range(1, max_row + 1):
                    # Column A pre-filter: skip rows where column A is empty (parity with UNO)
                    col_a_value = sheet.cell(row=row_num, column=1).value
                    if col_a_value is None or str(col_a_value).strip() == '':
                        continue

                    # Check columns dynamically based on filter_rules
                    all_match = True
                    
                    for rule in filter_rules:
                        column = rule.get('column')
                        expected_value = rule.get('value')
                        col_index = column_to_index(column)
                        cell_val = sheet.cell(row=row_num, column=col_index).value
                        
                        # Always check for empty/zero (parity with UNO behavior)
                        if not is_empty_or_zero(cell_val):
                            all_match = False
                            break
                    
                    if all_match:
                        rows_to_delete.append(row_num)
                        
                        # ← NEW: Capture row data for report
                        if sheet_name not in deleted_rows_data:
                            deleted_rows_data[sheet_name] = []
                        
                        deleted_rows_data[sheet_name].append({
                            'row_number': row_num,
                            'data': capture_row_data(sheet, row_num)
                        })
                        
                        if len(rows_to_delete) <= 5:
                            processing_log.append(f"Row {row_num} marked for deletion")
                
                if len(rows_to_delete) > 5:
                    processing_log.append(f"... and {len(rows_to_delete) - 5} more rows")
                
                if rows_to_delete:
                    rows_to_delete_by_sheet[sheet_name] = rows_to_delete
                    total_rows_to_delete += len(rows_to_delete)
                    processing_log.append(f"Found {len(rows_to_delete)} rows to delete in '{sheet_name}'")
                else:
                    processing_log.append(f"No rows to delete in '{sheet_name}'")
            
            wb_calc.close()
            
            if total_rows_to_delete == 0:
                return jsonify({
                    'message': 'No rows found for deletion',
                    'total_rows_to_delete': 0,
                    'processing_log': processing_log + ["Analysis complete - no changes needed"]
                }), 200
            
            # Generate macro
            macro_content = generate_libreoffice_macro(
                file_dict['original_filename'], 
                rows_to_delete_by_sheet,
                filter_rules
            )
            
            macro_filename = f"macro_{uuid.uuid4().hex[:8]}.bas"
            macro_path = os.path.join(app.config['MACROS_FOLDER'], macro_filename)
            with open(macro_path, 'w', encoding='utf-8') as f:
                f.write(macro_content)
            
            # Generate instructions
            instructions = generate_instructions(
                file_dict['original_filename'],
                total_rows_to_delete,
                list(rows_to_delete_by_sheet.keys()),
                filter_rules
            )
            
            instructions_filename = f"instructions_{uuid.uuid4().hex[:8]}.txt"
            instructions_path = os.path.join(app.config['MACROS_FOLDER'], instructions_filename)
            with open(instructions_path, 'w', encoding='utf-8') as f:
                f.write(instructions)
            
            # Record macro and instructions in database
            macro_file_id = conn.execute(
                '''INSERT INTO files (user_id, original_filename, stored_filename, file_size, processed, file_type) 
                   VALUES (?, ?, ?, ?, ?, ?)''',
                (file_dict['user_id'], f"Macro_{file_dict['original_filename']}.bas", 
                 macro_filename, os.path.getsize(macro_path), True, 'macro')
            ).lastrowid
            
            instructions_file_id = conn.execute(
                '''INSERT INTO files (user_id, original_filename, stored_filename, file_size, processed, file_type)
                   VALUES (?, ?, ?, ?, ?, ?)''',
                (file_dict['user_id'], f"Instructions_{file_dict['original_filename']}.txt",
                 instructions_filename, os.path.getsize(instructions_path), True, 'instructions')
            ).lastrowid

            # Generate deletion report (parity with UNO automated path)
            report_file_id = None
            if deleted_rows_data:
                report_stored_filename = f"deletion_report_{uuid.uuid4().hex[:8]}.xlsx"
                report_path = os.path.join(app.config['REPORTS_FOLDER'], report_stored_filename)
                report_result = generate_deletion_report(deleted_rows_data, report_path)
                if report_result:
                    report_file_id = conn.execute(
                        '''INSERT INTO files (user_id, original_filename, stored_filename, file_size, processed, file_type, parent_file_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?)''',
                        (file_dict['user_id'], f"DeletionReport_{file_dict['original_filename']}.xlsx",
                         report_stored_filename, os.path.getsize(report_path), True, 'report', file_id)
                    ).lastrowid

            conn.execute('UPDATE files SET processed = TRUE WHERE id = ?', (file_id,))
            conn.commit()
            conn.close()
            
            processing_log.append("Analysis complete - all files generated")
            
            downloads = {
                'macro': {
                    'file_id': macro_file_id,
                    'filename': f"Macro_{file_dict['original_filename']}.bas"
                },
                'instructions': {
                    'file_id': instructions_file_id,
                    'filename': f"Instructions_{file_dict['original_filename']}.txt"
                }
            }
            if report_file_id:
                downloads['report'] = {
                    'file_id': report_file_id,
                    'filename': f"DeletionReport_{file_dict['original_filename']}.xlsx"
                }

            return jsonify({
                'message': 'Analysis complete',
                'total_rows_to_delete': total_rows_to_delete,
                'sheets_affected': list(rows_to_delete_by_sheet.keys()),
                'processing_log': processing_log,
                'report_file_id': report_file_id,
                'downloads': downloads
            }), 200
            
        except Exception as processing_error:
            processing_log.append(f"Analysis error: {str(processing_error)}")
            return jsonify({
                'error': f'Analysis failed: {str(processing_error)}',
                'processing_log': processing_log
            }), 500
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-with-token/<int:file_id>', methods=['GET'])
def download_file_with_token(file_id):
    """Download file using temporary token (for GitHub Actions)"""
    try:
        print(f"DEBUG: Download request for file_id={file_id}")
        
        # Get token from query parameter or Authorization header
        token = request.args.get('token')
        if not token:
            auth_header = request.headers.get('Authorization')
            if auth_header and auth_header.startswith('Bearer '):
                token = auth_header[7:]
        
        if not token:
            print("DEBUG: No token provided")
            return jsonify({'error': 'Download token required'}), 401
        
        # Verify the token
        token_payload = verify_download_token(token)
        if not token_payload:
            print("DEBUG: Token verification failed")
            return jsonify({'error': 'Invalid or expired download token'}), 401
        
        # Check if token is for this specific file
        if token_payload.get('file_id') != file_id:
            print(f"DEBUG: Token file_id mismatch: {token_payload.get('file_id')} != {file_id}")
            return jsonify({'error': 'Token not valid for this file'}), 403
        
        # Get file info
        conn = get_db()
        file_info = conn.execute(
            '''SELECT * FROM files WHERE id = ? AND user_id = ?''',
            (file_id, token_payload.get('user_id'))
        ).fetchone()
        conn.close()
        
        if not file_info:
            print(f"DEBUG: File not found in database: file_id={file_id}, user_id={token_payload.get('user_id')}")
            return jsonify({'error': 'File not found'}), 404
        
        # Get correct file path based on type
        file_dict = dict(file_info)
        file_path = get_file_path(
            file_dict.get('file_type'),
            file_dict['stored_filename']
        )
        
        if not os.path.exists(file_path):
            print(f"DEBUG: File not found on disk: {file_path}")
            return jsonify({'error': 'File not found on disk'}), 404
        
        file_size = os.path.getsize(file_path)
        print(f"DEBUG: Serving file via token: {file_path} (size: {file_size} bytes)")
        
        # Determine MIME type
        original_filename = file_info['original_filename']
        if original_filename.endswith(('.xlsx', '.xls')):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mimetype = 'application/octet-stream'
        
        # Stream the file with explicit headers
        response = send_file(
            file_path,
            as_attachment=True,
            download_name=original_filename,
            mimetype=mimetype
        )
        
        # Add headers for better download handling
        response.headers['Content-Length'] = file_size
        response.headers['Cache-Control'] = 'no-cache'
        
        print(f"DEBUG: File response prepared, sending {file_size} bytes")
        return response
        
    except Exception as e:
        print(f"DEBUG: Token download error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/<int:file_id>', methods=['GET'])
@jwt_required()
def download_file(file_id):
    try:
        current_user_email = get_jwt_identity()
        
        # Get file info and verify ownership
        conn = get_db()
        file_info = conn.execute(
            '''SELECT f.* FROM files f
               JOIN users u ON f.user_id = u.id
               WHERE f.id = ? AND u.email = ?''',
            (file_id, current_user_email)
        ).fetchone()
        conn.close()
        
        if not file_info:
            return jsonify({'error': 'File not found'}), 404
        
        # Get correct file path based on type
        file_dict = dict(file_info)
        file_path = get_file_path(
            file_dict.get('file_type'),
            file_dict['stored_filename']
        )
        
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found on disk'}), 404
        
        # Determine MIME type based on file extension
        original_filename = file_info['original_filename']
        if original_filename.endswith('.bas'):
            mimetype = 'text/plain'
        elif original_filename.endswith('.txt'):
            mimetype = 'text/plain'
        elif original_filename.endswith(('.xlsx', '.xls')):
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            mimetype = 'application/octet-stream'
            
        return send_file(
            file_path,
            as_attachment=True,
            download_name=original_filename,
            mimetype=mimetype
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Add file cleanup and sync endpoints
@app.route('/api/cleanup-files', methods=['POST'])
@jwt_required()
def cleanup_files():
    """Remove database entries for files AND their jobs that no longer exist on disk"""
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        files = conn.execute(
            'SELECT id, stored_filename, file_type FROM files WHERE user_id = ?', 
            (user['id'],)
        ).fetchall()
        
        removed_count = 0
        
        # Clean up files that don't exist on disk
        for file in files:
            file_dict = dict(file)
            file_type = file_dict.get('file_type') or 'original'
            stored_filename = file_dict.get('stored_filename', '')
            
            if file_type == 'original':
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], stored_filename)
            elif file_type == 'processed':
                file_path = os.path.join(app.config['PROCESSED_FOLDER'], stored_filename)
            elif file_type in ['macro', 'instructions']:
                file_path = os.path.join(app.config['MACROS_FOLDER'], stored_filename)
            elif file_type == 'report':
                file_path = os.path.join(app.config['REPORTS_FOLDER'], stored_filename)
            else:
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], stored_filename)
            
            if not os.path.exists(file_path):
                print(f"DEBUG: Removing missing file from DB: {stored_filename}")
                
                # Delete related processing jobs FIRST (before deleting file)
                if file_type == 'processed':
                    conn.execute(
                        'DELETE FROM processing_jobs WHERE result_file_id = ?',
                        (file_dict['id'],)
                    )
                    print(f"DEBUG: Deleted processing job for file_id {file_dict['id']}")
                
                # Then delete the file record
                conn.execute('DELETE FROM files WHERE id = ?', (file_dict['id'],))
                removed_count += 1
        
        # NEW: Also clean up orphaned processing jobs (where result_file_id doesn't exist)
        orphaned_jobs = conn.execute(
            '''SELECT pj.job_id, pj.result_file_id 
               FROM processing_jobs pj
               WHERE pj.user_id = ? AND pj.result_file_id IS NOT NULL
               AND NOT EXISTS (
                   SELECT 1 FROM files f WHERE f.id = pj.result_file_id
               )''',
            (user['id'],)
        ).fetchall()

        for job in orphaned_jobs:
            print(f"DEBUG: Removing orphaned job {job['job_id']} with missing file_id {job['result_file_id']}")
            conn.execute('DELETE FROM processing_jobs WHERE job_id = ?', (job['job_id'],))
            removed_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Cleaned up {removed_count} missing files and orphaned jobs',
            'removed_count': removed_count
        }), 200
        
    except Exception as e:
        print(f"DEBUG: cleanup_files error: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Add admin endpoint to see file organization
@app.route('/api/debug/storage', methods=['GET'])
@jwt_required()
def debug_storage():
    """Debug endpoint to see file organization"""
    try:
        current_user_email = get_jwt_identity()
        
        # Get storage info
        storage_info = {
            'uploads': [],
            'processed': [],
            'macros': []
        }
        
        # List files in each directory
        if os.path.exists(app.config['UPLOAD_FOLDER']):
            storage_info['uploads'] = os.listdir(app.config['UPLOAD_FOLDER'])
        
        if os.path.exists(app.config['PROCESSED_FOLDER']):
            storage_info['processed'] = os.listdir(app.config['PROCESSED_FOLDER'])
            
        if os.path.exists(app.config['MACROS_FOLDER']):
            storage_info['macros'] = os.listdir(app.config['MACROS_FOLDER'])
        
        # Get database info
        conn = get_db()
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        db_files = conn.execute(
            '''SELECT id, original_filename, stored_filename, file_type, processed
               FROM files WHERE user_id = ?
               ORDER BY upload_date DESC''',
            (user['id'],)
        ).fetchall()
        
        conn.close()
        
        return jsonify({
            'user': current_user_email,
            'storage_folders': storage_info,
            'database_files': [dict(f) for f in db_files]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/test-github', methods=['GET'])
@jwt_required()
def test_github_auth():
    """Test GitHub App authentication"""
    try:
        print("DEBUG: Testing GitHub App authentication...")
        
        # Check environment variables
        app_id = os.getenv('GITHUB_APP_ID')
        private_key = os.getenv('GITHUB_PRIVATE_KEY', '').replace('\\n', '\n')
        installation_id = os.getenv('GITHUB_INSTALLATION_ID')
        github_repo = os.getenv('GITHUB_REPO', 'jewseppi/xlsvc')
        
        env_status = {
            'app_id': bool(app_id),
            'private_key': bool(private_key),
            'installation_id': bool(installation_id),
            'github_repo': github_repo
        }
        
        print(f"DEBUG: Environment variables: {env_status}")
        
        if not all([app_id, private_key, installation_id]):
            return jsonify({
                'status': 'error',
                'error': 'Missing required environment variables',
                'env_status': env_status
            }), 400
        
        # Test GitHub App authentication
        github_auth = GitHubAppAuth()
        
        # Test JWT generation
        app_token = github_auth.get_app_token()
        print(f"DEBUG: Generated app token: {app_token[:50]}...")
        
        # Test installation token
        installation_token = github_auth.get_installation_token()
        print(f"DEBUG: Generated installation token: {installation_token[:50]}...")
        
        # Test repository access
        headers = {
            'Authorization': f'Bearer {installation_token}',
            'Accept': 'application/vnd.github.v3+json',
            'User-Agent': 'Excel-Processor-App/1.0'
        }
        
        # Check if we can access the repository
        repo_url = f'https://api.github.com/repos/{github_repo}'
        print(f"DEBUG: Testing repo access: {repo_url}")
        
        repo_response = requests.get(repo_url, headers=headers, timeout=10)
        print(f"DEBUG: Repo access response: {repo_response.status_code}")
        
        if repo_response.status_code == 200:
            repo_data = repo_response.json()
            return jsonify({
                'status': 'success',
                'message': 'GitHub App authentication working',
                'env_status': env_status,
                'repo_access': True,
                'repo_name': repo_data.get('full_name'),
                'repo_private': repo_data.get('private'),
                'app_token_length': len(app_token),
                'installation_token_length': len(installation_token)
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'error': 'Cannot access repository',
                'env_status': env_status,
                'repo_response_code': repo_response.status_code,
                'repo_response_text': repo_response.text
            }), 400
            
    except Exception as e:
        print(f"DEBUG: GitHub test error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500
    
@app.route('/api/test-dispatch', methods=['POST'])
@jwt_required()
def test_dispatch():
    """Test GitHub repository dispatch specifically"""
    try:
        print("DEBUG: Testing GitHub repository dispatch...")
        
        # Get GitHub token
        github_auth = GitHubAppAuth()
        github_token = github_auth.get_installation_token()
        github_repo = os.getenv('GITHUB_REPO', 'jewseppi/xlsvc')
        
        print(f"DEBUG: Using repo: {github_repo}")
        print(f"DEBUG: Token length: {len(github_token)}")
        
        # Test payload - minimal test dispatch
        test_payload = {
            "event_type": "test-connection",
            "client_payload": {
                "test": True,
                "timestamp": datetime.utcnow().isoformat()
            }
        }
        
        headers = {
            'Authorization': f'Bearer {github_token}',
            'Accept': 'application/vnd.github.v3+json',
            'Content-Type': 'application/json',
            'User-Agent': 'Excel-Processor-App/1.0'
        }
        
        dispatch_url = f'https://api.github.com/repos/{github_repo}/dispatches'
        print(f"DEBUG: Dispatch URL: {dispatch_url}")
        print(f"DEBUG: Headers: {headers}")
        print(f"DEBUG: Payload: {test_payload}")
        
        # First, check if the repo exists and we have access
        repo_url = f'https://api.github.com/repos/{github_repo}'
        repo_response = requests.get(repo_url, headers=headers, timeout=10)
        print(f"DEBUG: Repo check - Status: {repo_response.status_code}")
        
        if repo_response.status_code != 200:
            return jsonify({
                'status': 'error',
                'error': 'Cannot access repository',
                'repo_status': repo_response.status_code,
                'repo_response': repo_response.text
            }), 400
        
        # Try the dispatch
        response = requests.post(
            dispatch_url,
            headers=headers,
            json=test_payload,
            timeout=30
        )
        
        print(f"DEBUG: Dispatch response status: {response.status_code}")
        print(f"DEBUG: Dispatch response headers: {dict(response.headers)}")
        print(f"DEBUG: Dispatch response text: {response.text}")
        
        # GitHub API permissions check
        if response.status_code == 403:
            # Check what permissions we actually have
            permissions_url = f'https://api.github.com/repos/{github_repo}/installation'
            perm_response = requests.get(permissions_url, headers=headers, timeout=10)
            
            print(f"DEBUG: Permissions check status: {perm_response.status_code}")
            if perm_response.status_code == 200:
                perm_data = perm_response.json()
                print(f"DEBUG: Installation permissions: {perm_data}")
            
        result = {
            'status': 'success' if response.status_code == 204 else 'error',
            'dispatch_status': response.status_code,
            'dispatch_response': response.text,
            'dispatch_headers': dict(response.headers),
            'expected_status': 204,
            'repo_access': True,
            'github_repo': github_repo,
            'test_payload': test_payload
        }
        
        if response.status_code == 204:
            result['message'] = 'Repository dispatch successful'
        else:
            result['error'] = f'Dispatch failed with status {response.status_code}'
            
        return jsonify(result), 200 if response.status_code == 204 else 400
        
    except Exception as e:
        print(f"DEBUG: Test dispatch error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return jsonify({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500

@app.route('/api/process-automated/<int:file_id>', methods=['POST'])
@jwt_required()
def process_file_automated(file_id):
    """Trigger GitHub Actions processing using GitHub App authentication"""
    print(f"DEBUG: Starting automated processing for file {file_id}")

    try:
        current_user_email = get_jwt_identity()
        print(f"DEBUG: User: {current_user_email}")
        
        # Get filter rules from request body
        data = request.get_json() or {}
        filter_rules = data.get('filter_rules')
        
        # Validate filter rules
        if not filter_rules or not isinstance(filter_rules, list) or len(filter_rules) == 0:
            return jsonify({
                'error': 'filter_rules required and must be a non-empty array'
            }), 400
        
        print(f"DEBUG: Processing with {len(filter_rules)} filter rules")
        for i, rule in enumerate(filter_rules, 1):
            print(f"DEBUG: Rule {i}: Column {rule.get('column')} = '{rule.get('value')}'")
        
        # Get file info and verify ownership
        conn = get_db()
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        if not user:
            print("DEBUG: User not found in database")
            return jsonify({'error': 'User not found'}), 404
        
        print(f"DEBUG: User ID: {user['id']}")
        
        file_info = conn.execute(
            '''SELECT f.* FROM files f
               WHERE f.id = ? AND f.user_id = ?''',
            (file_id, user['id'])
        ).fetchone()
        
        if not file_info:
            print("DEBUG: File not found or not owned by user")
            return jsonify({'error': 'File not found'}), 404
            
        print(f"DEBUG: File found: {file_info['original_filename']}")
        
        # Check if file exists on disk
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], file_info['stored_filename'])
        if not os.path.exists(input_path):
            print(f"DEBUG: File not found on disk: {input_path}")
            return jsonify({'error': 'File not found on disk'}), 404
        
        file_dict = dict(file_info)
        
        # Create job ID and callback token
        job_id = secrets.token_urlsafe(16)
        callback_token = secrets.token_urlsafe(32)
        
        # Generate download token for the file
        download_token = generate_download_token(file_id, user['id'], expires_in_minutes=30)
        
        print(f"DEBUG: Generated job_id: {job_id}")
        print(f"DEBUG: Generated download token: {download_token[:20]}...")
        
        # Store job in database WITH filter rules
        conn.execute(
            '''INSERT INTO processing_jobs (job_id, user_id, original_file_id, status, filter_rules_json)
               VALUES (?, ?, ?, ?, ?)''',
            (job_id, user['id'], file_id, 'pending', json.dumps(filter_rules))
        )
        conn.commit()
        conn.close()
        
        print("DEBUG: Job stored in database with filter rules")
        
        # Get GitHub credentials
        app_id = os.getenv('GITHUB_APP_ID')
        private_key = os.getenv('GITHUB_PRIVATE_KEY', '').replace('\\n', '\n')
        installation_id = os.getenv('GITHUB_INSTALLATION_ID')
        github_repo = os.getenv('GITHUB_REPO', 'jewseppi/xlsvc')
        
        if not app_id or not private_key or not installation_id:
            return jsonify({
                'error': 'GitHub App configuration missing'
            }), 500
        
        # Get GitHub token
        github_auth = GitHubAppAuth()
        github_token = github_auth.get_installation_token()
        
        # Prepare GitHub Actions payload with filter_rules
        base_url = request.host_url.rstrip('/')
        github_payload = {
            "event_type": "process-excel",
            "client_payload": {
                "file_id": str(file_id),
                "file_url": f"{base_url}/api/download-with-token/{file_id}?token={download_token}",
                "download_token": download_token,
                "callback_url": f"{base_url}/api/processing-callback",
                "callback_token": callback_token,
                "job_id": job_id,
                "filter_rules": json.dumps(filter_rules)
            }
        }
        
        headers = {
            'Authorization': f'Bearer {github_token}',
            'Accept': 'application/vnd.github.v3+json',
            'Content-Type': 'application/json'
        }
        
        dispatch_url = f'https://api.github.com/repos/{github_repo}/dispatches'
        response = requests.post(
            dispatch_url,
            headers=headers,
            json=github_payload,
            timeout=30
        )
        
        if response.status_code == 204:
            return jsonify({
                'message': 'Processing started via GitHub Actions',
                'job_id': job_id,
                'status': 'pending',
                'estimated_time': '2-3 minutes'
            }), 202
        else:
            return jsonify({
                'error': 'Failed to start GitHub Actions processing',
                'details': response.text
            }), 500
            
    except Exception as e:
        print(f"DEBUG: Exception in process_file_automated: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e)
        }), 500

@app.route('/api/processing-callback', methods=['POST'])
def processing_callback():
    """Receive results from GitHub Actions"""
    try:
        print("=== PROCESSING CALLBACK RECEIVED ===")
        
        # Check authentication
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({'error': 'Unauthorized'}), 401
        
        if request.content_type and 'application/json' in request.content_type:
            # Status update (failure notification)
            data = request.get_json()
            job_id = data.get('job_id')
            status = data.get('status')
            error = data.get('error')
            
            conn = get_db()
            if status == 'failed':
                conn.execute(
                    '''UPDATE processing_jobs 
                       SET status = ?, error_message = ?
                       WHERE job_id = ?''',
                    ('failed', error, job_id)
                )
            conn.commit()
            conn.close()
            
            return jsonify({'status': 'updated'}), 200
        
        # File upload (success)
        job_id = request.form.get('job_id')
        uploaded_file = request.files.get('file')
        report_file = request.files.get('report')  # ← NEW
        deleted_rows = request.form.get('deleted_rows', 0)
        
        if not job_id or not uploaded_file:
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Get job info
        conn = get_db()
        job = conn.execute(
            'SELECT * FROM processing_jobs WHERE job_id = ?',
            (job_id,)
        ).fetchone()
        
        if not job:
            conn.close()
            return jsonify({'error': 'Job not found'}), 404
        
        job = dict(job)
        
        # Get original file info
        original_file = conn.execute(
            'SELECT * FROM files WHERE id = ?',
            (job['original_file_id'],)
        ).fetchone()
        
        if not original_file:
            conn.close()
            return jsonify({'error': 'Original file not found'}), 404
        
        original_filename = original_file['original_filename']
        
        # Save processed file
        file_extension = 'xlsx'
        processed_filename = f"processed_{original_filename}"
        output_filename = f"processed_{uuid.uuid4().hex[:8]}.{file_extension}"
        output_path = os.path.join(app.config['PROCESSED_FOLDER'], output_filename)
        
        uploaded_file.save(output_path)
        file_size = os.path.getsize(output_path)
        
        print(f"Processed file saved: {output_path}, size: {file_size} bytes")
        
        # Create file record for processed file
        file_id = conn.execute(
            '''INSERT INTO files (user_id, original_filename, stored_filename, file_size, processed, file_type, parent_file_id) 
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (job['user_id'], processed_filename, output_filename, file_size, True, 'processed', job['original_file_id'])
        ).lastrowid
        
        # ← NEW: Save deletion report if provided
        report_file_id = None
        if report_file:
            report_filename = f"deletion_report_{uuid.uuid4().hex[:8]}.xlsx"
            report_path = os.path.join(app.config['REPORTS_FOLDER'], report_filename)
            
            report_file.save(report_path)
            report_size = os.path.getsize(report_path)
            print(f"Deletion report saved: {report_path}, size: {report_size} bytes")
            
            # Create file record for report
            report_file_id = conn.execute(
                '''INSERT INTO files (user_id, original_filename, stored_filename, file_size, processed, file_type, parent_file_id) 
                   VALUES (?, ?, ?, ?, ?, ?, ?)''',
                (job['user_id'], f"DeletionReport_Automated_{original_filename}", report_filename, report_size, True, 'report', job['original_file_id'])
            ).lastrowid
        
        # Update job status with report_file_id
        conn.execute(
            '''UPDATE processing_jobs 
               SET status = ?, result_file_id = ?, deleted_rows = ?, report_file_id = ?
               WHERE job_id = ?''',
            ('completed', file_id, deleted_rows, report_file_id, job_id)
        )
        
        conn.commit()
        conn.close()
        
        print(f"=== CALLBACK SUCCESSFUL ===")
        print(f"Job {job_id} completed, file_id={file_id}, report_file_id={report_file_id}, deleted_rows={deleted_rows}")
        
        # Delete GitHub artifact after successful upload
        try:
            github_auth = GitHubAppAuth()
            github_auth.delete_artifact_by_job_id(job_id)
        except Exception as e:
            print(f"DEBUG: Failed to delete GitHub artifact (non-critical): {str(e)}")
            # Don't fail the callback if artifact deletion fails
        
        return jsonify({
            'status': 'success',
            'file_id': file_id,
            'filename': processed_filename,
            'report_file_id': report_file_id
        }), 200
        
    except Exception as e:
        print(f"=== CALLBACK ERROR ===")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/<int:file_id>/history', methods=['GET'])
@jwt_required()
def get_file_history(file_id):
    """Get all processed versions of a file"""
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        
        # Verify user owns the original file
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        if not user:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        
        # Check if the file exists and belongs to the user
        original_file = conn.execute(
            '''SELECT * FROM files 
               WHERE id = ? AND user_id = ? AND (file_type = 'original' OR file_type IS NULL)''',
            (file_id, user['id'])
        ).fetchone()
        
        if not original_file:
            conn.close()
            return jsonify({'error': 'File not found'}), 404
        
        # Get all processing jobs for this file with their results
        history = conn.execute(
            '''SELECT 
                pj.job_id,
                pj.created_at as processed_at,
                pj.status,
                pj.deleted_rows,
                pj.filter_rules_json,
                pj.result_file_id,
                pj.report_file_id,
                f.original_filename as processed_filename,
                f.file_size as processed_file_size
            FROM processing_jobs pj
            LEFT JOIN files f ON pj.result_file_id = f.id
            WHERE pj.original_file_id = ? AND pj.user_id = ?
            ORDER BY pj.created_at DESC''',
            (file_id, user['id'])
        ).fetchall()
        
        conn.close()
        
        # Format the history
        history_list = []
        for job in history:
            job_dict = dict(job)
            
            # Parse filter rules if present
            if job_dict.get('filter_rules_json'):
                try:
                    job_dict['filter_rules'] = json.loads(job_dict['filter_rules_json'])
                except:
                    job_dict['filter_rules'] = []
            else:
                job_dict['filter_rules'] = []
            
            history_list.append(job_dict)
        
        return jsonify({
            'original_file': dict(original_file),
            'history': history_list
        }), 200
        
    except Exception as e:
        print(f"ERROR: get_file_history: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/<int:file_id>/history/<job_id>', methods=['DELETE'])
@jwt_required()
def delete_history_item(file_id, job_id):
    """Delete a single processing job from history"""
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        if not user:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        
        # Verify user owns the file and the job
        job = conn.execute(
            '''SELECT pj.* FROM processing_jobs pj
               JOIN files f ON pj.original_file_id = f.id
               WHERE pj.job_id = ? AND pj.original_file_id = ? AND pj.user_id = ?''',
            (job_id, file_id, user['id'])
        ).fetchone()
        
        if not job:
            conn.close()
            return jsonify({'error': 'Job not found or access denied'}), 404
        
        # Delete the processing job
        conn.execute('DELETE FROM processing_jobs WHERE job_id = ?', (job_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'History item deleted successfully'}), 200
        
    except Exception as e:
        print(f"ERROR: delete_history_item: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/<int:file_id>/history', methods=['DELETE'])
@jwt_required()
def clear_file_history(file_id):
    """Clear all processing history for a file (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        user = conn.execute(
            'SELECT id, is_admin FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        if not user:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        
        # Check if user is admin
        if not user['is_admin']:
            conn.close()
            return jsonify({'error': 'Admin access required'}), 403
        
        # Verify the file exists
        file_info = conn.execute(
            'SELECT id FROM files WHERE id = ?', (file_id,)
        ).fetchone()
        
        if not file_info:
            conn.close()
            return jsonify({'error': 'File not found'}), 404
        
        # Delete all processing jobs for this file
        deleted_count = conn.execute(
            'DELETE FROM processing_jobs WHERE original_file_id = ?',
            (file_id,)
        ).rowcount
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': f'Cleared {deleted_count} history items',
            'deleted_count': deleted_count
        }), 200
        
    except Exception as e:
        print(f"ERROR: clear_file_history: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/job-status/<job_id>', methods=['GET'])
@jwt_required()
def get_job_status(job_id):
    """Check processing job status"""
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        job = conn.execute(
            '''SELECT pj.*, 
               f.original_filename as result_filename,
               rf.original_filename as report_filename
               FROM processing_jobs pj
               LEFT JOIN files f ON pj.result_file_id = f.id
               LEFT JOIN files rf ON pj.report_file_id = rf.id
               WHERE pj.job_id = ? AND pj.user_id = ?''',
            (job_id, user['id'])
        ).fetchone()
        conn.close()
        
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        
        response = {
            'job_id': job['job_id'],
            'status': job['status'],
            'created_at': job['created_at']
        }
        
        if job['status'] == 'completed':
            response['download_file_id'] = job['result_file_id']
            response['download_filename'] = job['result_filename']
            if job['report_file_id']:
                response['report_file_id'] = job['report_file_id']
                response['report_filename'] = job['report_filename']
        elif job['status'] == 'failed':
            response['error'] = job['error_message']
        
        return jsonify(response), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/get-macro/<int:file_id>', methods=['GET'])
def get_macro_for_file(file_id):
    """Get the macro file content for a processed file"""
    try:
        conn = get_db()
        
        # Get the original file info first
        original_file = conn.execute(
            'SELECT user_id, original_filename FROM files WHERE id = ?', 
            (file_id,)
        ).fetchone()
        
        if not original_file:
            return jsonify({'error': 'Original file not found'}), 404
        
        # Find the macro file for this original file
        macro_file = conn.execute(
            '''SELECT stored_filename, original_filename FROM files 
               WHERE user_id = ? AND file_type = 'macro'
               AND original_filename LIKE ?
               ORDER BY upload_date DESC LIMIT 1''',
            (original_file['user_id'], f'Macro_{original_file["original_filename"]}%')
        ).fetchone()
        
        conn.close()
        
        if not macro_file:
            return jsonify({'error': 'Macro not found for this file'}), 404
        
        # Read the macro file content
        macro_path = os.path.join(app.config['MACROS_FOLDER'], macro_file['stored_filename'])
        
        if not os.path.exists(macro_path):
            return jsonify({'error': 'Macro file not found on disk'}), 404
        
        with open(macro_path, 'r', encoding='utf-8') as f:
            macro_content = f.read()
        
        return jsonify({
            'macro_content': macro_content,
            'filename': macro_file['original_filename']
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/files/<int:file_id>/generated', methods=['GET'])
@jwt_required()
def get_generated_files(file_id):
    """Get all generated files (macros, instructions, reports) for a file"""
    try:
        current_user_email = get_jwt_identity()
        
        conn = get_db()
        
        # Verify user owns the file
        user = conn.execute(
            'SELECT id FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        
        if not user:
            conn.close()
            return jsonify({'error': 'User not found'}), 404
        
        # Check if file exists and belongs to user
        original_file = conn.execute(
            '''SELECT * FROM files 
               WHERE id = ? AND user_id = ? AND (file_type = 'original' OR file_type IS NULL)''',
            (file_id, user['id'])
        ).fetchone()
        
        if not original_file:
            conn.close()
            return jsonify({'error': 'File not found'}), 404
        
        # Get all generated files for this parent file
        # This includes files directly referencing this file_id in various ways
        
        # Get macros, instructions, reports, and processed files that have parent_file_id set
        generated_with_parent = conn.execute(
            '''SELECT id, original_filename, file_type, file_size, upload_date 
               FROM files 
               WHERE parent_file_id = ? AND user_id = ?
               AND file_type IN ('macro', 'instructions', 'report', 'processed')
               ORDER BY upload_date DESC''',
            (file_id, user['id'])
        ).fetchall()
        
        # Also get files that match the naming pattern (for older files without parent_file_id)
        original_filename = original_file['original_filename']
        generated_by_name = conn.execute(
            '''SELECT id, original_filename, file_type, file_size, upload_date 
               FROM files 
               WHERE user_id = ? 
               AND file_type IN ('macro', 'instructions', 'report', 'processed')
               AND (original_filename LIKE ? OR original_filename LIKE ? OR original_filename LIKE ? OR original_filename LIKE ?)
               AND (parent_file_id IS NULL OR parent_file_id != ?)
               ORDER BY upload_date DESC''',
            (user['id'], f'Macro_{original_filename}%', f'Instructions_{original_filename}%', 
             f'DeletionReport_{original_filename}%', f'processed_{original_filename}%', file_id)
        ).fetchall()
        
        conn.close()
        
        # Combine and deduplicate
        all_generated = list(generated_with_parent) + list(generated_by_name)
        seen_ids = set()
        unique_files = []
        
        for file in all_generated:
            if file['id'] not in seen_ids:
                seen_ids.add(file['id'])
                unique_files.append(dict(file))
        
        # Organize by type
        macros = [f for f in unique_files if f['file_type'] == 'macro']
        instructions = [f for f in unique_files if f['file_type'] == 'instructions']
        reports = [f for f in unique_files if f['file_type'] == 'report']
        processed = [f for f in unique_files if f['file_type'] == 'processed']
        
        return jsonify({
            'macros': macros,
            'instructions': instructions,
            'reports': reports,
            'processed': processed
        }), 200
        
    except Exception as e:
        print(f"ERROR: get_generated_files: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Landing page and app routing now handled by Apache .htaccess

# Health check
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()})

# Protected route example
@app.route('/api/profile', methods=['GET'])
@jwt_required()
def get_profile():
    current_user_email = get_jwt_identity()
    conn = get_db()
    try:
        user = conn.execute(
            'SELECT email, is_admin FROM users WHERE email = ?', (current_user_email,)
        ).fetchone()
        if user:
            return jsonify({
                'email': user['email'],
                'is_admin': bool(user['is_admin']) if user['is_admin'] is not None else False
            })
        else:
            return jsonify({'error': 'User not found'}), 404
    finally:
        conn.close()

# Endpoint to validate invitation token and get email
@app.route('/api/validate-invitation', methods=['POST'])
def validate_invitation():
    """Validate invitation token and return email (public endpoint)"""
    try:
        data = request.get_json()
        token = data.get('token')
        
        if not token:
            return jsonify({'error': 'Token is required'}), 400
        
        is_valid, email, error_msg = validate_invitation_token(token)
        if not is_valid:
            return jsonify({'error': error_msg or 'Invalid invitation token'}), 400
        
        return jsonify({
            'valid': True,
            'email': email
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Admin endpoint to create invitations
@app.route('/api/admin/create-invitation', methods=['POST'])
@jwt_required()
def create_invitation():
    """Create an invitation token for a new user (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        # Check if user is admin
        if not is_admin_user(current_user_email):
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        # Validate email format
        email = email.strip().lower()
        if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Check if user already exists
        conn = get_db()
        try:
            existing_user = conn.execute(
                'SELECT id FROM users WHERE email = ?', (email,)
            ).fetchone()
            if existing_user:
                return jsonify({'error': 'User with this email already exists'}), 409
            
            # Check if there's already a pending invitation for this email
            pending_invitation = conn.execute(
                '''SELECT id FROM invitation_tokens 
                   WHERE email = ? AND used_at IS NULL AND expires_at > ?''',
                (email, datetime.utcnow().isoformat())
            ).fetchone()
            if pending_invitation:
                return jsonify({'error': 'A pending invitation already exists for this email'}), 409
            
            # Delete any expired or used invitations for this email to avoid UNIQUE constraint issues
            conn.execute(
                '''DELETE FROM invitation_tokens 
                   WHERE email = ? AND (used_at IS NOT NULL OR expires_at <= ?)''',
                (email, datetime.utcnow().isoformat())
            )
            
            # Generate invitation token (JWT with 7-day expiration)
            # Use time.time() for consistency with download tokens
            now = int(time.time())
            expires_at = datetime.utcnow() + timedelta(days=7)  # For database storage
            payload = {
                'email': email,
                'purpose': 'invitation',
                'exp': now + (7 * 24 * 60 * 60),  # 7 days in seconds
                'iat': now
            }
            secret = app.config['JWT_SECRET_KEY']
            token = jwt_lib.encode(payload, secret, algorithm='HS256')
            
            # Store token in database
            conn.execute(
                '''INSERT INTO invitation_tokens (email, token, expires_at, created_by)
                   VALUES (?, ?, ?, ?)''',
                (email, token, expires_at.isoformat(), current_user_email)
            )
            conn.commit()
            
            # Generate invitation URL - use frontend URL, not backend
            # Frontend is at https://xlsvc.jsilverman.ca
            # Backend is at https://api.xlsvc.jsilverman.ca
            frontend_url = os.environ.get('FRONTEND_URL', 'https://xlsvc.jsilverman.ca')
            if request.host.startswith('localhost'):
                frontend_url = 'http://localhost:5173'
            invitation_url = f"{frontend_url}/app?register=1&token={token}"
            
            return jsonify({
                'success': True,
                'email': email,
                'token': token,
                'invitation_url': invitation_url,
                'expires_at': expires_at.isoformat()
            }), 201
            
        except sqlite3.IntegrityError as e:
            # This should rarely happen now since we delete old invitations
            # But handle it gracefully if it does
            error_msg = str(e)
            if 'UNIQUE constraint' in error_msg and 'email' in error_msg:
                return jsonify({'error': 'An invitation for this email already exists. Please expire the existing invitation first or wait for it to expire.'}), 409
            elif 'UNIQUE constraint' in error_msg and 'token' in error_msg:  # pragma: no cover -- token collision extremely rare
                return jsonify({'error': 'Token collision occurred. Please try again.'}), 409
            else:  # pragma: no cover -- other constraint violations
                return jsonify({'error': 'Database constraint violation. Please try again.'}), 409
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error creating invitation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Admin endpoint to list all invitations
@app.route('/api/admin/invitations', methods=['GET'])
@jwt_required()
def list_invitations():
    """List all invitations (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        # Check if user is admin
        if not is_admin_user(current_user_email):
            return jsonify({'error': 'Admin access required'}), 403
        
        conn = get_db()
        try:
            # Get all invitations
            invitations = conn.execute(
                '''SELECT id, email, token, created_at, expires_at, used_at, created_by
                   FROM invitation_tokens
                   ORDER BY created_at DESC'''
            ).fetchall()
            
            # Format invitations with status
            result = []
            now = datetime.utcnow()
            for inv in invitations:
                inv_dict = dict(inv)
                expires_at = datetime.fromisoformat(inv_dict['expires_at'])
                
                # Determine status
                if inv_dict['used_at']:
                    status = 'used'
                elif now > expires_at:
                    status = 'expired'
                else:
                    status = 'pending'
                
                inv_dict['status'] = status
                result.append(inv_dict)
            
            return jsonify({'invitations': result}), 200
            
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error listing invitations: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Admin endpoint to expire/revoke an invitation
@app.route('/api/admin/invitations/<int:invitation_id>/expire', methods=['POST'])
@jwt_required()
def expire_invitation(invitation_id):
    """Expire/revoke a pending invitation (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        # Check if user is admin
        if not is_admin_user(current_user_email):
            return jsonify({'error': 'Admin access required'}), 403
        
        conn = get_db()
        try:
            # Get invitation
            invitation = conn.execute(
                'SELECT id, email, used_at, expires_at FROM invitation_tokens WHERE id = ?',
                (invitation_id,)
            ).fetchone()
            
            if not invitation:
                return jsonify({'error': 'Invitation not found'}), 404
            
            inv_dict = dict(invitation)
            
            # Check if already used
            if inv_dict['used_at']:
                return jsonify({'error': 'Invitation has already been used'}), 400
            
            # Check if already expired
            expires_at = datetime.fromisoformat(inv_dict['expires_at'])
            if datetime.utcnow() > expires_at:
                return jsonify({'error': 'Invitation has already expired'}), 400
            
            # Mark as used (effectively revoking it)
            conn.execute(
                'UPDATE invitation_tokens SET used_at = ? WHERE id = ?',
                (datetime.utcnow().isoformat(), invitation_id)
            )
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'Invitation for {inv_dict["email"]} has been revoked'
            }), 200
            
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error expiring invitation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Admin endpoint to list all users
@app.route('/api/admin/users', methods=['GET'])
@jwt_required()
def list_users():
    """List all users with file counts (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        # Check if user is admin
        if not is_admin_user(current_user_email):
            return jsonify({'error': 'Admin access required'}), 403
        
        conn = get_db()
        try:
            # Get all users
            users = conn.execute(
                '''SELECT id, email, created_at, is_admin 
                   FROM users 
                   ORDER BY created_at DESC'''
            ).fetchall()
            
            # Get file counts for each user
            result = []
            for user in users:
                user_dict = dict(user)
                user_id = user_dict['id']
                
                # Count files for this user
                file_count = conn.execute(
                    'SELECT COUNT(*) as count FROM files WHERE user_id = ?',
                    (user_id,)
                ).fetchone()['count']
                
                user_dict['file_count'] = file_count
                user_dict['is_admin'] = bool(user_dict['is_admin']) if user_dict['is_admin'] is not None else False
                result.append(user_dict)
            
            return jsonify({'users': result}), 200
            
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error listing users: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Admin endpoint to get user details
@app.route('/api/admin/users/<int:user_id>', methods=['GET'])
@jwt_required()
def get_user_details(user_id):
    """Get detailed information about a user (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        # Check if user is admin
        if not is_admin_user(current_user_email):
            return jsonify({'error': 'Admin access required'}), 403
        
        conn = get_db()
        try:
            # Get user info
            user = conn.execute(
                '''SELECT id, email, created_at, is_admin 
                   FROM users WHERE id = ?''',
                (user_id,)
            ).fetchone()
            
            if not user:
                return jsonify({'error': 'User not found'}), 404
            
            user_dict = dict(user)
            
            # Count files
            file_count = conn.execute(
                'SELECT COUNT(*) as count FROM files WHERE user_id = ?',
                (user_id,)
            ).fetchone()['count']
            
            # Count processing jobs
            job_count = conn.execute(
                'SELECT COUNT(*) as count FROM processing_jobs WHERE user_id = ?',
                (user_id,)
            ).fetchone()['count']
            
            user_dict['file_count'] = file_count
            user_dict['job_count'] = job_count
            user_dict['is_admin'] = bool(user_dict['is_admin']) if user_dict['is_admin'] is not None else False
            
            return jsonify(user_dict), 200
            
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error getting user details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Admin endpoint to delete a user
@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
def delete_user(user_id):
    """Delete a user and all their data (admin only)"""
    try:
        current_user_email = get_jwt_identity()
        
        # Check if user is admin
        if not is_admin_user(current_user_email):
            return jsonify({'error': 'Admin access required'}), 403
        
        conn = get_db()
        try:
            # Get current user's ID
            current_user = conn.execute(
                'SELECT id FROM users WHERE email = ?',
                (current_user_email,)
            ).fetchone()
            
            if not current_user:
                return jsonify({'error': 'Current user not found'}), 404
            
            current_user_id = current_user['id']
            
            # Cannot delete yourself
            if user_id == current_user_id:
                return jsonify({'error': 'Cannot delete your own account'}), 400
            
            # Get user to delete
            user_to_delete = conn.execute(
                'SELECT id, email, is_admin FROM users WHERE id = ?',
                (user_id,)
            ).fetchone()
            
            if not user_to_delete:
                return jsonify({'error': 'User not found'}), 404
            
            # Check if this is the last admin
            admin_count = conn.execute(
                'SELECT COUNT(*) as count FROM users WHERE is_admin = 1'
            ).fetchone()['count']
            
            if user_to_delete['is_admin'] and admin_count <= 1:
                return jsonify({'error': 'Cannot delete the last admin user'}), 400
            
            # Get all files for this user
            user_files = conn.execute(
                '''SELECT id, stored_filename, file_type 
                   FROM files WHERE user_id = ?''',
                (user_id,)
            ).fetchall()
            
            # Delete files from disk
            for file_record in user_files:
                file_dict = dict(file_record)
                file_path = get_file_path(
                    file_dict.get('file_type'),
                    file_dict['stored_filename']
                )
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                        print(f"Deleted file from disk: {file_path}")
                    except Exception as e:
                        print(f"Warning: Could not delete file {file_path}: {e}")
            
            # Delete all user's files from database
            conn.execute('DELETE FROM files WHERE user_id = ?', (user_id,))
            
            # Delete all user's processing jobs
            conn.execute('DELETE FROM processing_jobs WHERE user_id = ?', (user_id,))
            
            # Delete all invitation tokens created by this user
            conn.execute('DELETE FROM invitation_tokens WHERE created_by = ?', (user_to_delete['email'],))
            
            # Delete the user
            conn.execute('DELETE FROM users WHERE id = ?', (user_id,))
            
            conn.commit()
            
            return jsonify({
                'success': True,
                'message': f'User {user_to_delete["email"]} and all associated data have been deleted'
            }), 200
            
        finally:
            conn.close()
            
    except Exception as e:
        print(f"Error deleting user: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':  # pragma: no cover
    app.run(debug=True)