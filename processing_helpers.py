import io
import re
import zipfile

from openpyxl import load_workbook

# LibreOffice/Calc can write a <cellStyles> table whose entries reference xfId
# values beyond the <cellStyleXfs> array. openpyxl's named-style expansion then
# raises IndexError when loading. This matches that block so we can drop it.
_CELLSTYLES_RE = re.compile(rb"<cellStyles\b.*?</cellStyles>", re.S)


def load_workbook_resilient(path, **kwargs):
    """
    Load an .xlsx workbook, tolerating the malformed <cellStyles> table that
    LibreOffice emits when saving (the named styles reference xfId values past
    the end of cellStyleXfs, which makes openpyxl raise IndexError/KeyError).

    On that failure we strip the <cellStyles> block — only named-style *names*,
    not the cell formatting xfs — from an in-memory copy and load that. Cell
    values, sheets, rows and columns are unaffected, so this is safe for data
    comparison and re-processing. Normal workbooks load on the first attempt.
    """
    try:
        return load_workbook(path, **kwargs)
    except (IndexError, KeyError):
        sanitized = _strip_named_styles(path)
        if sanitized is None:
            raise
        return load_workbook(sanitized, **kwargs)


def _strip_named_styles(path):
    """Return a BytesIO copy of the xlsx with <cellStyles> removed, or None."""
    with zipfile.ZipFile(path) as src:
        if "xl/styles.xml" not in src.namelist():
            return None
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = _CELLSTYLES_RE.sub(b"", data)
                dst.writestr(item, data)
    buf.seek(0)
    return buf


def evaluate_cell_value(cell):
    """
    Get the actual value of a cell, evaluating formulas if necessary.
    Returns the calculated value for formulas, or the raw value otherwise.
    """
    if hasattr(cell, 'value'):
        value = cell.value
        
        if isinstance(value, str) and value.startswith('='):
            return None  # Return None to indicate we need data_only mode
        
        return value
    return None


def is_empty_or_zero(val):
    """
    Check if a value is empty, None, zero, or blank.
    Handles various data types appropriately.
    """
    if val is None:
        return True
    if val == 0:
        return True
    if val == "":
        return True
    if isinstance(val, str) and val.strip() == '':
        return True
    if isinstance(val, str) and val.strip() == '0':
        return True
    return False


def column_to_index(col):
    """Convert column letter (A, F, Z) or number (1, 6, 26) to 1-based column index"""
    col = str(col).strip().upper()
    
    # If it's already a number, return it as int
    if col.isdigit():
        return int(col)
    
    # Convert letter(s) to number (A=1, B=2, ... Z=26, AA=27, etc.)
    index = 0
    for char in col:
        if char.isalpha():
            index = index * 26 + (ord(char) - ord('A') + 1)
    return index
