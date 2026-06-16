#!/usr/bin/env python3
"""
Parity toolkit for comparing manual (macro/openpyxl) vs automated (UNO) output.

This does NOT run LibreOffice/UNO. It (1) generates a deterministic test
workbook with the tricky cases both paths must agree on, printing the rows that
SHOULD be deleted per the shared is_empty_or_zero predicate, and (2) diffs two
processed workbooks (or two deletion reports) cell-by-cell so a human can run
both pipelines on the fixture and confirm identical results.

Usage:
    python tools/parity_check.py generate [out.xlsx]
    python tools/parity_check.py compare <a.xlsx> <b.xlsx>

The default filter used for the fixture is F,G,H,I empty/zero (the "Silver"
template), and rows with an empty Column A are skipped (parity rule).
"""
import os
import sys
from openpyxl import Workbook, load_workbook

# Allow running as `python tools/parity_check.py` from the repo root.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from processing_helpers import is_empty_or_zero

FILTER_COLUMNS = ["F", "G", "H", "I"]


def _expected_delete(row_cells):
    """row_cells: dict col_letter -> value. Mirror both paths' decision."""
    a = row_cells.get("A")
    if a is None or str(a).strip() == "":
        return False  # Column A empty -> skipped
    return all(is_empty_or_zero(row_cells.get(c)) for c in FILTER_COLUMNS)


def generate(path="parity_test.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parity"
    ws.append(["A", "B", "C", "D", "E", "F", "G", "H", "I"])

    # (label in A, then F..I values) — clear cases both paths must agree on.
    rows = [
        ("all_zero", 0, 0, 0, 0),            # delete
        ("all_empty", None, None, None, None),  # delete
        ("partial", 1, 0, 0, 0),             # keep (F not empty/zero)
        ("", 0, 0, 0, 0),                    # skip (Column A empty) -> keep
        ("whitespace", "  ", "\t", " ", ""),  # delete
        ("str_zero", "0", "0", "0", "0"),    # delete
        ("text", "x", "y", "z", "w"),        # keep
        ("mixed", 0, "0", "", " "),          # delete
    ]
    for label, f, g, h, i in rows:
        ws.append([label, "", "", "", "", f, g, h, i])

    # Bulk rows beyond the old 1000-row cap, alternating delete/keep, to verify
    # the automated path now scans the full used range.
    for n in range(2000):
        if n % 2 == 0:
            ws.append([f"bulk_del_{n}", "", "", "", "", 0, 0, 0, 0])
        else:
            ws.append([f"bulk_keep_{n}", "", "", "", "", 1, 0, 0, 0])

    wb.save(path)

    # Compute expected deletions (1-based row numbers, header is row 1).
    expected = []
    for idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = {chr(65 + c): r[c] for c in range(len(r))}
        if _expected_delete(cells):
            expected.append(idx)
    print(f"Wrote {path} ({ws.max_row} rows incl header).")
    print(f"Expected rows to DELETE (filter {FILTER_COLUMNS}, skip empty A): {len(expected)}")
    print(f"  first 12: {expected[:12]}")
    print("Both the manual macro output and the automated output should delete "
          "exactly these rows and produce matching deletion reports.")
    return expected


def compare(path_a, path_b):
    wb_a = load_workbook(path_a, data_only=True)
    wb_b = load_workbook(path_b, data_only=True)
    diffs = []
    if wb_a.sheetnames != wb_b.sheetnames:
        diffs.append(f"sheet names differ: {wb_a.sheetnames} vs {wb_b.sheetnames}")
    for name in set(wb_a.sheetnames) & set(wb_b.sheetnames):
        sa, sb = wb_a[name], wb_b[name]
        if (sa.max_row, sa.max_column) != (sb.max_row, sb.max_column):
            diffs.append(
                f"[{name}] dimensions differ: "
                f"{sa.max_row}x{sa.max_column} vs {sb.max_row}x{sb.max_column}"
            )
        rows = min(sa.max_row, sb.max_row)
        cols = min(sa.max_column, sb.max_column)
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                va = sa.cell(row=r, column=c).value
                vb = sb.cell(row=r, column=c).value
                if va != vb:
                    diffs.append(f"[{name}] R{r}C{c}: {va!r} != {vb!r}")
                    if len(diffs) > 50:
                        diffs.append("... (more than 50 differences; stopping)")
                        _report(diffs)
                        return 1
    _report(diffs)
    return 1 if diffs else 0


def _report(diffs):
    if not diffs:
        print("✅ IDENTICAL — the two workbooks match cell-for-cell.")
    else:
        print(f"❌ {len(diffs)} difference(s):")
        for d in diffs:
            print(f"  - {d}")


def main(argv):
    if len(argv) >= 2 and argv[1] == "generate":
        generate(argv[2] if len(argv) > 2 else "parity_test.xlsx")
        return 0
    if len(argv) == 4 and argv[1] == "compare":
        return compare(argv[2], argv[3])
    print(__doc__)
    return 2


if __name__ == "__main__":
    sys.exit(main(sys.argv))
