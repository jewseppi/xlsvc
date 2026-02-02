"""
Pytest configuration and fixtures for xlsvc tests
"""
import pytest
import os
import tempfile
import shutil
import sqlite3
from datetime import datetime
from werkzeug.security import generate_password_hash
from flask import Flask
import sys

# Add parent directory to path to import main
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from main import app, init_db, get_db, ensure_directories


@pytest.fixture(scope='session')
def test_db_path():
    """Create a temporary database file for testing"""
    fd, path = tempfile.mkstemp(suffix='.db')
    os.close(fd)
    yield path
    if os.path.exists(path):
        os.remove(path)


@pytest.fixture(scope='session')
def test_directories():
    """Create temporary directories for test files"""
    base_dir = tempfile.mkdtemp()
    dirs = {
        'uploads': os.path.join(base_dir, 'uploads'),
        'processed': os.path.join(base_dir, 'processed'),
        'macros': os.path.join(base_dir, 'macros'),
        'reports': os.path.join(base_dir, 'reports')
    }
    
    for dir_path in dirs.values():
        os.makedirs(dir_path, exist_ok=True)
    
    yield dirs
    
    # Cleanup
    shutil.rmtree(base_dir, ignore_errors=True)


@pytest.fixture(scope='function')
def test_app(test_db_path, test_directories, monkeypatch):
    """Create a Flask app instance for testing"""
    # Set environment variables
    test_secret = 'test-secret-key-for-testing-only-12345678901234567890'
    monkeypatch.setenv('SECRET_KEY', test_secret)
    
    # Configure app for testing
    app.config['TESTING'] = True
    app.config['JWT_SECRET_KEY'] = test_secret
    app.config['UPLOAD_FOLDER'] = test_directories['uploads']
    app.config['PROCESSED_FOLDER'] = test_directories['processed']
    app.config['MACROS_FOLDER'] = test_directories['macros']
    app.config['REPORTS_FOLDER'] = test_directories['reports']
    app.config['DATABASE'] = test_db_path
    
    # Monkey patch get_db to use test database
    original_get_db = None
    if hasattr(app, 'get_db'):
        original_get_db = app.get_db
    
    def get_test_db():
        conn = sqlite3.connect(test_db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    # Replace get_db in main module
    import main
    monkeypatch.setattr(main, 'get_db', get_test_db)
    
    # Initialize test database
    conn = sqlite3.connect(test_db_path)
    cursor = conn.cursor()
    
    # Create tables (simplified version of init_db)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_admin BOOLEAN DEFAULT 0
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            original_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            file_size INTEGER,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            processed BOOLEAN DEFAULT FALSE,
            file_hash TEXT,
            file_type TEXT DEFAULT "original",
            parent_file_id INTEGER,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS processing_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id TEXT UNIQUE NOT NULL,
            user_id INTEGER,
            original_file_id INTEGER,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            result_file_id INTEGER,
            report_file_id INTEGER,
            error_message TEXT,
            deleted_rows INTEGER DEFAULT 0,
            filter_rules_json TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id),
            FOREIGN KEY (original_file_id) REFERENCES files (id),
            FOREIGN KEY (result_file_id) REFERENCES files (id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subscribers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL,
            notified_at TEXT DEFAULT NULL
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS invitation_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            token TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP NOT NULL,
            used_at TIMESTAMP DEFAULT NULL,
            created_by TEXT
        )
    ''')
    
    conn.commit()
    conn.close()
    
    ensure_directories()
    
    yield app
    
    # Cleanup test database
    if os.path.exists(test_db_path):
        os.remove(test_db_path)


@pytest.fixture(scope='function')
def client(test_app):
    """Create a test client"""
    return test_app.test_client()


@pytest.fixture(scope='function')
def db_connection(test_db_path):
    """Get a database connection for testing"""
    conn = sqlite3.connect(test_db_path)
    conn.row_factory = sqlite3.Row
    yield conn
    conn.close()


@pytest.fixture(scope='function')
def test_user(test_app):
    """Create a test user using the same database connection that the app uses"""
    from main import get_db
    conn = get_db()
    
    email = 'test@example.com'
    password_hash = generate_password_hash('testpassword123')
    
    try:
        # Delete existing user if any
        conn.execute('DELETE FROM users WHERE email = ?', (email,))
        
        # Insert new user
        cursor = conn.execute(
            'INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, ?)',
            (email, password_hash, 0)
        )
        conn.commit()
        
        # Get the user ID
        user = conn.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
        user_id = user['id'] if user else cursor.lastrowid
        
        yield {
            'id': user_id,
            'email': email,
            'password': 'testpassword123'
        }
    finally:
        # Cleanup
        conn.execute('DELETE FROM users WHERE email = ?', (email,))
        conn.commit()
        conn.close()


@pytest.fixture(scope='function')
def test_admin_user(test_app):
    """Create a test admin user using the same database connection that the app uses"""
    from main import get_db
    conn = get_db()
    
    email = 'admin@example.com'
    password_hash = generate_password_hash('adminpassword123')
    
    try:
        # Delete existing user if any
        conn.execute('DELETE FROM users WHERE email = ?', (email,))
        
        # Insert new admin user
        cursor = conn.execute(
            'INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, ?)',
            (email, password_hash, 1)
        )
        conn.commit()
        
        # Get the user ID
        user = conn.execute('SELECT id FROM users WHERE email = ?', (email,)).fetchone()
        user_id = user['id'] if user else cursor.lastrowid
        
        yield {
            'id': user_id,
            'email': email,
            'password': 'adminpassword123',
            'is_admin': True
        }
    finally:
        # Cleanup
        conn.execute('DELETE FROM users WHERE email = ?', (email,))
        conn.commit()
        conn.close()


@pytest.fixture(scope='function', params=[True, False], ids=['auth_ok', 'auth_none'])
def auth_token(request, client, test_user):
    """Get an authentication token for a test user. When request.param is False, returns None to cover skip branches."""
    if not request.param:
        return None
    response = client.post('/api/login', json={
        'email': test_user['email'],
        'password': test_user['password']
    })
    if response.status_code == 200:
        data = response.get_json()
        token = data.get('access_token')
        if token:
            return token
    if response.status_code != 200:
        print(f"DEBUG: Login failed with status {response.status_code}: {response.get_json()}")
    return None


@pytest.fixture(scope='function')
def sample_excel_file(test_directories):
    """Create a sample Excel file for testing"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add header row
    ws['A1'] = 'Name'
    ws['B1'] = 'Value1'
    ws['C1'] = 'Value2'
    ws['F1'] = 'ColF'
    ws['G1'] = 'ColG'
    ws['H1'] = 'ColH'
    ws['I1'] = 'ColI'
    
    # Add some test data
    ws['A2'] = 'Row1'
    ws['F2'] = 0
    ws['G2'] = 0
    ws['H2'] = 0
    ws['I2'] = 0
    
    ws['A3'] = 'Row2'
    ws['F3'] = 1
    ws['G3'] = 0
    ws['H3'] = 0
    ws['I3'] = 0
    
    ws['A4'] = 'Row3'
    ws['F4'] = 0
    ws['G4'] = 0
    ws['H4'] = 0
    ws['I4'] = 0
    
    # Save to test uploads directory
    file_path = os.path.join(test_directories['uploads'], 'test_file.xlsx')
    wb.save(file_path)
    
    yield file_path
    
    # Cleanup
    if os.path.exists(file_path):
        os.remove(file_path)


@pytest.fixture(scope='function')
def comprehensive_test_excel(test_directories):
    """Create a comprehensive test Excel file with various scenarios"""
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Basic test cases
    ws1 = wb.create_sheet("TestSheet1")
    
    # Headers
    headers = ['Name', 'ColA', 'ColB', 'ColC', 'ColD', 'ColE', 'ColF', 'ColG', 'ColH', 'ColI']
    ws1.append(headers)
    
    # Row 2: All zeros (should be deleted with default filters F,G,H,I = 0)
    ws1.append(['Row2_AllZeros', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Row 3: All blanks (should be deleted)
    ws1.append(['Row3_AllBlanks', '', '', '', '', '', '', '', '', ''])
    
    # Row 4: Mixed zeros and blanks in F,G,H,I (should be deleted)
    ws1.append(['Row4_MixedZeros', 1, 2, 3, 4, 5, 0, '', 0, ''])
    
    # Row 5: Has values in F,G,H,I (should NOT be deleted)
    ws1.append(['Row5_HasValues', 1, 2, 3, 4, 5, 1, 2, 3, 4])
    
    # Row 6: All zeros again (should be deleted)
    ws1.append(['Row6_AllZeros', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Row 7: String "0" in F,G,H,I (should be deleted - treated as zero)
    ws1.append(['Row7_StringZeros', 1, 2, 3, 4, 5, '0', '0', '0', '0'])
    
    # Row 8: Whitespace in F,G,H,I (should be deleted)
    ws1.append(['Row8_Whitespace', 1, 2, 3, 4, 5, ' ', '  ', '\t', '\n'])
    
    # Row 9: Has value in F but zeros in G,H,I (should NOT be deleted - all must match)
    ws1.append(['Row9_PartialMatch', 1, 2, 3, 4, 5, 1, 0, 0, 0])
    
    # Row 10: All zeros (should be deleted)
    ws1.append(['Row10_AllZeros', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Sheet 2: Edge cases
    ws2 = wb.create_sheet("TestSheet2")
    ws2.append(headers)
    
    # Row 2: Float zeros
    ws2.append(['Row2_FloatZeros', 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0])
    
    # Row 3: Has actual data
    ws2.append(['Row3_Data', 10, 20, 30, 40, 50, 60, 70, 80, 90])
    
    # Sheet 3: Empty sheet (just header)
    ws3 = wb.create_sheet("EmptySheet")
    ws3.append(headers)
    
    # Save to test uploads directory
    file_path = os.path.join(test_directories['uploads'], 'comprehensive_test.xlsx')
    wb.save(file_path)
    
    yield file_path
    
    # Cleanup
    if os.path.exists(file_path):
        os.remove(file_path)
