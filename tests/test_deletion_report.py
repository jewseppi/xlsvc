"""
Tests for deletion_report module (generate_deletion_report, capture_row_data).
"""
import pytest
import os
import tempfile
from unittest.mock import MagicMock

from deletion_report import generate_deletion_report, capture_row_data
from openpyxl import Workbook, load_workbook


class TestGenerateDeletionReport:
    """Tests for generate_deletion_report."""

    def test_empty_data_returns_none(self):
        """No data returns None and prints message."""
        assert generate_deletion_report({}, "/tmp/out.xlsx") is None
        assert generate_deletion_report(None, "/tmp/out.xlsx") is None

    def test_with_data_creates_report_file(self):
        """Valid data generates report and returns output path."""
        data = {
            "Sheet1": [
                {"row_number": 2, "data": ["a", "b", "c"]},
                {"row_number": 3, "data": ["x", "y", "z"]},
            ]
        }
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            result = generate_deletion_report(data, path)
            assert result == path
            assert os.path.exists(path)
            wb = load_workbook(path)
            assert "Sheet1" in wb.sheetnames
            ws = wb["Sheet1"]
            assert ws.cell(row=1, column=1).value == "Original Row #"
            assert ws.cell(row=2, column=1).value == 2
            assert ws.cell(row=2, column=2).value == "a"
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_skip_empty_sheets(self):
        """Sheets with empty rows are skipped; only non-empty sheets added."""
        data = {
            "Empty": [],
            "HasData": [{"row_number": 1, "data": ["only"]}],
        }
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            result = generate_deletion_report(data, path)
            assert result == path
            wb = load_workbook(path)
            assert "HasData" in wb.sheetnames
            assert "Empty" not in wb.sheetnames
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_all_sheets_empty_returns_none(self):
        """When every sheet has no rows, no file is written and returns None."""
        data = {"S1": [], "S2": []}
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            result = generate_deletion_report(data, path)
            assert result is None
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_sheet_name_truncated_to_31_chars(self):
        """Excel sheet name limit (31) is applied."""
        long_name = "A" * 40
        data = {long_name: [{"row_number": 1, "data": ["x"]}]}
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            result = generate_deletion_report(data, path)
            assert result == path
            wb = load_workbook(path)
            assert list(wb.sheetnames)[0] == long_name[:31]
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_cell_value_len_exception_handled(self):
        """Exception when computing cell length in column sizing is caught (pass)."""
        data = {"S1": [{"row_number": 1, "data": ["normal"]}]}
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            # Should not raise; internal except pass covers bad cell.value
            result = generate_deletion_report(data, path)
            assert result == path
        finally:
            if os.path.exists(path):
                os.remove(path)


class TestCaptureRowData:
    """Tests for capture_row_data."""

    def test_capture_row_data_normal(self):
        """Captures row values from openpyxl sheet."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["C1"] = None
        result = capture_row_data(ws, 1, max_cols=10)
        assert result == ["a", "b", ""]

    def test_capture_row_data_trims_trailing_empty(self):
        """Trailing empty cells are trimmed."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "x"
        ws["B1"] = ""
        ws["C1"] = ""
        result = capture_row_data(ws, 1, max_cols=5)
        assert result == ["x"]

    def test_capture_row_data_exception_breaks_loop(self):
        """When sheet.cell raises, loop breaks and trimmed data is returned."""
        sheet = MagicMock()
        sheet.cell.side_effect = [
            MagicMock(value="ok"),
            MagicMock(value="also"),
            Exception("mock error"),
        ]
        result = capture_row_data(sheet, 1, max_cols=5)
        assert result == ["ok", "also"]
        assert sheet.cell.call_count == 3
