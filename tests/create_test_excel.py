"""
Script to create a comprehensive test Excel file with various scenarios
This file will be used for testing row deletion logic
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os

def create_test_excel_file(output_path):
    """Create a test Excel file with various row deletion scenarios"""
    wb = Workbook()
    
    # Remove default sheet and create our test sheets
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: Basic test cases
    ws1 = wb.create_sheet("TestSheet1")
    
    # Headers
    headers = ['Name', 'ColA', 'ColB', 'ColC', 'ColD', 'ColE', 'ColF', 'ColG', 'ColH', 'ColI']
    ws1.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Row 2: All zeros (should be deleted with default filters)
    ws1.append(['Row2_AllZeros', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Row 3: All blanks (should be deleted with default filters)
    ws1.append(['Row3_AllBlanks', '', '', '', '', '', '', '', '', ''])
    
    # Row 4: Mixed zeros and blanks in F,G,H,I (should be deleted)
    ws1.append(['Row4_MixedZeros', 1, 2, 3, 4, 5, 0, '', 0, ''])
    
    # Row 5: Has values in F,G,H,I (should NOT be deleted)
    ws1.append(['Row5_HasValues', 1, 2, 3, 4, 5, 1, 2, 3, 4])
    
    # Row 6: All zeros again (should be deleted)
    ws1.append(['Row6_AllZeros', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Row 7: String "0" in F,G,H,I (should be deleted - treated as zero)
    ws1.append(['Row7_StringZeros', 1, 2, 3, 4, 5, '0', '0', '0', '0'])
    
    # Row 8: Whitespace in F,G,H,I (should be deleted)
    ws1.append(['Row8_Whitespace', 1, 2, 3, 4, 5, ' ', '  ', '\t', '\n'])
    
    # Row 9: Has value in F but zeros in G,H,I (should NOT be deleted - all must match)
    ws1.append(['Row9_PartialMatch', 1, 2, 3, 4, 5, 1, 0, 0, 0])
    
    # Row 10: All zeros (should be deleted)
    ws1.append(['Row10_AllZeros', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Sheet 2: Edge cases
    ws2 = wb.create_sheet("TestSheet2")
    ws2.append(headers)
    
    # Style headers
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Row 2: Very large row number test
    ws2.append(['Row2_Large', 0, 0, 0, 0, 0, 0, 0, 0, 0])
    
    # Row 3: Float zeros
    ws2.append(['Row3_FloatZeros', 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0])
    
    # Row 4: Mixed float and int zeros
    ws2.append(['Row4_MixedZeros', 0, 0.0, 0, 0.0, 0, 0, 0, 0, 0])
    
    # Row 5: Has actual data
    ws2.append(['Row5_Data', 10, 20, 30, 40, 50, 60, 70, 80, 90])
    
    # Sheet 3: Empty sheet (edge case)
    ws3 = wb.create_sheet("EmptySheet")
    ws3.append(headers)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 4: Single row (just header)
    ws4 = wb.create_sheet("HeaderOnly")
    ws4.append(headers)
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Save the workbook
    wb.save(output_path)
    print(f"Test Excel file created: {output_path}")
    print(f"Sheet1: 10 rows (header + 9 data rows)")
    print(f"Sheet2: 5 rows (header + 4 data rows)")
    print(f"Sheet3: 1 row (header only)")
    print(f"Sheet4: 1 row (header only)")

if __name__ == '__main__':
    # Create test file in tests directory
    test_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(test_dir, 'test_data', 'test_file.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    create_test_excel_file(output_path)
