"""
Unit tests for processing helper functions
"""
import pytest
from main import is_empty_or_zero, column_to_index


class TestIsEmptyOrZero:
    """Tests for is_empty_or_zero function"""
    
    def test_none_value(self):
        """Test that None returns True"""
        assert is_empty_or_zero(None) is True
    
    def test_zero_integer(self):
        """Test that integer 0 returns True"""
        assert is_empty_or_zero(0) is True
    
    def test_zero_float(self):
        """Test that float 0.0 returns True"""
        assert is_empty_or_zero(0.0) is True
    
    def test_empty_string(self):
        """Test that empty string returns True"""
        assert is_empty_or_zero("") is True
    
    def test_whitespace_string(self):
        """Test that whitespace-only string returns True"""
        assert is_empty_or_zero("   ") is True
        assert is_empty_or_zero("\t\n") is True
    
    def test_string_zero(self):
        """Test that string "0" returns True"""
        assert is_empty_or_zero("0") is True
        assert is_empty_or_zero(" 0 ") is True
    
    def test_non_zero_integer(self):
        """Test that non-zero integer returns False"""
        assert is_empty_or_zero(1) is False
        assert is_empty_or_zero(-1) is False
        assert is_empty_or_zero(100) is False
    
    def test_non_zero_float(self):
        """Test that non-zero float returns False"""
        assert is_empty_or_zero(0.1) is False
        assert is_empty_or_zero(-0.1) is False
        assert is_empty_or_zero(1.5) is False
    
    def test_non_empty_string(self):
        """Test that non-empty string returns False"""
        assert is_empty_or_zero("hello") is False
        assert is_empty_or_zero("1") is False
        assert is_empty_or_zero("abc") is False


class TestColumnToIndex:
    """Tests for column_to_index function"""
    
    def test_single_letter_columns(self):
        """Test single letter column conversion"""
        assert column_to_index('A') == 1
        assert column_to_index('B') == 2
        assert column_to_index('Z') == 26
        assert column_to_index('a') == 1  # lowercase
        assert column_to_index('F') == 6
    
    def test_double_letter_columns(self):
        """Test double letter column conversion"""
        assert column_to_index('AA') == 27
        assert column_to_index('AB') == 28
        assert column_to_index('AZ') == 52
        assert column_to_index('BA') == 53
    
    def test_numeric_input(self):
        """Test numeric column input"""
        assert column_to_index('1') == 1
        assert column_to_index('5') == 5
        assert column_to_index('26') == 26
        assert column_to_index(1) == 1
        assert column_to_index(5) == 5
    
    def test_whitespace_handling(self):
        """Test that whitespace is handled correctly"""
        assert column_to_index(' A ') == 1
        assert column_to_index('  F  ') == 6
        assert column_to_index(' AA ') == 27
    
    def test_edge_cases(self):
        """Test edge cases"""
        assert column_to_index('IV') == 256  # Excel max column in older versions
        assert column_to_index('XFD') == 16384  # Excel max column in newer versions


class TestLoadWorkbookResilient:
    """Tests for load_workbook_resilient (tolerates LibreOffice's bad cellStyles)."""

    @staticmethod
    def _normal_xlsx_bytes():
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "hello"
        ws["B1"] = 42
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    @classmethod
    def _bad_cellstyles_xlsx(cls):
        """A workbook whose <cellStyles> references a missing xfId -> IndexError."""
        import io, re, zipfile
        src = zipfile.ZipFile(io.BytesIO(cls._normal_xlsx_bytes()))
        out = io.BytesIO()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if item.filename == "xl/styles.xml":
                    s = data.decode("utf-8")
                    bad = '<cellStyle name="Bad" xfId="999" builtinId="0"/>'
                    s = re.sub(r'<cellStyles count="\d+">', '<cellStyles count="2">' + bad, s, count=1)
                    data = s.encode("utf-8")
                dst.writestr(item, data)
        out.seek(0)
        return out.read()

    def test_normal_workbook_loads(self, tmp_path):
        from processing_helpers import load_workbook_resilient
        p = tmp_path / "ok.xlsx"
        p.write_bytes(self._normal_xlsx_bytes())
        wb = load_workbook_resilient(str(p), data_only=True)
        assert wb["Data"]["A1"].value == "hello"
        wb.close()

    def test_bad_cellstyles_recovered(self, tmp_path):
        import warnings
        from openpyxl import load_workbook
        from processing_helpers import load_workbook_resilient
        p = tmp_path / "bad.xlsx"
        p.write_bytes(self._bad_cellstyles_xlsx())
        # The unguarded loader raises on this file...
        with pytest.raises((IndexError, KeyError)):
            load_workbook(str(p))
        # ...but the resilient loader recovers and the data is intact.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook_resilient(str(p), data_only=True)
        assert wb["Data"]["A1"].value == "hello"
        assert wb["Data"]["B1"].value == 42
        wb.close()

    def test_reraises_when_no_styles_to_strip(self, tmp_path, monkeypatch):
        """If loading fails but there's no styles.xml to fix, the error propagates."""
        import io, zipfile
        import processing_helpers
        # A valid zip with no xl/styles.xml entry.
        p = tmp_path / "nostyles.xlsx"
        with zipfile.ZipFile(p, "w") as z:
            z.writestr("dummy.txt", "x")
        monkeypatch.setattr(
            processing_helpers, "load_workbook",
            lambda *a, **k: (_ for _ in ()).throw(IndexError("boom")),
        )
        with pytest.raises(IndexError):
            processing_helpers.load_workbook_resilient(str(p))
