"""
Tests for numbers_export.to_numbers_compatible — producing an Apple Numbers
readable copy of a LibreOffice-style workbook while preserving images.
"""
import io
import re
import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from numbers_export import to_numbers_compatible

# 1x1 transparent PNG
_PNG = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
    "890000000a49444154789c6360000002000154a24f5f0000000049454e44ae426082"
)

_DRAWING_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
    ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
    '<xdr:oneCellAnchor><xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>'
    '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
    '<xdr:ext cx="500000" cy="500000"/>'
    '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="img"/><xdr:cNvPicPr/></xdr:nvPicPr>'
    '<xdr:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:embed="rId1"/></xdr:blipFill>'
    '<xdr:spPr/></xdr:pic><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>'
)


def _libreoffice_like_source(tmp_path, with_drawing=True, extra_media_ext=None):
    """Build an .xlsx that looks like LibreOffice output: openpyxl data plus a
    grafted-in drawing on the first sheet (so it exercises the image branch)."""
    wb = Workbook()
    a = wb.active
    a.title = "Alpha"
    a.append(["SKU", "Qty"])
    a.append(["X-1", 5])
    b = wb.create_sheet("Beta")
    b.append(["only", "data"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if not with_drawing:
        path = tmp_path / "src.xlsx"
        path.write_bytes(buf.read())
        return str(path)

    src = zipfile.ZipFile(buf)
    # find Alpha's worksheet part
    import xml.etree.ElementTree as ET
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
          "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    wbxml = ET.fromstring(src.read("xl/workbook.xml"))
    rels = ET.fromstring(src.read("xl/_rels/workbook.xml.rels"))
    rid2t = {r.get("Id"): r.get("Target") for r in rels}
    alpha_part = None
    for sh in wbxml.find("m:sheets", ns):
        if sh.get("name") == "Alpha":
            alpha_part = rid2t[sh.get("{%s}id" % ns["r"])].split("/")[-1]

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w") as z:
        for it in src.infolist():
            z.writestr(it, src.read(it.filename))
        z.writestr("xl/media/image1.png", _PNG)
        if extra_media_ext:
            z.writestr(f"xl/media/image2.{extra_media_ext}", b"rawbytes")
        z.writestr("xl/drawings/drawing1.xml", _DRAWING_XML)
        z.writestr(
            "xl/drawings/_rels/drawing1.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.png"/></Relationships>',
        )
        z.writestr(
            f"xl/worksheets/_rels/{alpha_part}.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>',
        )
    path = tmp_path / "src.xlsx"
    path.write_bytes(out.getvalue())
    return str(path), alpha_part


class TestToNumbersCompatible:
    def test_grafts_images_and_data(self, tmp_path):
        src, alpha_part = _libreoffice_like_source(tmp_path, extra_media_ext="xyz")
        dst = str(tmp_path / "out.xlsx")
        assert to_numbers_compatible(src, dst) == dst

        with zipfile.ZipFile(dst) as z:
            names = z.namelist()
            # images copied verbatim
            assert "xl/media/image1.png" in names
            assert "xl/media/image2.xyz" in names
            assert "xl/drawings/drawing1.xml" in names
            # worksheet wired to the drawing
            rels = z.read(f"xl/worksheets/_rels/{alpha_part}.rels").decode()
            assert "drawings/drawing1.xml" in rels
            sheet = z.read(f"xl/worksheets/{alpha_part}").decode()
            assert "<drawing" in sheet
            # content types augmented (known + fallback ext + drawing override)
            ct = z.read("[Content_Types].xml").decode()
            assert 'Extension="png" ContentType="image/png"' in ct
            assert 'Extension="xyz" ContentType="application/octet-stream"' in ct
            assert "drawing+xml" in ct

        # the result is a valid workbook openpyxl (and Numbers) can read
        wb = load_workbook(dst)
        assert wb.sheetnames == ["Alpha", "Beta"]
        assert wb["Alpha"]["A2"].value == "X-1"
        wb.close()

    def test_no_drawings_plain_rebuild(self, tmp_path):
        src = _libreoffice_like_source(tmp_path, with_drawing=False)
        dst = str(tmp_path / "out.xlsx")
        to_numbers_compatible(src, dst)
        with zipfile.ZipFile(dst) as z:
            assert not any(n.startswith("xl/media/") for n in z.namelist())
            assert not any(n.startswith("xl/worksheets/_rels/") for n in z.namelist())
        wb = load_workbook(dst)
        assert wb.sheetnames == ["Alpha", "Beta"]
        wb.close()
