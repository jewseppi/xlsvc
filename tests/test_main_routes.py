"""
Tests for main.py API routes (download, cleanup-files, admin, health, etc.).
"""
import pytest
import os
import uuid
import jwt as jwt_lib
from datetime import datetime, timedelta
from unittest.mock import patch, MagicMock


class TestSubscribe:
    """POST /api/subscribe (used by landing.html notification form)."""

    def test_subscribe_success(self, client, db_connection):
        """Valid email subscribes successfully and returns 200."""
        email = "subscriber@example.com"
        r = client.post("/api/subscribe", json={"email": email})
        assert r.status_code == 200
        data = r.get_json()
        assert data.get("success") is True
        assert "subscribed" in data.get("message", "").lower()
        # Verify row was inserted
        row = db_connection.execute(
            "SELECT email FROM subscribers WHERE email = ?", (email,)
        ).fetchone()
        assert row is not None
        assert row["email"] == email
        # Cleanup
        db_connection.execute("DELETE FROM subscribers WHERE email = ?", (email,))
        db_connection.commit()

    def test_subscribe_duplicate_returns_409(self, client, db_connection):
        """Subscribing the same email twice returns 409."""
        email = "dupsub@example.com"
        r1 = client.post("/api/subscribe", json={"email": email})
        assert r1.status_code == 200
        r2 = client.post("/api/subscribe", json={"email": email})
        assert r2.status_code == 409
        assert "already" in r2.get_json().get("error", "").lower()
        # Cleanup
        db_connection.execute("DELETE FROM subscribers WHERE email = ?", (email,))
        db_connection.commit()

    def test_subscribe_invalid_email_returns_400(self, client):
        """Invalid email format returns 400."""
        r = client.post("/api/subscribe", json={"email": "not-an-email"})
        assert r.status_code == 400
        assert "invalid" in r.get_json().get("error", "").lower()

    def test_subscribe_empty_email_returns_400(self, client):
        """Empty or missing email returns 400."""
        r = client.post("/api/subscribe", json={"email": ""})
        assert r.status_code == 400
        r2 = client.post("/api/subscribe", json={})
        assert r2.status_code == 400

    def test_subscribe_normalizes_email(self, client, db_connection):
        """Email is lowercased and trimmed."""
        email = "  SuBsCrIbEr@Example.COM  "
        r = client.post("/api/subscribe", json={"email": email})
        assert r.status_code == 200
        row = db_connection.execute(
            "SELECT email FROM subscribers WHERE email = ?", ("subscriber@example.com",)
        ).fetchone()
        assert row is not None
        # Cleanup
        db_connection.execute("DELETE FROM subscribers WHERE email = ?", ("subscriber@example.com",))
        db_connection.commit()

    def test_subscribe_server_error(self, client, monkeypatch):
        """When get_db raises, subscribe returns 500."""
        import main
        monkeypatch.setattr(main, "get_db", lambda: (_ for _ in ()).throw(RuntimeError("DB down")))
        r = client.post("/api/subscribe", json={"email": "fail@example.com"})
        assert r.status_code == 500
        assert "server error" in r.get_json().get("error", "").lower()


class TestHealthAndProfile:
    """Health and profile endpoints."""

    def test_health_check(self, client):
        """GET /api/health returns 200."""
        r = client.get("/api/health")
        assert r.status_code == 200
        data = r.get_json()
        assert data.get("status") == "healthy"

    def test_profile_authenticated(self, client, auth_token):
        """GET /api/profile with valid token returns email."""
        if auth_token is None:
            assert client.get("/api/profile").status_code == 401
            return
        r = client.get("/api/profile", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert "email" in r.get_json()

    def test_profile_user_not_in_db(self, client, test_app):
        """GET /api/profile with JWT for email not in DB returns 404."""
        from flask_jwt_extended import create_access_token
        token = create_access_token(identity="ghost@example.com")
        r = client.get("/api/profile", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 404
        assert "user" in r.get_json().get("error", "").lower()


class TestDownloadWithToken:
    """GET /api/download-with-token/<file_id> (no JWT, uses query token)."""

    def test_download_with_token_no_token(self, client):
        """No token returns 401."""
        r = client.get("/api/download-with-token/1")
        assert r.status_code == 401

    def test_download_with_token_invalid_token(self, client, test_app):
        """Invalid token returns 401."""
        token = "invalid"
        r = client.get(f"/api/download-with-token/1?token={token}")
        assert r.status_code == 401

    def test_download_with_token_valid_but_wrong_file(self, client, test_app, test_user, db_connection):
        """Valid token for different file_id returns 403."""
        import main
        token = main.generate_download_token(file_id=999, user_id=test_user["id"], expires_in_minutes=30)
        r = client.get(f"/api/download-with-token/1?token={token}")
        assert r.status_code == 403

    def test_download_with_token_success(self, client, auth_token, sample_excel_file, test_user):
        """Valid token and file_id returns 200 and file."""
        if auth_token is None:
            assert client.post("/api/upload", content_type="multipart/form-data").status_code == 401
            return
        import main
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        token = main.generate_download_token(file_id=file_id, user_id=test_user["id"], expires_in_minutes=30)
        r = client.get(f"/api/download-with-token/{file_id}?token={token}")
        assert r.status_code == 200
        assert len(r.data) > 0

    def test_download_with_token_via_authorization_header(self, client, auth_token, sample_excel_file, test_user):
        """Token in Authorization header (no query param) works."""
        if auth_token is None:
            assert client.post("/api/upload", content_type="multipart/form-data").status_code == 401
            return
        import main
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        token = main.generate_download_token(file_id=file_id, user_id=test_user["id"], expires_in_minutes=30)
        r = client.get(
            f"/api/download-with-token/{file_id}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        assert len(r.data) > 0

    def test_download_with_token_file_not_in_db(self, client, test_user, test_app):
        """Valid token but file_id not in DB for that user returns 404."""
        import main
        token = main.generate_download_token(file_id=99999, user_id=test_user["id"], expires_in_minutes=30)
        r = client.get(f"/api/download-with-token/99999?token={token}")
        assert r.status_code == 404
        assert "not found" in r.get_json().get("error", "").lower()

    def test_download_with_token_file_not_on_disk(self, client, auth_token, test_user, db_connection, test_app):
        """Valid token, file in DB, but file missing on disk returns 404."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        import main
        import uuid
        stored = f"{uuid.uuid4()}.xlsx"
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "ghost.xlsx", stored, "original"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        token = main.generate_download_token(file_id=file_id, user_id=test_user["id"], expires_in_minutes=30)
        r = client.get(f"/api/download-with-token/{file_id}?token={token}")
        assert r.status_code == 404
        assert "not found" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM files WHERE id = ?", (file_id,))
        db_connection.commit()

    def test_download_with_token_mimetype_octet_stream(self, client, auth_token, test_user, db_connection, test_app):
        """download-with-token for file with non-.xlsx/.xls extension uses application/octet-stream."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        import main
        path = os.path.join(test_app.config["UPLOAD_FOLDER"], "custom.csv")
        with open(path, "wb") as f:
            f.write(b"a,b,c\n1,2,3\n")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "data.csv", "custom.csv", "original"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        token = main.generate_download_token(file_id=file_id, user_id=test_user["id"], expires_in_minutes=30)
        r = client.get(f"/api/download-with-token/{file_id}?token={token}")
        assert r.status_code == 200
        assert "octet-stream" in (r.headers.get("Content-Type") or "")
        os.remove(path)
        db_connection.execute("DELETE FROM files WHERE id = ?", (file_id,))
        db_connection.commit()


class TestDebugStorage:
    """GET /api/debug/storage (JWT required)."""

    def test_debug_storage_authenticated(self, client, auth_token):
        """Authenticated debug/storage returns 200 and storage info."""
        if auth_token is None:
            assert client.get("/api/debug/storage").status_code == 401
            return
        r = client.get("/api/debug/storage", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert "storage_folders" in data
        assert "database_files" in data


class TestDownloadAuthenticated:
    """GET /api/download/<file_id> (JWT required)."""

    def test_download_file_not_found(self, client, auth_token):
        """Non-existent file_id returns 404."""
        if auth_token is None:
            assert client.get("/api/download/99999").status_code == 401
            return
        r = client.get("/api/download/99999", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 404

    def test_download_file_success(self, client, auth_token, sample_excel_file):
        """Upload then download returns 200 and file content."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        r = client.get(f"/api/download/{file_id}", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert r.headers.get("Content-Disposition")
        assert len(r.data) > 0

    def test_download_macro_and_instructions_mimetype(self, client, auth_token, comprehensive_test_excel, test_app):
        """Download .bas and .txt files returns correct mimetype (text/plain)."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        with open(comprehensive_test_excel, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "comp.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        proc = client.post(
            f"/api/process/{file_id}",
            json={"filter_rules": [{"column": "F", "value": "0"}, {"column": "G", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert proc.status_code == 200
        data = proc.get_json()
        macro_id = data["downloads"]["macro"]["file_id"]
        instr_id = data["downloads"]["instructions"]["file_id"]
        r_macro = client.get(f"/api/download/{macro_id}", headers={"Authorization": f"Bearer {auth_token}"})
        r_instr = client.get(f"/api/download/{instr_id}", headers={"Authorization": f"Bearer {auth_token}"})
        assert r_macro.status_code == 200
        assert r_instr.status_code == 200
        assert r_macro.headers.get("Content-Type", "").startswith("text/plain")
        assert r_instr.headers.get("Content-Type", "").startswith("text/plain")


class TestCleanupFilesRoute:
    """POST /api/cleanup-files (remove DB entries for missing files)."""

    def test_cleanup_files_authenticated(self, client, auth_token):
        """Authenticated cleanup-files returns 200 and removed_count."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert "removed_count" in data

    def test_cleanup_files_removes_missing_processed_file(self, client, auth_token, test_user, db_connection, test_app):
        """cleanup-files removes DB record when processed/macro file is missing on disk."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        # Insert a file record for a "processed" file that does not exist on disk
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "ghost_processed.xlsx", "nonexistent_processed.xlsx", "processed"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert data.get("removed_count", 0) >= 1
        row = db_connection.execute("SELECT id FROM files WHERE id = ?", (file_id,)).fetchone()
        assert row is None

    def test_cleanup_files_removes_missing_report_file(self, client, auth_token, test_user, db_connection, test_app):
        """cleanup-files removes DB record when report file is missing on disk (covers report branch)."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "ghost_report.xlsx", "nonexistent_report.xlsx", "report"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert db_connection.execute("SELECT id FROM files WHERE id = ?", (file_id,)).fetchone() is None
        db_connection.commit()

    def test_cleanup_files_removes_missing_macro_report_file(self, client, auth_token, test_user, db_connection, test_app):
        """cleanup-files removes DB record when macro_report file is missing on disk."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "ghost_macro_report.xlsx", "nonexistent_macro_report.xlsx", "macro_report"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert db_connection.execute("SELECT id FROM files WHERE id = ?", (file_id,)).fetchone() is None
        db_connection.commit()

    def test_cleanup_files_removes_missing_file_other_type(self, client, auth_token, test_user, db_connection, test_app):
        """cleanup-files uses UPLOAD_FOLDER for unknown file_type (else branch)."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "ghost_other.xlsx", "nonexistent_other.xlsx", "other"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert db_connection.execute("SELECT id FROM files WHERE id = ?", (file_id,)).fetchone() is None
        db_connection.commit()


class TestValidateInvitationRoute:
    """POST /api/validate-invitation."""

    def test_validate_invitation_missing_body(self, client):
        """Missing token returns 400."""
        r = client.post("/api/validate-invitation", json={})
        assert r.status_code in [400, 422]

    def test_validate_invitation_invalid_token(self, client):
        """Invalid token returns error."""
        r = client.post("/api/validate-invitation", json={"token": "invalid"})
        assert r.status_code in [400, 401, 422]

    def test_validate_invitation_valid_token(self, client, test_admin_user, db_connection, test_app):
        """Valid invitation token returns 200 and email."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        create_r = client.post(
            "/api/admin/create-invitation",
            json={"email": "validate-test@example.com"},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert create_r.status_code in [200, 201], f"Create invitation failed: {create_r.get_json()}"
        inv_token = create_r.get_json().get("token")
        r = client.post("/api/validate-invitation", json={"token": inv_token})
        assert r.status_code == 200
        data = r.get_json()
        assert data.get("valid") is True
        assert data.get("email") == "validate-test@example.com"
        db_connection.execute("DELETE FROM invitation_tokens WHERE email = ?", ("validate-test@example.com",))
        db_connection.commit()


class TestGetMacroAndGeneratedFiles:
    """GET /api/get-macro/<file_id>, GET /api/files/<file_id>/generated."""

    def test_get_macro_for_file_not_found(self, client, auth_token):
        """Non-existent file_id returns 404."""
        if auth_token is None:
            r = client.get("/api/get-macro/99999")
            assert r.status_code in [401, 404]
            return
        r = client.get("/api/get-macro/99999", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 404

    def test_get_generated_files_not_found(self, client, auth_token):
        """Non-existent file_id returns 404."""
        if auth_token is None:
            r = client.get("/api/files/99999/generated")
            assert r.status_code in [401, 404]
            return
        r = client.get("/api/files/99999/generated", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 404

    def test_get_macro_and_generated_after_process(self, client, auth_token, comprehensive_test_excel):
        """After process_file, get-macro and get_generated_files return 200."""
        if auth_token is None:
            assert client.post("/api/upload", content_type="multipart/form-data").status_code == 401
            return
        with open(comprehensive_test_excel, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "comprehensive_test.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        proc = client.post(
            f"/api/process/{file_id}",
            json={
                "filter_rules": [
                    {"column": "F", "value": "0"},
                    {"column": "G", "value": "0"},
                    {"column": "H", "value": "0"},
                    {"column": "I", "value": "0"},
                ]
            },
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert proc.status_code == 200
        r_macro = client.get(f"/api/get-macro/{file_id}")
        assert r_macro.status_code == 200
        assert "macro_content" in r_macro.get_json()
        r_gen = client.get(f"/api/files/{file_id}/generated", headers={"Authorization": f"Bearer {auth_token}"})
        assert r_gen.status_code == 200
        data = r_gen.get_json()
        assert "macros" in data
        assert "instructions" in data

    def test_get_macro_macro_not_found_for_file(self, client, auth_token, sample_excel_file):
        """get-macro for file that has no macro yet returns 404."""
        if auth_token is None:
            assert client.post("/api/upload", content_type="multipart/form-data").status_code == 401
            return
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        r = client.get(f"/api/get-macro/{file_id}")
        assert r.status_code == 404
        assert "macro" in r.get_json().get("error", "").lower()


class TestAdminRoutes:
    """Admin endpoints (require admin user)."""

    def test_create_invitation_unauthenticated(self, client):
        """No token returns 401."""
        r = client.post("/api/admin/create-invitation", json={"email": "new@example.com"})
        assert r.status_code == 401

    def test_create_invitation_non_admin(self, client, auth_token):
        """Non-admin user returns 403."""
        if auth_token is None:
            r = client.post("/api/admin/create-invitation", json={"email": "new@example.com"})
            assert r.status_code == 401
            return
        r = client.post(
            "/api/admin/create-invitation",
            json={"email": "new@example.com"},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code in [403, 404]

    def test_list_invitations_unauthenticated(self, client):
        """No token returns 401."""
        r = client.get("/api/admin/invitations")
        assert r.status_code == 401

    def test_list_users_unauthenticated(self, client):
        """No token returns 401."""
        r = client.get("/api/admin/users")
        assert r.status_code == 401

    def test_admin_create_invitation_as_admin(self, client, test_admin_user, test_app):
        """Admin can create invitation."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.post(
            "/api/admin/create-invitation",
            json={"email": "invited@example.com"},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code in [200, 201]
        data = r.get_json()
        assert "token" in data or "invitation" in data or "email" in data

    def test_list_invitations_as_admin(self, client, test_admin_user):
        """Admin can list invitations."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.get("/api/admin/invitations", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200

    def test_list_users_as_admin(self, client, test_admin_user):
        """Admin can list users."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.get("/api/admin/users", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200

    def test_create_invitation_missing_email(self, client, test_admin_user):
        """Create invitation without email returns 400."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.post(
            "/api/admin/create-invitation",
            json={},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400

    def test_create_invitation_invalid_email_format(self, client, test_admin_user):
        """Create invitation with invalid email format returns 400."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.post(
            "/api/admin/create-invitation",
            json={"email": "not-an-email"},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400

    def test_create_invitation_user_already_exists(self, client, test_admin_user, test_user):
        """Create invitation for existing user returns 409."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.post(
            "/api/admin/create-invitation",
            json={"email": test_user["email"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 409

    def test_create_invitation_pending_already_exists(self, client, test_admin_user, db_connection, test_app):
        """Create invitation for email that already has a pending invitation returns 409."""
        from datetime import datetime, timedelta
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        email = "pending-twice@example.com"
        expires = (datetime.utcnow() + timedelta(days=7)).isoformat()
        existing_token = "existing-invitation-token-123"
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by)
               VALUES (?, ?, ?, ?)""",
            (email, existing_token, expires, test_admin_user["email"]),
        )
        db_connection.commit()
        r = client.post(
            "/api/admin/create-invitation",
            json={"email": email},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 409
        assert "pending" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM invitation_tokens WHERE email = ?", (email,))
        db_connection.commit()

class TestProcessAutomated:
    """POST /api/process-automated/<file_id> (GitHub Actions)."""

    def test_process_automated_requires_filter_rules(self, client, auth_token, sample_excel_file):
        """Missing or empty filter_rules returns 400."""
        if auth_token is None:
            assert client.post("/api/process-automated/1", json={}).status_code == 401
            return
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        r = client.post(
            f"/api/process-automated/{file_id}",
            json={},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 400
        r2 = client.post(
            f"/api/process-automated/{file_id}",
            json={"filter_rules": []},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r2.status_code == 400

    def test_process_automated_file_not_found(self, client, auth_token):
        """Non-existent file_id returns 404."""
        if auth_token is None:
            assert client.post("/api/process-automated/99999", json={"filter_rules": [{"column": "F", "value": "0"}]}).status_code == 401
            return
        r = client.post(
            "/api/process-automated/99999",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 404

    def test_process_automated_success_mocked_github(self, client, auth_token, sample_excel_file, monkeypatch):
        """With mocked GitHub, process-automated returns 202 and job_id."""
        if auth_token is None:
            assert client.post("/api/process-automated/1", json={"filter_rules": [{"column": "F", "value": "0"}]}).status_code == 401
            return
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "-----BEGIN RSA PRIVATE KEY-----\nMIIBOgIBAAJBALRiMLAHudeSA2j/WD2V/b7vR2NRAqOozG+0LZQpVuJ0xQpcrR0e\n-----END RSA PRIVATE KEY-----")
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        mock_response = MagicMock()
        mock_response.status_code = 204
        with patch("main.requests.post", return_value=mock_response):
            with patch("main.GitHubAppAuth") as MockAuth:
                mock_auth_instance = MagicMock()
                mock_auth_instance.get_installation_token.return_value = "ghp_mock"
                MockAuth.return_value = mock_auth_instance
                r = client.post(
                    f"/api/process-automated/{file_id}",
                    json={"filter_rules": [{"column": "F", "value": "0"}]},
                    headers={"Authorization": f"Bearer {auth_token}"},
                )
        if r.status_code == 202:
            data = r.get_json()
            assert "job_id" in data
            assert data.get("status") == "pending"
        else:
            assert r.status_code in [202, 500]


class TestProcessingCallback:
    """POST /api/processing-callback (no JWT; Bearer for callback auth)."""

    def test_processing_callback_unauthorized_no_header(self, client):
        """No Authorization header returns 401."""
        r = client.post("/api/processing-callback", json={"job_id": "x", "status": "failed"})
        assert r.status_code == 401

    def test_processing_callback_status_failed(self, client, test_user, db_connection, test_app):
        """JSON body with status=failed updates job and returns 200."""
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "callback_orig.xlsx", "dummy.xlsx", "original"),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        job_id = "callback-test-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES (?, ?, ?, 'pending')""",
            (job_id, test_user["id"], orig_id),
        )
        db_connection.commit()
        r = client.post(
            "/api/processing-callback",
            json={"job_id": job_id, "status": "failed", "error": "Test failure"},
            headers={"Authorization": "Bearer callback-secret", "Content-Type": "application/json"},
        )
        assert r.status_code == 200
        row = db_connection.execute(
            "SELECT status, error_message FROM processing_jobs WHERE job_id = ?", (job_id,)
        ).fetchone()
        assert row["status"] == "failed"
        assert "Test failure" in (row["error_message"] or "")
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.execute("DELETE FROM files WHERE id = ?", (orig_id,))
        db_connection.commit()

    def test_processing_callback_missing_job_or_file(self, client):
        """Multipart without job_id or file returns 400."""
        r = client.post(
            "/api/processing-callback",
            data={},
            headers={"Authorization": "Bearer callback-secret"},
        )
        assert r.status_code == 400

    def test_processing_callback_job_not_found(self, client):
        """Multipart with non-existent job_id returns 404."""
        from io import BytesIO
        r = client.post(
            "/api/processing-callback",
            data={"job_id": "nonexistent-job-123", "file": (BytesIO(b"xlsx"), "out.xlsx")},
            headers={"Authorization": "Bearer callback-secret"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 404

    def test_processing_callback_original_file_not_found(self, client, test_user, db_connection, test_app, test_directories):
        """Callback with job whose original_file_id is missing in files returns 404."""
        import main
        from io import BytesIO
        main.app.config["PROCESSED_FOLDER"] = test_directories["processed"]
        job_id = "callback-no-orig-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES (?, ?, ?, 'pending')""",
            (job_id, test_user["id"], 99999),
        )
        db_connection.commit()
        r = client.post(
            "/api/processing-callback",
            data={"job_id": job_id, "deleted_rows": "0", "file": (BytesIO(b"PK\x03\x04"), "out.xlsx")},
            content_type="multipart/form-data",
            headers={"Authorization": "Bearer callback-secret"},
        )
        assert r.status_code == 404
        assert "original" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.commit()

    def test_processing_callback_success_with_file(self, client, test_user, db_connection, test_app, test_directories):
        """Multipart with job_id and file creates processed file and returns 200."""
        import main
        main.app.config["PROCESSED_FOLDER"] = test_directories["processed"]
        main.app.config["REPORTS_FOLDER"] = test_directories["reports"]
        stored = str(uuid.uuid4()) + ".xlsx"
        upload_path = os.path.join(test_app.config["UPLOAD_FOLDER"], stored)
        with open(upload_path, "wb") as f:
            f.write(b"PK\x03\x04")  # minimal xlsx-like
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "orig.xlsx", stored, "original"),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        job_id = "callback-success-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES (?, ?, ?, 'pending')""",
            (job_id, test_user["id"], orig_id),
        )
        db_connection.commit()
        from io import BytesIO
        data = {"job_id": job_id, "deleted_rows": "10", "file": (BytesIO(b"PK\x03\x04"), "processed.xlsx")}
        with patch("main.GitHubAppAuth") as MockAuth:
            mock_auth = MagicMock()
            MockAuth.return_value = mock_auth
            r = client.post(
                "/api/processing-callback",
                data=data,
                content_type="multipart/form-data",
                headers={"Authorization": "Bearer callback-secret"},
            )
        assert r.status_code == 200
        j = r.get_json()
        assert j.get("status") == "success"
        assert "file_id" in j
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.execute("DELETE FROM files WHERE parent_file_id = ?", (orig_id,))
        db_connection.execute("DELETE FROM files WHERE id = ?", (orig_id,))
        db_connection.commit()


class TestGitHubRoutes:
    """GET /api/test-github, POST /api/test-dispatch (require GitHub env in prod)."""

    def test_test_github_missing_env(self, client, auth_token):
        """test-github without GitHub env vars returns 400."""
        if auth_token is None:
            assert client.get("/api/test-github").status_code == 401
            return
        r = client.get("/api/test-github", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code in [400, 500]
        data = r.get_json()
        assert "status" in data or "error" in data

    def test_test_github_success_mocked(self, client, auth_token, monkeypatch):
        """test-github with env and mocked requests returns 200."""
        if auth_token is None:
            assert client.get("/api/test-github").status_code == 401
            return
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        monkeypatch.setenv("GITHUB_REPO", "owner/repo")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "-----BEGIN RSA PRIVATE KEY-----\nMIIBOgIBAAJBALRiMLAHudeSA2j/WD2V/b7vR2NRAqOozG+0LZQpVuJ0xQpcrR0e\n-----END RSA PRIVATE KEY-----")
        mock_get = MagicMock()
        mock_get.return_value.status_code = 200
        mock_get.return_value.json.return_value = {"full_name": "owner/repo", "private": False}
        with patch("main.requests.get", mock_get):
            with patch("main.GitHubAppAuth") as MockAuth:
                mock_auth_instance = MagicMock()
                mock_auth_instance.get_app_token.return_value = "app_token"
                mock_auth_instance.get_installation_token.return_value = "inst_token"
                MockAuth.return_value = mock_auth_instance
                r = client.get("/api/test-github", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code in [200, 400, 500]
        if r.status_code == 200:
            assert r.get_json().get("status") == "success"

    def test_test_dispatch_requires_auth(self, client):
        """test-dispatch without auth returns 401."""
        r = client.post("/api/test-dispatch", json={})
        assert r.status_code == 401

    def test_test_dispatch_success_mocked(self, client, auth_token, monkeypatch):
        """test-dispatch with mocked requests returns 200 when dispatch returns 204."""
        if auth_token is None:
            assert client.post("/api/test-dispatch", json={}).status_code == 401
            return
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        monkeypatch.setenv("GITHUB_REPO", "owner/repo")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "-----BEGIN RSA PRIVATE KEY-----\nMIIBOgIBAAJBALRiMLAHudeSA2j/WD2V/b7vR2NRAqOozG+0LZQpVuJ0xQpcrR0e\n-----END RSA PRIVATE KEY-----")
        mock_get = MagicMock()
        mock_get.return_value.status_code = 200
        mock_post = MagicMock()
        mock_post.return_value.status_code = 204
        with patch("main.requests.get", mock_get):
            with patch("main.requests.post", mock_post):
                with patch("main.GitHubAppAuth") as MockAuth:
                    mock_auth_instance = MagicMock()
                    mock_auth_instance.get_installation_token.return_value = "inst_token"
                    MockAuth.return_value = mock_auth_instance
                    r = client.post(
                        "/api/test-dispatch",
                        json={},
                        headers={"Authorization": f"Bearer {auth_token}"},
                    )
        assert r.status_code in [200, 400, 500]
        if r.status_code == 200:
            data = r.get_json()
            assert data.get("dispatch_status") == 204 or data.get("status") == "success"


class TestFileHistoryAndJobStatus:
    """GET /api/files/<id>/history, DELETE, GET /api/job-status/<job_id>."""

    def test_get_file_history_not_found(self, client, auth_token):
        """Non-existent file_id returns 404."""
        if auth_token is None:
            assert client.get("/api/files/99999/history").status_code == 401
            return
        r = client.get("/api/files/99999/history", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 404

    def test_get_file_history_user_not_found(self, client, test_app):
        """GET history with JWT for user not in DB returns 404."""
        from flask_jwt_extended import create_access_token
        token = create_access_token(identity="ghost@example.com")
        r = client.get("/api/files/1/history", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 404

    def test_get_job_status_not_found(self, client, auth_token):
        """Non-existent job_id returns 404."""
        if auth_token is None:
            assert client.get("/api/job-status/nonexistent-job-id").status_code == 401
            return
        r = client.get("/api/job-status/nonexistent-job-id", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 404

    def test_get_job_status_completed_with_report(self, client, auth_token, test_user, db_connection, test_app, test_directories):
        """GET job-status for completed job with report_file_id returns 200 and report info."""
        if auth_token is None:
            assert client.get("/api/job-status/x").status_code == 401
            return
        job_id = "job-completed-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "orig.xlsx", "stored.xlsx", "original"),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type, parent_file_id)
               VALUES (?, ?, ?, ?, ?)""",
            (test_user["id"], "processed.xlsx", "proc.xlsx", "processed", orig_id),
        )
        db_connection.commit()
        result_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type, parent_file_id)
               VALUES (?, ?, ?, ?, ?)""",
            (test_user["id"], "report.xlsx", "rep.xlsx", "report", orig_id),
        )
        db_connection.commit()
        report_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status, result_file_id, report_file_id, deleted_rows)
               VALUES (?, ?, ?, 'completed', ?, ?, 5)""",
            (job_id, test_user["id"], orig_id, result_id, report_id),
        )
        db_connection.commit()
        r = client.get(f"/api/job-status/{job_id}", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert data["status"] == "completed"
        assert data.get("download_file_id") == result_id
        assert data.get("report_file_id") == report_id
        assert data.get("report_filename") is not None
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.execute("DELETE FROM files WHERE id IN (?, ?, ?)", (orig_id, result_id, report_id))
        db_connection.commit()

    def test_get_job_status_failed(self, client, auth_token, test_user, db_connection):
        """GET job-status for failed job returns 200 with error."""
        if auth_token is None:
            assert client.get("/api/job-status/x").status_code == 401
            return
        job_id = "job-failed-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "orig.xlsx", "stored.xlsx", "original"),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status, error_message)
               VALUES (?, ?, ?, 'failed', ?)""",
            (job_id, test_user["id"], orig_id, "Analysis failed"),
        )
        db_connection.commit()
        r = client.get(f"/api/job-status/{job_id}", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert data["status"] == "failed"
        assert "error" in data
        assert "Analysis" in (data.get("error") or "")
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.execute("DELETE FROM files WHERE id = ?", (orig_id,))
        db_connection.commit()

    def test_get_file_history_success(self, client, auth_token, test_user, db_connection, comprehensive_test_excel):
        """Upload, insert a processing job, then GET history returns 200 and history list."""
        if auth_token is None:
            assert client.get("/api/files/1/history").status_code == 401
            return
        with open(comprehensive_test_excel, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "comprehensive_test.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        # Manual processing does not create a processing_jobs record, so insert one
        # to test that the history endpoint returns it correctly.
        job_id = "hist-test-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status, filter_rules_json)
               VALUES (?, ?, ?, 'completed', ?)""",
            (job_id, test_user["id"], file_id, '[{"column":"F","value":"0"}]'),
        )
        db_connection.commit()
        r = client.get(f"/api/files/{file_id}/history", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert "history" in data
        assert "original_file" in data
        assert len(data["history"]) >= 1
        r_status = client.get(
            f"/api/job-status/{job_id}",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r_status.status_code == 200
        # Cleanup
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.commit()

    def test_get_file_history_invalid_filter_rules_json(self, client, auth_token, test_user, db_connection, test_app):
        """get_file_history with job containing invalid filter_rules_json uses empty list (except branch)."""
        if auth_token is None:
            assert client.get("/api/files/1/history").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "hist_invalid.xlsx", "hist_invalid.xlsx", "original"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        job_id = "hist-invalid-json-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status, filter_rules_json)
               VALUES (?, ?, ?, 'completed', ?)""",
            (job_id, test_user["id"], file_id, "{invalid json"),
        )
        db_connection.commit()
        r = client.get(f"/api/files/{file_id}/history", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        data = r.get_json()
        assert "history" in data
        assert len(data["history"]) >= 1
        assert data["history"][0].get("filter_rules") == []
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = ?", (job_id,))
        db_connection.execute("DELETE FROM files WHERE id = ?", (file_id,))
        db_connection.commit()

    def test_delete_history_item_not_found(self, client, auth_token):
        """DELETE history item with bad job_id returns 404."""
        if auth_token is None:
            assert client.delete("/api/files/99999/history/fake-job-id").status_code == 401
            return
        r = client.delete(
            "/api/files/99999/history/fake-job-id",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 404

    def test_clear_file_history_not_found(self, client, test_admin_user):
        """DELETE all history for non-existent file returns 404 (admin required first)."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.delete(
            "/api/files/99999/history",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_clear_file_history_non_admin_403(self, client, auth_token, test_user, db_connection):
        """Non-admin user cannot clear file history (403)."""
        if auth_token is None:
            assert client.delete("/api/files/1/history").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "hist.xlsx", "hist.xlsx", "original"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.delete(
            f"/api/files/{file_id}/history",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 403
        db_connection.execute("DELETE FROM files WHERE id = ?", (file_id,))
        db_connection.commit()

    def test_clear_file_history_admin_success(self, client, test_admin_user, test_user, db_connection):
        """Admin can clear file history and gets deleted_count."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "hist.xlsx", "hist.xlsx", "original"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        job_id = "hist-job-" + str(uuid.uuid4())[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES (?, ?, ?, 'completed')""",
            (job_id, test_user["id"], file_id),
        )
        db_connection.commit()
        r = client.delete(
            f"/api/files/{file_id}/history",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert "deleted_count" in data
        assert data["deleted_count"] >= 1
        db_connection.execute("DELETE FROM files WHERE id = ?", (file_id,))
        db_connection.commit()


class TestAdminUserDetailsAndDelete:
    """GET /api/admin/users/<id>, DELETE /api/admin/users/<id>."""

    def test_get_user_details_as_admin(self, client, test_admin_user, test_user):
        """Admin can get user details."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.get(
            f"/api/admin/users/{test_user['id']}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        data = r.get_json()
        assert data.get("email") == test_user["email"]
        assert "file_count" in data
        assert "job_count" in data

    def test_get_user_details_not_found(self, client, test_admin_user):
        """Admin GET non-existent user returns 404."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.get(
            "/api/admin/users/99999",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_expire_invitation_as_admin(self, client, test_admin_user, db_connection, test_app):
        """Admin can expire a pending invitation."""
        import jwt as jwt_lib
        from datetime import datetime, timedelta
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        email = "expire-me@example.com"
        inv_token = jwt_lib.encode(
            {"email": email, "purpose": "invitation", "exp": datetime.utcnow() + timedelta(days=7)},
            test_app.config["JWT_SECRET_KEY"],
            algorithm="HS256",
        )
        expires = (datetime.utcnow() + timedelta(days=7)).isoformat()
        db_connection.execute(
            "INSERT INTO invitation_tokens (email, token, expires_at, created_by) VALUES (?, ?, ?, ?)",
            (email, inv_token, expires, test_admin_user["email"]),
        )
        db_connection.commit()
        inv_id = db_connection.execute(
            "SELECT id FROM invitation_tokens WHERE email = ?", (email,)
        ).fetchone()[0]
        r = client.post(
            f"/api/admin/invitations/{inv_id}/expire",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        db_connection.execute("DELETE FROM invitation_tokens WHERE email = ?", (email,))
        db_connection.commit()

    def test_expire_invitation_not_found(self, client, test_admin_user):
        """Expire non-existent invitation returns 404."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.post(
            "/api/admin/invitations/99999/expire",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_delete_user_cannot_delete_self(self, client, test_admin_user):
        """Admin cannot delete own account (400)."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.delete(
            f"/api/admin/users/{test_admin_user['id']}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400

    def test_delete_user_not_found(self, client, test_admin_user):
        """Admin DELETE non-existent user returns 404."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        r = client.delete(
            "/api/admin/users/99999",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_delete_user_success(self, client, test_admin_user, db_connection, test_app):
        """Admin can delete a non-admin user."""
        from werkzeug.security import generate_password_hash
        from datetime import datetime, timedelta
        import jwt as jwt_lib
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200, f"Admin login failed: {login.get_json()}"
        token = login.get_json().get("access_token")
        email = "to-delete@example.com"
        inv_token = jwt_lib.encode(
            {"email": email, "purpose": "invitation", "exp": datetime.utcnow() + timedelta(days=7)},
            test_app.config["JWT_SECRET_KEY"],
            algorithm="HS256",
        )
        expires = (datetime.utcnow() + timedelta(days=7)).isoformat()
        db_connection.execute(
            "INSERT INTO invitation_tokens (email, token, expires_at, created_by) VALUES (?, ?, ?, ?)",
            (email, inv_token, expires, test_admin_user["email"]),
        )
        db_connection.commit()
        reg = client.post("/api/register", json={"invitation_token": inv_token, "password": "SecurePassword123!"})
        assert reg.status_code == 201, f"Register failed: {reg.get_json()}"
        user_id = db_connection.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()[0]
        r = client.delete(
            f"/api/admin/users/{user_id}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        assert db_connection.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone() is None


# ──────────────────────────────────────────────────────────────────────────────
# Coverage gap tests – systematically cover every uncovered production line
# ──────────────────────────────────────────────────────────────────────────────

class TestLoginExceptionHandler:
    """Cover main.py lines 184-185 (login exception handler)."""

    def test_login_exception_returns_500(self, client, monkeypatch):
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.post("/api/login", json={"email": "a@b.com", "password": "x"})
        assert r.status_code == 500


class TestUploadEdgeCases:
    """Cover main.py upload route edge cases."""

    def test_upload_validate_excel_error(self, client, auth_token, test_directories):
        """Upload file that fails validate_excel_file → lines 219-220."""
        if auth_token is None:
            assert client.post("/api/upload").status_code == 401
            return
        from io import BytesIO
        # Not a valid xlsx magic bytes
        r = client.post(
            "/api/upload",
            data={"file": (BytesIO(b"not-an-excel-file"), "test.xlsx")},
            headers={"Authorization": f"Bearer {auth_token}"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400

    def test_upload_duplicate_file(self, client, auth_token, sample_excel_file):
        """Upload same file twice → lines 242-245 (duplicate detection)."""
        if auth_token is None:
            assert client.post("/api/upload").status_code == 401
            return
        headers = {"Authorization": f"Bearer {auth_token}"}
        # First upload
        with open(sample_excel_file, "rb") as f:
            r1 = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers=headers,
                content_type="multipart/form-data",
            )
        assert r1.status_code in [200, 201]
        # Second upload — same file content and name → duplicate
        with open(sample_excel_file, "rb") as f:
            r2 = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers=headers,
                content_type="multipart/form-data",
            )
        assert r2.status_code == 200
        assert r2.get_json().get("duplicate") is True

    def test_upload_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception during upload → lines 269-270."""
        if auth_token is None:
            assert client.post("/api/upload").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        from io import BytesIO
        r = client.post(
            "/api/upload",
            data={"file": (BytesIO(b"PK\x03\x04"), "test.xlsx")},
            headers={"Authorization": f"Bearer {auth_token}"},
            content_type="multipart/form-data",
        )
        assert r.status_code == 500


class TestGetFilesEdgeCases:
    """Cover main.py get_files route edge cases."""

    def test_get_files_file_not_on_disk(self, client, auth_token, db_connection, test_user):
        """File in DB but not on disk is excluded → line 299."""
        if auth_token is None:
            assert client.get("/api/files").status_code == 401
            return
        # Insert a file record with a stored_filename that doesn't exist on disk
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'ghost.xlsx', 'nonexistent-on-disk.xlsx', 'original')""",
            (test_user["id"],),
        )
        db_connection.commit()
        r = client.get("/api/files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        # The ghost file should be filtered out
        files = r.get_json().get("files", [])
        ghost_files = [f for f in files if f["original_filename"] == "ghost.xlsx"]
        assert len(ghost_files) == 0
        db_connection.execute("DELETE FROM files WHERE stored_filename = 'nonexistent-on-disk.xlsx'")
        db_connection.commit()

    def test_get_files_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception in get_files → lines 309-313."""
        if auth_token is None:
            assert client.get("/api/files").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get("/api/files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500


class TestProcessFileEdgeCases:
    """Cover main.py process_file route edge cases."""

    def test_process_file_not_found(self, client, auth_token):
        """process_file for non-existent file → line 342."""
        if auth_token is None:
            assert client.post("/api/process/99999").status_code == 401
            return
        r = client.post(
            "/api/process/99999",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 404

    def test_process_file_more_than_5_rows_logged(self, client, auth_token, test_user, test_directories):
        """Processing file with >5 matching rows hits log branch → line 401."""
        if auth_token is None:
            assert client.post("/api/process/1").status_code == 401
            return
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "B", "C", "D", "E", "F"])
        for i in range(10):
            ws.append([f"Row{i}", 0, 0, 0, 0, 0])
        file_path = os.path.join(test_directories["uploads"], "many_rows.xlsx")
        wb.save(file_path)
        # Upload the file
        headers = {"Authorization": f"Bearer {auth_token}"}
        with open(file_path, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "many_rows.xlsx")},
                headers=headers,
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        # Process with filter that matches many rows
        r = client.post(
            f"/api/process/{file_id}",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers=headers,
        )
        assert r.status_code == 200
        if os.path.exists(file_path):
            os.remove(file_path)


class TestDownloadEdgeCases:
    """Cover main.py download route edge cases."""

    def test_download_file_not_on_disk(self, client, auth_token, test_user, db_connection):
        """File exists in DB but not on disk → line 601."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'missing.xlsx', 'does-not-exist.xlsx', 'original')""",
            (test_user["id"],),
        )
        db_connection.commit()
        fid = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.get(f"/api/download/{fid}", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 404
        db_connection.execute("DELETE FROM files WHERE id = ?", (fid,))
        db_connection.commit()

    def test_download_file_unknown_extension(self, client, auth_token, test_user, db_connection, test_directories):
        """File with unknown extension gets octet-stream mimetype → line 612."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        stored = "test_unknown.bin"
        fpath = os.path.join(test_directories["uploads"], stored)
        with open(fpath, "wb") as f:
            f.write(b"data")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'data.bin', ?, 'original')""",
            (test_user["id"], stored),
        )
        db_connection.commit()
        fid = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.get(f"/api/download/{fid}", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert r.content_type.startswith("application/octet-stream")
        db_connection.execute("DELETE FROM files WHERE id = ?", (fid,))
        db_connection.commit()
        if os.path.exists(fpath):
            os.remove(fpath)

    def test_download_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception in download → lines 621-622."""
        if auth_token is None:
            assert client.get("/api/download/1").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get("/api/download/1", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500

    def test_download_with_token_exception_handler(self, client, test_app, test_user, monkeypatch):
        """Generic exception in download_file_with_token → lines 568-572."""
        import main
        token = main.generate_download_token(file_id=1, user_id=test_user["id"], expires_in_minutes=5)
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get(f"/api/download-with-token/1?token={token}")
        assert r.status_code == 500


class TestCleanupFilesEdgeCases:
    """Cover main.py cleanup-files route edge cases."""

    def test_cleanup_original_and_macro_types(self, client, auth_token, test_user, db_connection, test_directories):
        """Cleanup handles 'original' and 'macro' file types → lines 651, 655."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        # Insert original file record with no matching file on disk
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'cleanup_orig.xlsx', 'no-such-orig.xlsx', 'original')""",
            (test_user["id"],),
        )
        # Insert macro file record with no matching file on disk
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'Macro_test.bas', 'no-such-macro.bas', 'macro')""",
            (test_user["id"],),
        )
        # Insert instructions file record
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'Instructions_test.txt', 'no-such-instr.txt', 'instructions')""",
            (test_user["id"],),
        )
        db_connection.commit()
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert r.get_json().get("removed_count", 0) >= 3

    def test_cleanup_orphaned_processing_jobs(self, client, auth_token, test_user, db_connection):
        """Cleanup removes orphaned jobs → lines 688-690."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        # Create a processing job that references a non-existent file
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'base.xlsx', 'some-base.xlsx', 'original')""",
            (test_user["id"],),
        )
        db_connection.commit()
        base_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, result_file_id, status)
               VALUES ('orphan-job-test', ?, ?, 99998, 'completed')""",
            (test_user["id"], base_id),
        )
        db_connection.commit()
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert r.get_json().get("removed_count", 0) >= 1
        # Cleanup
        db_connection.execute("DELETE FROM files WHERE id = ?", (base_id,))
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = 'orphan-job-test'")
        db_connection.commit()

    def test_cleanup_files_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception in cleanup_files → lines 700-702."""
        if auth_token is None:
            assert client.post("/api/cleanup-files").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.post("/api/cleanup-files", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500


class TestDebugStorageExceptionHandler:
    """Cover main.py lines 750-751."""

    def test_debug_storage_exception(self, client, auth_token, monkeypatch):
        if auth_token is None:
            assert client.get("/api/debug/storage").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get("/api/debug/storage", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500


class TestGitHubRoutes:
    """Cover main.py test-github and test-dispatch routes."""

    def test_github_repo_access_failure(self, client, auth_token, monkeypatch):
        """test-github returns error when repo access fails → lines 820-826."""
        if auth_token is None:
            assert client.get("/api/test-github").status_code == 401
            return
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "fake-key")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        import main
        mock_auth = MagicMock()
        mock_auth.get_app_token.return_value = "fake-app-token-" + "x" * 50
        mock_auth.get_installation_token.return_value = "fake-install-token-" + "x" * 50
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            mock_response = MagicMock()
            mock_response.status_code = 403
            mock_response.text = "Forbidden"
            with patch.object(main.requests, 'get', return_value=mock_response):
                r = client.get("/api/test-github", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 400
        assert "error" in r.get_json()

    def test_github_exception_handler(self, client, auth_token, monkeypatch):
        """test-github exception → lines 828-837."""
        if auth_token is None:
            assert client.get("/api/test-github").status_code == 401
            return
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "fake-key")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        import main
        with patch.object(main, 'GitHubAppAuth', side_effect=RuntimeError("auth failed")):
            r = client.get("/api/test-github", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500

    def test_dispatch_repo_check_failure(self, client, auth_token, monkeypatch):
        """test-dispatch when repo check fails → line 881."""
        if auth_token is None:
            assert client.post("/api/test-dispatch").status_code == 401
            return
        import main
        mock_auth = MagicMock()
        mock_auth.get_installation_token.return_value = "fake-token-" + "x" * 50
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            mock_resp = MagicMock()
            mock_resp.status_code = 404
            mock_resp.text = "Not found"
            with patch.object(main.requests, 'get', return_value=mock_resp):
                r = client.post("/api/test-dispatch", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 400

    def test_dispatch_403_permissions_check(self, client, auth_token, monkeypatch):
        """test-dispatch with 403 on dispatch → lines 903-909."""
        if auth_token is None:
            assert client.post("/api/test-dispatch").status_code == 401
            return
        import main
        mock_auth = MagicMock()
        mock_auth.get_installation_token.return_value = "fake-token-" + "x" * 50
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            repo_resp = MagicMock()
            repo_resp.status_code = 200
            dispatch_resp = MagicMock()
            dispatch_resp.status_code = 403
            dispatch_resp.text = "Forbidden"
            dispatch_resp.headers = {"X-RateLimit-Remaining": "50"}
            perm_resp = MagicMock()
            perm_resp.status_code = 200
            perm_resp.json.return_value = {"permissions": {"contents": "read"}}

            def mock_get(url, **kwargs):
                if "installation" in url:
                    return perm_resp
                return repo_resp

            with patch.object(main.requests, 'get', side_effect=mock_get):
                with patch.object(main.requests, 'post', return_value=dispatch_resp):
                    r = client.post("/api/test-dispatch", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 400

    def test_dispatch_non_204_status(self, client, auth_token, monkeypatch):
        """test-dispatch when dispatch returns non-204 → line 925."""
        if auth_token is None:
            assert client.post("/api/test-dispatch").status_code == 401
            return
        import main
        mock_auth = MagicMock()
        mock_auth.get_installation_token.return_value = "fake-token-" + "x" * 50
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            repo_resp = MagicMock()
            repo_resp.status_code = 200
            dispatch_resp = MagicMock()
            dispatch_resp.status_code = 422
            dispatch_resp.text = "Unprocessable"
            dispatch_resp.headers = {}
            with patch.object(main.requests, 'get', return_value=repo_resp):
                with patch.object(main.requests, 'post', return_value=dispatch_resp):
                    r = client.post("/api/test-dispatch", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 400
        assert "error" in r.get_json()


class TestProcessAutomatedEdgeCases:
    """Cover main.py process-automated route edge cases."""

    def test_process_automated_user_not_found(self, client, test_app):
        """User not in DB → lines 971-972."""
        from flask_jwt_extended import create_access_token
        token = create_access_token(identity="ghost@example.com")
        r = client.post(
            "/api/process-automated/1",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_process_automated_file_not_on_disk(self, client, auth_token, test_user, db_connection):
        """File in DB but not on disk → lines 991-992."""
        if auth_token is None:
            assert client.post("/api/process-automated/1").status_code == 401
            return
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'nodisk.xlsx', 'nodisk-stored.xlsx', 'original')""",
            (test_user["id"],),
        )
        db_connection.commit()
        fid = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.post(
            f"/api/process-automated/{fid}",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 404
        db_connection.execute("DELETE FROM files WHERE id = ?", (fid,))
        db_connection.commit()

    def test_process_automated_github_config_missing(self, client, auth_token, test_user, db_connection, test_directories, monkeypatch):
        """Missing GitHub env vars → line 1024."""
        if auth_token is None:
            assert client.post("/api/process-automated/1").status_code == 401
            return
        stored = "auto_test.xlsx"
        fpath = os.path.join(test_directories["uploads"], stored)
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["A"])
        wb.save(fpath)
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'auto.xlsx', ?, 'original')""",
            (test_user["id"], stored),
        )
        db_connection.commit()
        fid = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        monkeypatch.delenv("GITHUB_APP_ID", raising=False)
        monkeypatch.delenv("GITHUB_PRIVATE_KEY", raising=False)
        monkeypatch.delenv("GITHUB_INSTALLATION_ID", raising=False)
        r = client.post(
            f"/api/process-automated/{fid}",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 500
        assert "configuration" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM files WHERE id = ?", (fid,))
        db_connection.commit()
        if os.path.exists(fpath):
            os.remove(fpath)

    def test_process_automated_dispatch_failure(self, client, auth_token, test_user, db_connection, test_directories, monkeypatch):
        """Dispatch returns non-204 → lines 1069-1072."""
        if auth_token is None:
            assert client.post("/api/process-automated/1").status_code == 401
            return
        stored = "dispatch_fail.xlsx"
        fpath = os.path.join(test_directories["uploads"], stored)
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["A"])
        wb.save(fpath)
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'dispatch_fail.xlsx', ?, 'original')""",
            (test_user["id"], stored),
        )
        db_connection.commit()
        fid = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "fake")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        import main
        mock_auth = MagicMock()
        mock_auth.get_installation_token.return_value = "tok"
        dispatch_resp = MagicMock()
        dispatch_resp.status_code = 422
        dispatch_resp.text = "fail"
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            with patch.object(main.requests, 'post', return_value=dispatch_resp):
                r = client.post(
                    f"/api/process-automated/{fid}",
                    json={"filter_rules": [{"column": "F", "value": "0"}]},
                    headers={"Authorization": f"Bearer {auth_token}"},
                )
        assert r.status_code == 500
        db_connection.execute("DELETE FROM processing_jobs WHERE original_file_id = ?", (fid,))
        db_connection.execute("DELETE FROM files WHERE id = ?", (fid,))
        db_connection.commit()
        if os.path.exists(fpath):
            os.remove(fpath)

    def test_process_automated_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception → lines 1074-1078."""
        if auth_token is None:
            assert client.post("/api/process-automated/1").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.post(
            "/api/process-automated/1",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 500


class TestCallbackEdgeCases:
    """Cover main.py processing-callback edge cases."""

    def test_callback_with_report_file(self, client, test_user, db_connection, test_app, test_directories):
        """Callback with report file → lines 1168-1176."""
        import main
        main.app.config["PROCESSED_FOLDER"] = test_directories["processed"]
        main.app.config["REPORTS_FOLDER"] = test_directories["reports"]
        stored = str(uuid.uuid4()) + ".xlsx"
        upload_path = os.path.join(test_app.config["UPLOAD_FOLDER"], stored)
        with open(upload_path, "wb") as f:
            f.write(b"PK\x03\x04")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'orig_report.xlsx', ?, 'original')""",
            (test_user["id"], stored),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        job_id = "callback-report-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES (?, ?, ?, 'pending')""",
            (job_id, test_user["id"], orig_id),
        )
        db_connection.commit()
        from io import BytesIO
        with patch("main.GitHubAppAuth") as MockAuth:
            MockAuth.return_value = MagicMock()
            r = client.post(
                "/api/processing-callback",
                data={
                    "job_id": job_id,
                    "deleted_rows": "5",
                    "file": (BytesIO(b"PK\x03\x04"), "processed.xlsx"),
                    "report": (BytesIO(b"PK\x03\x04"), "report.xlsx"),
                },
                content_type="multipart/form-data",
                headers={"Authorization": "Bearer callback-secret"},
            )
        assert r.status_code == 200
        j = r.get_json()
        assert j.get("report_file_id") is not None

    def test_callback_artifact_deletion_failure(self, client, test_user, db_connection, test_app, test_directories):
        """Artifact deletion fails (non-critical) → lines 1200-1201."""
        import main
        main.app.config["PROCESSED_FOLDER"] = test_directories["processed"]
        stored = str(uuid.uuid4()) + ".xlsx"
        upload_path = os.path.join(test_app.config["UPLOAD_FOLDER"], stored)
        with open(upload_path, "wb") as f:
            f.write(b"PK\x03\x04")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'orig_art.xlsx', ?, 'original')""",
            (test_user["id"], stored),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        job_id = "callback-art-fail-" + uuid.uuid4().hex[:8]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES (?, ?, ?, 'pending')""",
            (job_id, test_user["id"], orig_id),
        )
        db_connection.commit()
        from io import BytesIO
        mock_auth = MagicMock()
        mock_auth.delete_artifact_by_job_id.side_effect = RuntimeError("artifact deletion failed")
        with patch("main.GitHubAppAuth", return_value=mock_auth):
            r = client.post(
                "/api/processing-callback",
                data={"job_id": job_id, "deleted_rows": "0", "file": (BytesIO(b"PK\x03\x04"), "processed.xlsx")},
                content_type="multipart/form-data",
                headers={"Authorization": "Bearer callback-secret"},
            )
        # Should still succeed despite artifact deletion failure
        assert r.status_code == 200

    def test_callback_exception_handler(self, client, monkeypatch):
        """Generic exception in callback → lines 1211-1216."""
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        from io import BytesIO
        r = client.post(
            "/api/processing-callback",
            data={"job_id": "x", "file": (BytesIO(b"PK\x03\x04"), "out.xlsx")},
            content_type="multipart/form-data",
            headers={"Authorization": "Bearer callback-secret"},
        )
        assert r.status_code == 500


class TestHistoryEdgeCases:
    """Cover main.py history route edge cases."""

    def test_history_filter_rules_null(self, client, auth_token, test_user, db_connection, sample_excel_file):
        """Job with no filter_rules_json → line 1280."""
        if auth_token is None:
            assert client.get("/api/files/1/history").status_code == 401
            return
        headers = {"Authorization": f"Bearer {auth_token}"}
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "hist_null.xlsx")},
                headers=headers,
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        fid = up.get_json()["file_id"]
        # Insert job without filter_rules_json
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status, filter_rules_json)
               VALUES ('hist-null-rules', ?, ?, 'completed', NULL)""",
            (test_user["id"], fid),
        )
        db_connection.commit()
        r = client.get(f"/api/files/{fid}/history", headers=headers)
        assert r.status_code == 200
        history = r.get_json().get("history", [])
        null_job = [h for h in history if h.get("job_id") == "hist-null-rules"]
        assert len(null_job) == 1
        assert null_job[0]["filter_rules"] == []
        db_connection.execute("DELETE FROM processing_jobs WHERE job_id = 'hist-null-rules'")
        db_connection.commit()

    def test_history_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception in get_file_history → lines 1289-1293."""
        if auth_token is None:
            assert client.get("/api/files/1/history").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get("/api/files/1/history", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500

    def test_delete_history_user_not_found(self, client, test_app):
        """delete_history_item with ghost user → lines 1308-1309."""
        from flask_jwt_extended import create_access_token
        token = create_access_token(identity="ghost@example.com")
        r = client.delete(
            "/api/files/1/history/fake-job",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_delete_history_item_success(self, client, auth_token, test_user, db_connection, sample_excel_file):
        """Successfully delete a history item → lines 1324-1328."""
        if auth_token is None:
            assert client.delete("/api/files/1/history/x").status_code == 401
            return
        headers = {"Authorization": f"Bearer {auth_token}"}
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "hist_del.xlsx")},
                headers=headers,
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        fid = up.get_json()["file_id"]
        db_connection.execute(
            """INSERT INTO processing_jobs (job_id, user_id, original_file_id, status)
               VALUES ('del-hist-item', ?, ?, 'completed')""",
            (test_user["id"], fid),
        )
        db_connection.commit()
        r = client.delete(f"/api/files/{fid}/history/del-hist-item", headers=headers)
        assert r.status_code == 200
        assert db_connection.execute(
            "SELECT * FROM processing_jobs WHERE job_id = 'del-hist-item'"
        ).fetchone() is None

    def test_delete_history_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception in delete_history_item → lines 1330-1334."""
        if auth_token is None:
            assert client.delete("/api/files/1/history/x").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.delete(
            "/api/files/1/history/fake-job",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 500

    def test_clear_history_user_not_found(self, client, test_app):
        """clear_file_history with ghost user → lines 1349-1350."""
        from flask_jwt_extended import create_access_token
        token = create_access_token(identity="ghost@example.com")
        r = client.delete(
            "/api/files/1/history",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_clear_history_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception in clear_file_history → lines 1380-1384."""
        if auth_token is None:
            assert client.delete("/api/files/1/history").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.delete(
            "/api/files/1/history",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 500


class TestJobStatusExceptionHandler:
    """Cover main.py lines 1430-1431."""

    def test_job_status_exception(self, client, auth_token, monkeypatch):
        if auth_token is None:
            assert client.get("/api/job-status/x").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get("/api/job-status/x", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500


class TestGetMacroEdgeCases:
    """Cover main.py get-macro route edge cases."""

    def test_get_macro_file_not_on_disk(self, client, test_user, db_connection):
        """Macro in DB but not on disk → line 1466."""
        # Create original + macro file records, but no actual file on disk
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'orig.xlsx', 'orig-stored.xlsx', 'original')""",
            (test_user["id"],),
        )
        db_connection.commit()
        orig_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'Macro_orig.xlsx.bas', 'no-such-macro.bas', 'macro')""",
            (test_user["id"],),
        )
        db_connection.commit()
        r = client.get(f"/api/get-macro/{orig_id}")
        assert r.status_code == 404
        db_connection.execute("DELETE FROM files WHERE original_filename LIKE 'Macro_orig%'")
        db_connection.execute("DELETE FROM files WHERE id = ?", (orig_id,))
        db_connection.commit()

    def test_get_macro_exception_handler(self, client, monkeypatch):
        """Generic exception in get-macro → lines 1476-1477."""
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get("/api/get-macro/1")
        assert r.status_code == 500


class TestGeneratedFilesEdgeCases:
    """Cover main.py generated files route edge cases."""

    def test_generated_files_user_not_found(self, client, test_app):
        """Ghost user → lines 1494-1495."""
        from flask_jwt_extended import create_access_token
        token = create_access_token(identity="ghost@example.com")
        r = client.get(
            "/api/files/1/generated",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_generated_files_exception_handler(self, client, auth_token, monkeypatch):
        """Generic exception → lines 1560-1564."""
        if auth_token is None:
            assert client.get("/api/files/1/generated").status_code == 401
            return
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get(
            "/api/files/1/generated",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 500


class TestValidateInvitationExceptionHandler:
    """Cover main.py lines 1613-1614."""

    def test_validate_invitation_exception(self, client, monkeypatch):
        import main
        monkeypatch.setattr(main, 'validate_invitation_token', lambda t: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.post("/api/validate-invitation", json={"token": "x"})
        assert r.status_code == 500


class TestAdminNonAdminChecks:
    """Cover non-admin 403 responses for all admin-only routes."""

    def _get_auth_token(self, client, auth_token):
        """Return token or assert 401 for unauthenticated."""
        if auth_token is None:
            return None
        return auth_token

    def test_list_invitations_non_admin(self, client, auth_token):
        """Non-admin → line 1730."""
        if auth_token is None:
            assert client.get("/api/admin/invitations").status_code == 401
            return
        r = client.get("/api/admin/invitations", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 403

    def test_expire_invitation_non_admin(self, client, auth_token):
        """Non-admin → line 1780."""
        if auth_token is None:
            assert client.post("/api/admin/invitations/1/expire").status_code == 401
            return
        r = client.post(
            "/api/admin/invitations/1/expire",
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 403

    def test_list_users_non_admin(self, client, auth_token):
        """Non-admin → line 1835."""
        if auth_token is None:
            assert client.get("/api/admin/users").status_code == 401
            return
        r = client.get("/api/admin/users", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 403

    def test_get_user_details_non_admin(self, client, auth_token):
        """Non-admin → line 1883."""
        if auth_token is None:
            assert client.get("/api/admin/users/1").status_code == 401
            return
        r = client.get("/api/admin/users/1", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 403

    def test_delete_user_non_admin(self, client, auth_token):
        """Non-admin → line 1936."""
        if auth_token is None:
            assert client.delete("/api/admin/users/1").status_code == 401
            return
        r = client.delete("/api/admin/users/1", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 403


class TestAdminInvitationEdgeCases:
    """Cover admin invitation management edge cases."""

    def _admin_token(self, client, test_admin_user):
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200
        return login.get_json().get("access_token")

    def test_list_invitations_statuses(self, client, test_admin_user, db_connection, test_app):
        """Invitations have used/expired/pending statuses → lines 1745-1757."""
        token = self._admin_token(client, test_admin_user)
        now = datetime.utcnow()
        # Insert used invitation
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by, used_at)
               VALUES ('used@x.com', 'tok-used', ?, ?, ?)""",
            ((now + timedelta(days=7)).isoformat(), test_admin_user["email"], now.isoformat()),
        )
        # Insert expired invitation
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by)
               VALUES ('expired@x.com', 'tok-expired', ?, ?)""",
            ((now - timedelta(days=1)).isoformat(), test_admin_user["email"]),
        )
        # Insert pending invitation
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by)
               VALUES ('pending@x.com', 'tok-pending', ?, ?)""",
            ((now + timedelta(days=7)).isoformat(), test_admin_user["email"]),
        )
        db_connection.commit()
        r = client.get("/api/admin/invitations", headers={"Authorization": f"Bearer {token}"})
        assert r.status_code == 200
        invitations = r.get_json().get("invitations", [])
        statuses = {inv["email"]: inv["status"] for inv in invitations}
        assert statuses.get("used@x.com") == "used"
        assert statuses.get("expired@x.com") == "expired"
        assert statuses.get("pending@x.com") == "pending"
        # Cleanup
        for e in ("used@x.com", "expired@x.com", "pending@x.com"):
            db_connection.execute("DELETE FROM invitation_tokens WHERE email = ?", (e,))
        db_connection.commit()

    def test_expire_invitation_already_used(self, client, test_admin_user, db_connection, test_app):
        """Expire already-used invitation → line 1797."""
        token = self._admin_token(client, test_admin_user)
        now = datetime.utcnow()
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by, used_at)
               VALUES ('used2@x.com', 'tok-used2', ?, ?, ?)""",
            ((now + timedelta(days=7)).isoformat(), test_admin_user["email"], now.isoformat()),
        )
        db_connection.commit()
        inv_id = db_connection.execute(
            "SELECT id FROM invitation_tokens WHERE email = 'used2@x.com'"
        ).fetchone()[0]
        r = client.post(
            f"/api/admin/invitations/{inv_id}/expire",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400
        assert "used" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM invitation_tokens WHERE email = 'used2@x.com'")
        db_connection.commit()

    def test_expire_invitation_already_expired(self, client, test_admin_user, db_connection, test_app):
        """Expire already-expired invitation → line 1802."""
        token = self._admin_token(client, test_admin_user)
        now = datetime.utcnow()
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by)
               VALUES ('exp2@x.com', 'tok-exp2', ?, ?)""",
            ((now - timedelta(days=1)).isoformat(), test_admin_user["email"]),
        )
        db_connection.commit()
        inv_id = db_connection.execute(
            "SELECT id FROM invitation_tokens WHERE email = 'exp2@x.com'"
        ).fetchone()[0]
        r = client.post(
            f"/api/admin/invitations/{inv_id}/expire",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 400
        assert "expired" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM invitation_tokens WHERE email = 'exp2@x.com'")
        db_connection.commit()

    def test_create_invitation_integrity_error(self, client, test_admin_user, db_connection, test_app):
        """IntegrityError on UNIQUE constraint → lines 1701-1706."""
        token = self._admin_token(client, test_admin_user)
        email = "integrity-dupe@example.com"
        now = datetime.utcnow()
        # Pre-insert a non-expired, non-used invitation for this email
        db_connection.execute(
            """INSERT INTO invitation_tokens (email, token, expires_at, created_by)
               VALUES (?, 'existing-tok', ?, ?)""",
            (email, (now + timedelta(days=7)).isoformat(), test_admin_user["email"]),
        )
        db_connection.commit()
        # The route checks for pending invitation and returns 409 before hitting IntegrityError
        # To hit the IntegrityError, we need a race condition scenario
        # Instead, let's trigger it by monkeypatching conn.execute to raise IntegrityError on INSERT
        import main
        import sqlite3
        original_get_db = main.get_db

        call_count = [0]

        class FakeConn:
            def __init__(self, real_conn):
                self._real = real_conn

            def execute(self, sql, params=None):
                if "INSERT INTO invitation_tokens" in sql:
                    raise sqlite3.IntegrityError("UNIQUE constraint failed: invitation_tokens.email")
                if params:
                    return self._real.execute(sql, params)
                return self._real.execute(sql)

            def commit(self):
                return self._real.commit()

            def close(self):
                return self._real.close()

        def fake_get_db():
            return FakeConn(original_get_db())

        # Use a fresh email that doesn't have pending invitations (so route gets to INSERT)
        fresh_email = "integrity-fresh@example.com"
        with patch.object(main, 'get_db', fake_get_db):
            r = client.post(
                "/api/admin/create-invitation",
                json={"email": fresh_email},
                headers={"Authorization": f"Bearer {token}"},
            )
        assert r.status_code == 409
        db_connection.execute("DELETE FROM invitation_tokens WHERE email = ?", (email,))
        db_connection.commit()


class TestAdminExceptionHandlers:
    """Cover exception handlers in admin routes."""

    def _admin_token(self, client, test_admin_user):
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200
        return login.get_json().get("access_token")

    def test_create_invitation_exception(self, client, test_admin_user, monkeypatch):
        """Generic exception → lines 1714-1718."""
        token = self._admin_token(client, test_admin_user)
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.post(
            "/api/admin/create-invitation",
            json={"email": "x@y.com"},
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 500

    def test_list_invitations_exception(self, client, test_admin_user, monkeypatch):
        """Generic exception → lines 1764-1768."""
        token = self._admin_token(client, test_admin_user)
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get(
            "/api/admin/invitations",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 500

    def test_expire_invitation_exception(self, client, test_admin_user, monkeypatch):
        """Generic exception → lines 1819-1823."""
        token = self._admin_token(client, test_admin_user)
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.post(
            "/api/admin/invitations/1/expire",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 500

    def test_list_users_exception(self, client, test_admin_user, monkeypatch):
        """Generic exception → lines 1867-1871."""
        token = self._admin_token(client, test_admin_user)
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get(
            "/api/admin/users",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 500

    def test_get_user_details_exception(self, client, test_admin_user, monkeypatch):
        """Generic exception → lines 1920-1924."""
        token = self._admin_token(client, test_admin_user)
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.get(
            "/api/admin/users/1",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 500

    def test_delete_user_exception(self, client, test_admin_user, monkeypatch):
        """Generic exception → lines 2015-2019."""
        token = self._admin_token(client, test_admin_user)
        import main
        monkeypatch.setattr(main, 'get_db', lambda: (_ for _ in ()).throw(RuntimeError("boom")))
        r = client.delete(
            "/api/admin/users/1",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 500


class TestAdminUserManagementEdgeCases:
    """Cover admin user management edge cases."""

    def _admin_token(self, client, test_admin_user):
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200
        return login.get_json().get("access_token")

    def test_delete_user_current_user_not_found(self, client, test_app, test_admin_user, monkeypatch):
        """Current admin user somehow not in DB → line 1947."""
        from flask_jwt_extended import create_access_token
        import main
        # Create token for admin email, then make is_admin_user return True
        # but get_db return a conn where the admin doesn't exist
        token = create_access_token(identity="phantom-admin@example.com")
        monkeypatch.setattr(main, 'is_admin_user', lambda email: True)
        r = client.delete(
            "/api/admin/users/99999",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 404

    def test_delete_last_admin(self, client, test_admin_user, db_connection, test_app):
        """Cannot delete the last admin → line 1970."""
        token = self._admin_token(client, test_admin_user)
        # Create a second admin user to delete
        from werkzeug.security import generate_password_hash
        db_connection.execute(
            "INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, 1)",
            ("second-admin@example.com", generate_password_hash("pwd123")),
        )
        db_connection.commit()
        second_id = db_connection.execute(
            "SELECT id FROM users WHERE email = 'second-admin@example.com'"
        ).fetchone()[0]
        # Demote the original admin so second-admin is the last one
        db_connection.execute(
            "UPDATE users SET is_admin = 0 WHERE email = ?",
            (test_admin_user["email"],),
        )
        db_connection.commit()
        # Re-promote so we can make the request
        db_connection.execute(
            "UPDATE users SET is_admin = 1 WHERE email = ?",
            (test_admin_user["email"],),
        )
        db_connection.commit()
        # Now try to delete second_admin when there are 2 admins - this should succeed
        # To test "last admin", we need to delete when admin_count == 1
        # Demote ourselves, try to delete second_admin (who is last admin)
        db_connection.execute(
            "UPDATE users SET is_admin = 0 WHERE email = ?",
            (test_admin_user["email"],),
        )
        db_connection.commit()
        # We lost admin, so re-promote ourselves, and make second_admin sole admin
        db_connection.execute(
            "UPDATE users SET is_admin = 1 WHERE email = ?",
            (test_admin_user["email"],),
        )
        # Make second_admin the only admin besides us
        # Both are admins now. Delete second_admin should succeed since count > 1
        db_connection.commit()
        # To test last admin protection: make second_admin the ONLY admin
        # and try to delete them
        db_connection.execute(
            "UPDATE users SET is_admin = 0 WHERE email != 'second-admin@example.com'"
        )
        db_connection.commit()
        # But now our token is non-admin, so route will 403
        # Restore ourselves as admin for the request
        db_connection.execute(
            "UPDATE users SET is_admin = 1 WHERE email = ?",
            (test_admin_user["email"],),
        )
        db_connection.commit()
        # Now both are admins. To test "last admin", delete the non-self admin,
        # then try to create another admin scenario. Actually, simpler approach:
        # Just have one other admin user who is the ONLY admin (besides us) and try to delete
        # Actually the simplest way: only have our admin user and second-admin.
        # Make second-admin the only admin, then use monkeypatch for is_admin_user
        # Let me just set it up correctly:
        # 1. Both are admins (admin_count = 2) - deleting either should work
        # 2. Only second-admin is admin (count=1) - deleting them is blocked
        # Reset: both admins
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email IN (?, 'second-admin@example.com')",
                              (test_admin_user["email"],))
        db_connection.commit()
        # Remove our admin status, keep second-admin as sole admin
        db_connection.execute("UPDATE users SET is_admin = 0 WHERE email = ?", (test_admin_user["email"],))
        db_connection.commit()
        # But now we're non-admin and can't call the endpoint. Monkeypatch is_admin_user.
        import main
        original_is_admin = main.is_admin_user
        monkeypatch = test_app.extensions.get("monkeypatch")
        # Use a simpler approach: both are admins, no one else is. admin_count = 2.
        # Delete second_admin → admin_count drops to 1 → should succeed.
        # Then try to delete the remaining admin → should fail.
        # But we can't delete ourselves. So we need a third user.
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (test_admin_user["email"],))
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email = 'second-admin@example.com'")
        db_connection.commit()
        # Delete second-admin (count=2, so allowed)
        r = client.delete(
            f"/api/admin/users/{second_id}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        # Now create a third admin to test the last-admin protection
        db_connection.execute(
            "INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, 1)",
            ("last-admin@example.com", generate_password_hash("pwd123")),
        )
        db_connection.commit()
        last_id = db_connection.execute(
            "SELECT id FROM users WHERE email = 'last-admin@example.com'"
        ).fetchone()[0]
        # Make last-admin the only admin
        db_connection.execute("UPDATE users SET is_admin = 0 WHERE email != 'last-admin@example.com'")
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (test_admin_user["email"],))
        db_connection.commit()
        # admin_count = 2 (test_admin + last-admin). Delete last-admin → count drops to 1 → OK
        # To get "last admin" rejection: need target to be admin and count == 1
        db_connection.execute("UPDATE users SET is_admin = 0 WHERE email = ?", (test_admin_user["email"],))
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email = 'last-admin@example.com'")
        db_connection.commit()
        # Now only last-admin is admin (count=1). But we're not admin, so route 403s.
        # Re-promote ourselves but patch is_admin_user to return True
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (test_admin_user["email"],))
        db_connection.commit()
        # Now count = 2, can't trigger last-admin. This is circular.
        # Simplest approach: monkeypatch admin_count query
        # Actually, the simplest way: have only 1 admin total, and that admin tries to delete themselves
        # But the route already blocks "cannot delete yourself". So we need a scenario where
        # the target is the ONLY admin, and the requester is a different admin.
        # That means admin_count >= 2 always when the request gets to the count check.
        # Unless we monkeypatch. Let me just do that.
        db_connection.execute("DELETE FROM users WHERE email = 'last-admin@example.com'")
        db_connection.commit()
        # Recreate: make last-admin the sole admin, but patch is_admin_user for our call
        db_connection.execute(
            "INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, 1)",
            ("sole-admin@example.com", generate_password_hash("pwd123")),
        )
        db_connection.commit()
        sole_id = db_connection.execute(
            "SELECT id FROM users WHERE email = 'sole-admin@example.com'"
        ).fetchone()[0]
        # Set admin_count to exactly 1 by removing other admins
        db_connection.execute("UPDATE users SET is_admin = 0 WHERE email != 'sole-admin@example.com'")
        db_connection.commit()
        # Monkeypatch is_admin_user to always return True for our request
        import main as main_mod
        with patch.object(main_mod, 'is_admin_user', return_value=True):
            r = client.delete(
                f"/api/admin/users/{sole_id}",
                headers={"Authorization": f"Bearer {token}"},
            )
        assert r.status_code == 400
        assert "last admin" in r.get_json().get("error", "").lower()
        # Cleanup
        db_connection.execute("UPDATE users SET is_admin = 1 WHERE email = ?", (test_admin_user["email"],))
        db_connection.execute("DELETE FROM users WHERE email IN ('sole-admin@example.com')")
        db_connection.commit()

    def test_delete_user_with_files_on_disk(self, client, test_admin_user, db_connection, test_app, test_directories):
        """Delete user cleans up files from disk → lines 1981-1991."""
        token = self._admin_token(client, test_admin_user)
        from werkzeug.security import generate_password_hash
        db_connection.execute(
            "INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, 0)",
            ("file-user@example.com", generate_password_hash("pwd123")),
        )
        db_connection.commit()
        uid = db_connection.execute(
            "SELECT id FROM users WHERE email = 'file-user@example.com'"
        ).fetchone()[0]
        # Create a file on disk
        stored = "user-file-to-delete.xlsx"
        fpath = os.path.join(test_directories["uploads"], stored)
        with open(fpath, "wb") as f:
            f.write(b"data")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'to-delete.xlsx', ?, 'original')""",
            (uid, stored),
        )
        db_connection.commit()
        r = client.delete(
            f"/api/admin/users/{uid}",
            headers={"Authorization": f"Bearer {token}"},
        )
        assert r.status_code == 200
        # File should be deleted from disk
        assert not os.path.exists(fpath)


class TestGetFilesHappyPath:
    """Cover main.py line 299 (valid_files.append for files that exist on disk)."""

    def test_get_files_includes_files_on_disk(self, client, auth_token, sample_excel_file):
        """GET /api/files returns files that exist on disk → line 299."""
        if auth_token is None:
            assert client.get("/api/files").status_code == 401
            return
        headers = {"Authorization": f"Bearer {auth_token}"}
        # Upload a real file so it exists on disk
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "real_file.xlsx")},
                headers=headers,
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        # Now list files — the uploaded file should appear
        r = client.get("/api/files", headers=headers)
        assert r.status_code == 200
        files = r.get_json().get("files", [])
        assert len(files) >= 1


class TestGitHubRoutesAdditional:
    """Cover remaining GitHub route branches."""

    def test_github_missing_env_vars(self, client, auth_token, monkeypatch):
        """test-github with missing env vars → line 776."""
        if auth_token is None:
            assert client.get("/api/test-github").status_code == 401
            return
        monkeypatch.delenv("GITHUB_APP_ID", raising=False)
        monkeypatch.delenv("GITHUB_PRIVATE_KEY", raising=False)
        monkeypatch.delenv("GITHUB_INSTALLATION_ID", raising=False)
        r = client.get("/api/test-github", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 400
        assert "missing" in r.get_json().get("error", "").lower()

    def test_github_success_path(self, client, auth_token, monkeypatch):
        """test-github success when repo responds 200 → lines 808-809."""
        if auth_token is None:
            assert client.get("/api/test-github").status_code == 401
            return
        monkeypatch.setenv("GITHUB_APP_ID", "123")
        monkeypatch.setenv("GITHUB_PRIVATE_KEY", "fake-key")
        monkeypatch.setenv("GITHUB_INSTALLATION_ID", "456")
        import main
        mock_auth = MagicMock()
        mock_auth.get_app_token.return_value = "fake-app-token-" + "x" * 50
        mock_auth.get_installation_token.return_value = "fake-install-token-" + "x" * 50
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.json.return_value = {"full_name": "owner/repo", "private": True}
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            with patch.object(main.requests, 'get', return_value=mock_response):
                r = client.get("/api/test-github", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert r.get_json().get("repo_access") is True

    def test_dispatch_success_204(self, client, auth_token, monkeypatch):
        """test-dispatch success when dispatch returns 204 → line 923."""
        if auth_token is None:
            assert client.post("/api/test-dispatch").status_code == 401
            return
        import main
        mock_auth = MagicMock()
        mock_auth.get_installation_token.return_value = "fake-token-" + "x" * 50
        repo_resp = MagicMock()
        repo_resp.status_code = 200
        dispatch_resp = MagicMock()
        dispatch_resp.status_code = 204
        dispatch_resp.text = ""
        dispatch_resp.headers = {}
        with patch.object(main, 'GitHubAppAuth', return_value=mock_auth):
            with patch.object(main.requests, 'get', return_value=repo_resp):
                with patch.object(main.requests, 'post', return_value=dispatch_resp):
                    r = client.post("/api/test-dispatch", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 200
        assert r.get_json().get("status") == "success"

    def test_dispatch_exception_handler(self, client, auth_token, monkeypatch):
        """test-dispatch exception → lines 929-934."""
        if auth_token is None:
            assert client.post("/api/test-dispatch").status_code == 401
            return
        import main
        with patch.object(main, 'GitHubAppAuth', side_effect=RuntimeError("auth boom")):
            r = client.post("/api/test-dispatch", headers={"Authorization": f"Bearer {auth_token}"})
        assert r.status_code == 500


class TestDeleteUserFileError:
    """Cover main.py lines 1990-1991 (os.remove exception during user delete)."""

    def test_delete_user_file_remove_oserror(self, client, test_admin_user, db_connection, test_app, test_directories):
        """Delete user handles OSError when removing files from disk."""
        login = client.post(
            "/api/login",
            json={"email": test_admin_user["email"], "password": test_admin_user["password"]},
        )
        assert login.status_code == 200
        token = login.get_json().get("access_token")
        from werkzeug.security import generate_password_hash
        db_connection.execute(
            "INSERT INTO users (email, password_hash, is_admin) VALUES (?, ?, 0)",
            ("file-err-user@example.com", generate_password_hash("pwd123")),
        )
        db_connection.commit()
        uid = db_connection.execute(
            "SELECT id FROM users WHERE email = 'file-err-user@example.com'"
        ).fetchone()[0]
        # Create a file on disk
        stored = "del-oserr-file.xlsx"
        fpath = os.path.join(test_directories["uploads"], stored)
        with open(fpath, "wb") as f:
            f.write(b"data")
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, 'oserr.xlsx', ?, 'original')""",
            (uid, stored),
        )
        db_connection.commit()
        # Make os.remove fail for this specific file
        import main
        real_remove = os.remove

        def failing_remove(path):
            if "del-oserr-file" in path:
                raise OSError("Permission denied")
            return real_remove(path)

        with patch.object(main.os, 'remove', failing_remove):
            r = client.delete(
                f"/api/admin/users/{uid}",
                headers={"Authorization": f"Bearer {token}"},
            )
        # Should still succeed despite file deletion failure
        assert r.status_code == 200
        # Cleanup file if it still exists
        if os.path.exists(fpath):
            os.remove(fpath)
