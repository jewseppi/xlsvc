"""
Comprehensive tests for process_file endpoint and related functions
These tests ensure full coverage before refactoring
"""
import pytest
import os
import json
from unittest.mock import patch
from main import (
    generate_libreoffice_macro,
    generate_instructions,
    column_to_index,
    is_empty_or_zero
)


class TestProcessFileEndpoint:
    """Tests for the /api/process/<file_id> endpoint"""
    
    def test_process_file_without_filter_rules(self, client, auth_token, test_user, db_connection, sample_excel_file):
        """Test that process_file requires filter_rules"""
        if auth_token is None:
            r = client.post(f'/api/process/1', json={})
            assert r.status_code == 401
            return
        
        # Upload a file first
        with open(sample_excel_file, 'rb') as f:
            upload_response = client.post(
                '/api/upload',
                data={'file': (f, 'test_file.xlsx')},
                headers={'Authorization': f'Bearer {auth_token}'},
                content_type='multipart/form-data'
            )
        
        # Upload returns 201 for new files, 200 for duplicates
        assert upload_response.status_code in [200, 201], f"Upload failed: {upload_response.get_json()}"
        file_id = upload_response.get_json()['file_id']
        
        # Try to process without filter_rules
        response = client.post(
            f'/api/process/{file_id}',
            json={},
            headers={'Authorization': f'Bearer {auth_token}'}
        )
        
        assert response.status_code == 400
        assert 'filter_rules' in response.get_json()['error'].lower()
    
    def test_process_file_file_not_on_disk(self, client, auth_token, test_user, db_connection, test_app):
        """process_file when file record exists but file missing on disk returns 404."""
        if auth_token is None:
            r = client.post('/api/process/1', json={"filter_rules": [{"column": "F", "value": "0"}]})
            assert r.status_code == 401
            return
        import uuid
        stored = f"{uuid.uuid4()}.xlsx"
        db_connection.execute(
            """INSERT INTO files (user_id, original_filename, stored_filename, file_type)
               VALUES (?, ?, ?, ?)""",
            (test_user["id"], "ghost.xlsx", stored, "original"),
        )
        db_connection.commit()
        file_id = db_connection.execute("SELECT last_insert_rowid()").fetchone()[0]
        r = client.post(
            f"/api/process/{file_id}",
            json={"filter_rules": [{"column": "F", "value": "0"}]},
            headers={"Authorization": f"Bearer {auth_token}"},
        )
        assert r.status_code == 404
        assert "disk" in r.get_json().get("error", "").lower()
        db_connection.execute("DELETE FROM files WHERE id = ?", (file_id,))
        db_connection.commit()

    def test_process_file_with_empty_filter_rules(self, client, auth_token, sample_excel_file):
        """Test that process_file requires non-empty filter_rules"""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': []})
            assert r.status_code == 401
            return
        
        # Upload a file first
        with open(sample_excel_file, 'rb') as f:
            upload_response = client.post(
                '/api/upload',
                data={'file': (f, 'test_file.xlsx')},
                headers={'Authorization': f'Bearer {auth_token}'},
                content_type='multipart/form-data'
            )
        
        # Upload returns 201 for new files, 200 for duplicates
        assert upload_response.status_code in [200, 201], f"Upload failed: {upload_response.get_json()}"
        file_id = upload_response.get_json()['file_id']
        
        # Try to process with empty filter_rules
        response = client.post(
            f'/api/process/{file_id}',
            json={'filter_rules': []},
            headers={'Authorization': f'Bearer {auth_token}'}
        )
        
        assert response.status_code == 400
    
    def test_process_file_identifies_rows_to_delete(self, client, auth_token, comprehensive_test_excel, test_user, db_connection):
        """Test that process_file correctly identifies rows to delete"""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'F', 'value': '0'}]})
            assert r.status_code == 401
            return
        
        # Upload the comprehensive test file
        with open(comprehensive_test_excel, 'rb') as f:
            upload_response = client.post(
                '/api/upload',
                data={'file': (f, 'comprehensive_test.xlsx')},
                headers={'Authorization': f'Bearer {auth_token}'},
                content_type='multipart/form-data'
            )
        
        # Upload returns 201 for new files, 200 for duplicates
        assert upload_response.status_code in [200, 201], f"Upload failed: {upload_response.get_json()}"
        file_id = upload_response.get_json()['file_id']
        
        # Process with default filter rules (F, G, H, I = 0)
        filter_rules = [
            {'column': 'F', 'value': '0'},
            {'column': 'G', 'value': '0'},
            {'column': 'H', 'value': '0'},
            {'column': 'I', 'value': '0'}
        ]
        
        response = client.post(
            f'/api/process/{file_id}',
            json={'filter_rules': filter_rules},
            headers={'Authorization': f'Bearer {auth_token}'}
        )
        
        assert response.status_code == 200
        data = response.get_json()
        
        # Should find rows to delete
        assert 'total_rows_to_delete' in data
        assert data['total_rows_to_delete'] > 0
        assert 'sheets_affected' in data
        assert 'TestSheet1' in data['sheets_affected']
        assert 'downloads' in data
        assert 'macro' in data['downloads']
        assert 'instructions' in data['downloads']
    
    def test_process_file_no_rows_to_delete(self, client, auth_token, sample_excel_file):
        """Test process_file when no rows match filter criteria.
        
        All rules use is_empty_or_zero, so we filter on Column A which
        contains non-empty text in every data row -> no matches.
        """
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'A', 'value': '0'}]})
            assert r.status_code == 401
            return

        # Upload a file
        with open(sample_excel_file, 'rb') as f:
            upload_response = client.post(
                '/api/upload',
                data={'file': (f, 'test_file.xlsx')},
                headers={'Authorization': f'Bearer {auth_token}'},
                content_type='multipart/form-data'
            )

        # Upload returns 201 for new files, 200 for duplicates
        assert upload_response.status_code in [200, 201], f"Upload failed: {upload_response.get_json()}"
        file_id = upload_response.get_json()['file_id']
        
        # Process with filter on Column A (has text in every row -> never empty/zero)
        filter_rules = [
            {'column': 'A', 'value': '0'}
        ]

        response = client.post(
            f'/api/process/{file_id}',
            json={'filter_rules': filter_rules},
            headers={'Authorization': f'Bearer {auth_token}'}
        )

        assert response.status_code == 200
        data = response.get_json()
        assert data['total_rows_to_delete'] == 0
        assert 'No rows found for deletion' in data['message']
    
    def test_process_file_generates_macro_and_instructions(self, client, auth_token, comprehensive_test_excel, test_directories, db_connection):
        """Test that process_file generates macro and instructions files"""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'F', 'value': '0'}]})
            assert r.status_code == 401
            return
        
        # Upload file
        with open(comprehensive_test_excel, 'rb') as f:
            upload_response = client.post(
                '/api/upload',
                data={'file': (f, 'comprehensive_test.xlsx')},
                headers={'Authorization': f'Bearer {auth_token}'},
                content_type='multipart/form-data'
            )
        
        # Upload returns 201 for new files, 200 for duplicates
        assert upload_response.status_code in [200, 201], f"Upload failed: {upload_response.get_json()}"
        file_id = upload_response.get_json()['file_id']
        
        # Process file
        filter_rules = [
            {'column': 'F', 'value': '0'},
            {'column': 'G', 'value': '0'},
            {'column': 'H', 'value': '0'},
            {'column': 'I', 'value': '0'}
        ]
        
        response = client.post(
            f'/api/process/{file_id}',
            json={'filter_rules': filter_rules},
            headers={'Authorization': f'Bearer {auth_token}'}
        )
        
        assert response.status_code == 200
        data = response.get_json()
        
        # Check that macro and instructions files were created
        # The response contains file_ids, we need to check the database for stored filenames
        macro_file_id = data['downloads']['macro']['file_id']
        instructions_file_id = data['downloads']['instructions']['file_id']
        
        # Use db_connection fixture instead of get_db
        macro_info = db_connection.execute('SELECT stored_filename FROM files WHERE id = ?', (macro_file_id,)).fetchone()
        instructions_info = db_connection.execute('SELECT stored_filename FROM files WHERE id = ?', (instructions_file_id,)).fetchone()
        
        assert macro_info is not None
        assert instructions_info is not None
        
        macro_path = os.path.join(test_directories['macros'], macro_info['stored_filename'])
        instructions_path = os.path.join(test_directories['macros'], instructions_info['stored_filename'])
        
        assert os.path.exists(macro_path), f"Macro file not found: {macro_path}"
        assert os.path.exists(instructions_path), f"Instructions file not found: {instructions_path}"
        
        # Verify macro content
        with open(macro_path, 'r') as f:
            macro_content = f.read()
            assert 'Sub DeleteEmptyRows()' in macro_content
            assert 'TestSheet1' in macro_content or 'Sheet1' in macro_content
        
        # Verify instructions content
        with open(instructions_path, 'r') as f:
            instructions_content = f.read()
            assert 'EXCEL FILE CLEANUP INSTRUCTIONS' in instructions_content
            assert 'comprehensive_test.xlsx' in instructions_content

    def test_process_file_analysis_exception_returns_error(self, client, auth_token, sample_excel_file):
        """When load_workbook/analysis raises, process_file returns 500 with error message."""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'F', 'value': '0'}]})
            assert r.status_code == 401
            return
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        with patch("main.load_workbook") as mock_load:
            mock_load.side_effect = RuntimeError("Simulated analysis error")
            r = client.post(
                f"/api/process/{file_id}",
                json={"filter_rules": [{"column": "F", "value": "0"}]},
                headers={"Authorization": f"Bearer {auth_token}"},
            )
        assert r.status_code in [400, 500]
        data = r.get_json()
        assert "error" in data
        assert "analysis" in data.get("error", "").lower() or "failed" in data.get("error", "").lower()

    def test_process_file_outer_exception(self, client, auth_token, sample_excel_file):
        """When get_db raises, outer except returns 500."""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'F', 'value': '0'}]})
            assert r.status_code == 401
            return
        with open(sample_excel_file, "rb") as f:
            up = client.post(
                "/api/upload",
                data={"file": (f, "test_file.xlsx")},
                headers={"Authorization": f"Bearer {auth_token}"},
                content_type="multipart/form-data",
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()["file_id"]
        with patch("main.get_db", side_effect=RuntimeError("DB unavailable")):
            r = client.post(
                f"/api/process/{file_id}",
                json={"filter_rules": [{"column": "F", "value": "0"}]},
                headers={"Authorization": f"Bearer {auth_token}"},
            )
        assert r.status_code == 500
        assert "error" in r.get_json()

    def test_process_file_skips_empty_column_a_rows(self, client, auth_token, test_user, test_directories, db_connection):
        """Test that rows with empty Column A are skipped and NOT marked for deletion (UNO parity)"""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'F', 'value': '0'}]})
            assert r.status_code == 401
            return

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'TestSheet'
        # Header row (Column A has content)
        ws.append(['Name', 'B', 'C', 'D', 'E', 'F'])
        # Row 2: Column A has value, F=0 → SHOULD be marked for deletion
        ws.append(['DataRow', '', '', '', '', 0])
        # Row 3: Column A is EMPTY, F=0 → should be SKIPPED (not marked)
        ws.append([None, '', '', '', '', 0])
        # Row 4: Column A is blank string, F=0 → should be SKIPPED
        ws.append(['', '', '', '', '', 0])
        # Row 5: Column A has value, F=5 → should NOT match (F is not empty/zero)
        ws.append(['AnotherRow', '', '', '', '', 5])

        file_path = os.path.join(test_directories['uploads'], 'col_a_test.xlsx')
        wb.save(file_path)

        headers = {'Authorization': f'Bearer {auth_token}'}
        with open(file_path, 'rb') as f:
            up = client.post(
                '/api/upload',
                data={'file': (f, 'col_a_test.xlsx')},
                headers=headers,
                content_type='multipart/form-data',
            )
        assert up.status_code in [200, 201]
        file_id = up.get_json()['file_id']

        response = client.post(
            f'/api/process/{file_id}',
            json={'filter_rules': [{'column': 'F', 'value': '0'}]},
            headers=headers,
        )
        assert response.status_code == 200
        data = response.get_json()

        # Only row 2 should be marked (Column A has content AND F=0)
        # Rows 3 & 4 are skipped (empty Column A)
        # Row 5 is not marked (F=5, not empty/zero)
        # Row 1 (header) has "Name" in A and "F" in F — "F" is not empty/zero, so not marked
        assert data['total_rows_to_delete'] == 1

        if os.path.exists(file_path):
            os.remove(file_path)

    def test_process_file_generates_deletion_report(self, client, auth_token, comprehensive_test_excel, test_directories, db_connection):
        """Test that process_file generates a deletion report and returns report_file_id (UNO parity)"""
        if auth_token is None:
            r = client.post('/api/process/1', json={'filter_rules': [{'column': 'F', 'value': '0'}]})
            assert r.status_code == 401
            return

        with open(comprehensive_test_excel, 'rb') as f:
            upload_response = client.post(
                '/api/upload',
                data={'file': (f, 'comprehensive_test.xlsx')},
                headers={'Authorization': f'Bearer {auth_token}'},
                content_type='multipart/form-data'
            )
        assert upload_response.status_code in [200, 201]
        file_id = upload_response.get_json()['file_id']

        filter_rules = [
            {'column': 'F', 'value': '0'},
            {'column': 'G', 'value': '0'},
            {'column': 'H', 'value': '0'},
            {'column': 'I', 'value': '0'}
        ]

        response = client.post(
            f'/api/process/{file_id}',
            json={'filter_rules': filter_rules},
            headers={'Authorization': f'Bearer {auth_token}'}
        )
        assert response.status_code == 200
        data = response.get_json()

        # Must have rows to delete for report to be generated
        assert data['total_rows_to_delete'] > 0

        # Verify report_file_id is in response
        assert 'report_file_id' in data
        assert data['report_file_id'] is not None

        # Verify report is in downloads
        assert 'report' in data['downloads']
        assert 'file_id' in data['downloads']['report']
        assert data['downloads']['report']['file_id'] == data['report_file_id']

        # Verify DB record exists with correct file_type
        report_record = db_connection.execute(
            'SELECT file_type, stored_filename FROM files WHERE id = ?',
            (data['report_file_id'],)
        ).fetchone()
        assert report_record is not None
        assert report_record['file_type'] == 'report'

        # Verify file exists on disk
        report_path = os.path.join(test_directories['reports'], report_record['stored_filename'])
        assert os.path.exists(report_path), f"Report file not found: {report_path}"


class TestGenerateLibreOfficeMacro:
    """Tests for generate_libreoffice_macro function"""
    
    def test_generate_macro_basic(self):
        """Test basic macro generation"""
        rows_to_delete_by_sheet = {
            'Sheet1': [2, 3, 5]
        }
        filter_rules = [
            {'column': 'F', 'value': '0'},
            {'column': 'G', 'value': '0'}
        ]
        
        macro = generate_libreoffice_macro(
            'test.xlsx',
            rows_to_delete_by_sheet,
            filter_rules
        )
        
        assert 'Sub DeleteEmptyRows()' in macro
        assert 'test.xlsx' in macro
        assert 'Sheet1' in macro
        assert 'removeByIndex' in macro
    
    def test_generate_macro_multiple_sheets(self):
        """Test macro generation with multiple sheets"""
        rows_to_delete_by_sheet = {
            'Sheet1': [2, 3],
            'Sheet2': [5, 6, 7]
        }
        
        macro = generate_libreoffice_macro(
            'test.xlsx',
            rows_to_delete_by_sheet,
            []
        )
        
        assert 'Sheet1' in macro
        assert 'Sheet2' in macro
        assert macro.count('hasByName') == 2  # One for each sheet
    
    def test_generate_macro_consecutive_rows(self):
        """Test that macro groups consecutive rows efficiently"""
        rows_to_delete_by_sheet = {
            'Sheet1': [5, 4, 3, 2]  # Consecutive rows
        }
        
        macro = generate_libreoffice_macro(
            'test.xlsx',
            rows_to_delete_by_sheet,
            []
        )
        
        # Should group consecutive rows into one removeByIndex call
        assert 'removeByIndex(1, 4)' in macro or 'removeByIndex(2, 4)' in macro
    
    def test_generate_macro_non_consecutive_rows(self):
        """Test macro with non-consecutive rows"""
        rows_to_delete_by_sheet = {
            'Sheet1': [2, 5, 10]  # Non-consecutive
        }
        
        macro = generate_libreoffice_macro(
            'test.xlsx',
            rows_to_delete_by_sheet,
            []
        )
        
        # Should have multiple removeByIndex calls
        assert macro.count('removeByIndex') >= 3


class TestGenerateInstructions:
    """Tests for generate_instructions function"""
    
    def test_generate_instructions_basic(self):
        """Test basic instructions generation"""
        instructions = generate_instructions(
            'test.xlsx',
            5,
            ['Sheet1'],
            [{'column': 'F', 'value': '0'}]
        )
        
        assert 'EXCEL FILE CLEANUP INSTRUCTIONS' in instructions
        assert 'test.xlsx' in instructions
        assert '5 rows' in instructions or '5' in instructions
        assert 'Sheet1' in instructions
        assert 'Column F' in instructions
    
    def test_generate_instructions_multiple_sheets(self):
        """Test instructions with multiple sheets"""
        instructions = generate_instructions(
            'test.xlsx',
            10,
            ['Sheet1', 'Sheet2'],
            [{'column': 'F', 'value': '0'}, {'column': 'G', 'value': '0'}]
        )
        
        assert 'Sheet1' in instructions
        assert 'Sheet2' in instructions
        assert 'Column F' in instructions
        assert 'Column G' in instructions
    
    def test_generate_instructions_all_rules_empty_or_zero(self):
        """All rules are described as empty/zero checks (parity with UNO)."""
        instructions = generate_instructions(
            'test.xlsx',
            3,
            ['Sheet1'],
            [{'column': 'A', 'value': 'DELETE'}, {'column': 'B', 'value': '0'}]
        )
        
        # Both rules should say "empty or zero" regardless of the value field
        assert instructions.count('is empty or zero') == 2
        assert "equals" not in instructions


class TestProcessingLogic:
    """Tests for the core processing logic"""
    
    def test_row_identification_with_zeros(self, comprehensive_test_excel):
        """Test that rows with zeros are correctly identified"""
        from openpyxl import load_workbook
        from main import is_empty_or_zero, column_to_index
        
        wb = load_workbook(comprehensive_test_excel, data_only=True)
        sheet = wb['TestSheet1']
        
        filter_rules = [
            {'column': 'F', 'value': '0'},
            {'column': 'G', 'value': '0'},
            {'column': 'H', 'value': '0'},
            {'column': 'I', 'value': '0'}
        ]
        
        rows_to_delete = []
        for row_num in range(2, sheet.max_row + 1):  # Skip header
            all_match = True
            for rule in filter_rules:
                col_index = column_to_index(rule['column'])
                cell_val = sheet.cell(row=row_num, column=col_index).value
                
                if rule['value'] == '0':
                    if not is_empty_or_zero(cell_val):
                        all_match = False
                        break
            
            if all_match:
                rows_to_delete.append(row_num)
        
        # Should identify rows 2, 3, 4, 6, 7, 8, 10 (rows with all zeros in F,G,H,I)
        # Row 5 has values, Row 9 has partial match
        assert len(rows_to_delete) > 0
        assert 2 in rows_to_delete  # All zeros
        assert 5 not in rows_to_delete  # Has values
        assert 9 not in rows_to_delete  # Partial match
    
    def test_row_identification_always_uses_empty_or_zero(self, comprehensive_test_excel):
        """Test that filter rules always check for empty/zero regardless of value field"""
        from openpyxl import load_workbook
        from main import column_to_index, is_empty_or_zero

        wb = load_workbook(comprehensive_test_excel, data_only=True)
        sheet = wb['TestSheet1']

        # Even with value='Row5_HasValues', the evaluation should check empty/zero
        filter_rules = [
            {'column': 'A', 'value': 'Row5_HasValues'}
        ]

        rows_to_delete = []
        for row_num in range(2, sheet.max_row + 1):
            # Column A pre-filter
            col_a_value = sheet.cell(row=row_num, column=1).value
            if col_a_value is None or str(col_a_value).strip() == '':
                continue

            all_match = True
            for rule in filter_rules:
                col_index = column_to_index(rule['column'])
                cell_val = sheet.cell(row=row_num, column=col_index).value

                if not is_empty_or_zero(cell_val):
                    all_match = False
                    break

            if all_match:
                rows_to_delete.append(row_num)

        # Column A has non-empty values for data rows, so NO rows should match
        # (because is_empty_or_zero returns False for non-empty strings)
        assert len(rows_to_delete) == 0


class TestColumnAPreFilter:
    """Tests for Column A pre-filter behavior (parity with UNO)."""

    def test_rows_with_empty_col_a_are_skipped(self, client, test_user, test_directories, db_connection):
        """Rows where Column A is empty are NOT evaluated for deletion."""
        from openpyxl import Workbook

        # Create a workbook where some rows have empty column A
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws['A1'] = 'Header'
        ws['F1'] = 'ColF'
        # Row 2: Col A has data, F is zero → should be deleted
        ws['A2'] = 'DataRow'
        ws['F2'] = 0
        # Row 3: Col A is empty, F is zero → should be SKIPPED (not deleted)
        ws['A3'] = None
        ws['F3'] = 0
        # Row 4: Col A has whitespace only, F is zero → should be SKIPPED
        ws['A4'] = '   '
        ws['F4'] = 0
        # Row 5: Col A = 0 (numeric zero), F is zero → Column A is "0" which is truthy
        ws['A5'] = 0
        ws['F5'] = 0

        file_path = os.path.join(test_directories['uploads'], 'col_a_test.xlsx')
        wb.save(file_path)

        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        with open(file_path, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'col_a_test.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'filter_rules': [{'column': 'F', 'value': '0'}]
        }, headers={'Authorization': f'Bearer {token}'})

        assert response.status_code == 200
        data = response.get_json()
        # Row 1 (Header) has non-zero F, Row 3 and 4 skipped (empty A),
        # Row 2 and 5 should be deleted (A has data, F is zero)
        assert data['total_rows_to_delete'] == 2

        # Cleanup
        if os.path.exists(file_path):
            os.remove(file_path)


class TestColumnsToRemoveValidation:
    """Tests for columns_to_remove parameter in process_file."""

    def test_invalid_columns_to_remove_returns_400(self, client, test_user, sample_excel_file):
        """Invalid columns_to_remove returns 400."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'filter_rules': [{'column': 'F', 'value': '0'}],
            'columns_to_remove': ['123']  # invalid
        }, headers={'Authorization': f'Bearer {token}'})

        assert response.status_code == 400
        assert 'columns_to_remove' in response.get_json()['error']

    def test_valid_columns_to_remove_accepted(self, client, test_user, sample_excel_file):
        """Valid columns_to_remove is accepted and processed."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'filter_rules': [{'column': 'F', 'value': '0'}],
            'columns_to_remove': ['B', 'c']  # will be normalized to uppercase
        }, headers={'Authorization': f'Bearer {token}'})

        # Should succeed (rows are evaluated)
        assert response.status_code == 200


class TestDeletionReportGeneration:
    """Tests for deletion report generation in manual path."""

    def test_process_file_generates_deletion_report(self, client, test_user, comprehensive_test_excel, test_directories, db_connection):
        """Manual process_file generates deletion report and returns report_file_id."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        with open(comprehensive_test_excel, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'comprehensive_test.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'filter_rules': [
                {'column': 'G', 'value': '0'},
                {'column': 'H', 'value': '0'},
                {'column': 'I', 'value': '0'},
                {'column': 'F', 'value': '0'}
            ]
        }, headers={'Authorization': f'Bearer {token}'})

        assert response.status_code == 200
        data = response.get_json()
        assert data['total_rows_to_delete'] > 0
        # Deletion report should be generated and report_file_id returned
        assert 'report_file_id' in data
        assert isinstance(data['report_file_id'], int)

        # Verify the report file exists in DB
        report_record = db_connection.execute(
            "SELECT * FROM files WHERE id = ?", (data['report_file_id'],)
        ).fetchone()
        assert report_record is not None
        assert report_record['file_type'] == 'report'
        assert 'DeletionReport' in report_record['original_filename']

        # Verify the report file exists on disk
        report_path = os.path.join(test_directories['reports'], report_record['stored_filename'])
        assert os.path.exists(report_path)


class TestInstructionsWithColumnsToRemove:
    """Tests for generate_instructions with columns_to_remove."""

    def test_instructions_include_column_removal_info(self):
        """Instructions mention columns to be removed when provided."""
        instructions = generate_instructions(
            'test.xlsx', 5, ['Sheet1'],
            [{'column': 'F', 'value': '0'}],
            columns_to_remove=['B', 'D']
        )
        assert 'Column B' in instructions
        assert 'Column D' in instructions
        assert 'right-to-left' in instructions

    def test_instructions_no_column_removal_when_empty(self):
        """Instructions don't mention column removal when list is empty."""
        instructions = generate_instructions(
            'test.xlsx', 5, ['Sheet1'],
            [{'column': 'F', 'value': '0'}],
            columns_to_remove=[]
        )
        assert 'Columns to be removed' not in instructions

    def test_instructions_dynamic_manual_columns(self):
        """Manual deletion section uses actual rule columns, not hardcoded F,G,H,I."""
        instructions = generate_instructions(
            'test.xlsx', 5, ['Sheet1'],
            [{'column': 'X', 'value': '0'}, {'column': 'Y', 'value': '0'}]
        )
        assert 'X, Y' in instructions
        # Should NOT contain the old hardcoded F, G, H, I reference
        assert 'columns F, G, H, and I' not in instructions


class TestProfileAwareProcessing:
    """Tests for profile_id resolution in process_file and process-automated routes."""

    def test_process_file_with_profile_id(self, client, test_user, sample_excel_file, db_connection):
        """Process file using a saved profile (profile_id)."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        # Create a profile
        db_connection.execute(
            """INSERT INTO filter_profiles
               (id, user_id, name, filter_rules_json, columns_to_remove, is_system_template)
               VALUES (100, ?, 'TestProfile', '[{"column":"F","value":"0"}]', '[]', 0)""",
            (test_user['id'],)
        )
        db_connection.commit()

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        # Process with profile_id (no filter_rules in request)
        response = client.post(f'/api/process/{file_id}', json={
            'profile_id': 100
        }, headers={'Authorization': f'Bearer {token}'})

        assert response.status_code == 200
        data = response.get_json()
        # Profile has filter for column F = 0, sample has zeros in F → should find rows
        assert data['total_rows_to_delete'] > 0

        # Cleanup
        db_connection.execute("DELETE FROM filter_profiles WHERE id = 100")
        db_connection.commit()

    def test_process_file_with_system_template(self, client, test_user, sample_excel_file, db_connection):
        """Process file using a system template profile."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        db_connection.execute(
            """INSERT INTO filter_profiles
               (id, user_id, name, filter_rules_json, columns_to_remove, is_system_template)
               VALUES (101, NULL, 'SystemTmpl', '[{"column":"F","value":"0"}]', '[]', 1)"""
        )
        db_connection.commit()

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'profile_id': 101
        }, headers={'Authorization': f'Bearer {token}'})
        assert response.status_code == 200

        # Cleanup
        db_connection.execute("DELETE FROM filter_profiles WHERE id = 101")
        db_connection.commit()

    def test_process_file_profile_not_found(self, client, test_user, sample_excel_file):
        """Profile not found returns 404."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'profile_id': 99999
        }, headers={'Authorization': f'Bearer {token}'})
        assert response.status_code == 404

    def test_process_file_profile_access_denied(self, client, test_user, sample_excel_file, db_connection):
        """Cannot use another user's profile."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        # Create profile for a different user (id=99999)
        db_connection.execute(
            """INSERT INTO filter_profiles
               (id, user_id, name, filter_rules_json, is_system_template)
               VALUES (102, 99999, 'OtherUser', '[{"column":"F","value":"0"}]', 0)"""
        )
        db_connection.commit()

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process/{file_id}', json={
            'profile_id': 102
        }, headers={'Authorization': f'Bearer {token}'})
        assert response.status_code == 403

        # Cleanup
        db_connection.execute("DELETE FROM filter_profiles WHERE id = 102")
        db_connection.commit()

    def test_inline_rules_take_precedence_over_profile_id(self, client, test_user, sample_excel_file, db_connection):
        """When both profile_id and filter_rules are provided, filter_rules take precedence."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        db_connection.execute(
            """INSERT INTO filter_profiles
               (id, user_id, name, filter_rules_json, is_system_template)
               VALUES (103, ?, 'Unused', '[{"column":"A","value":"0"}]', 0)""",
            (test_user['id'],)
        )
        db_connection.commit()

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        # Send both profile_id and filter_rules — filter_rules should be used
        response = client.post(f'/api/process/{file_id}', json={
            'profile_id': 103,
            'filter_rules': [{'column': 'F', 'value': '0'}]
        }, headers={'Authorization': f'Bearer {token}'})
        assert response.status_code == 200
        data = response.get_json()
        # Column F has zeros → rows found (inline rules used, not profile's column A)
        assert data['total_rows_to_delete'] > 0

        # Cleanup
        db_connection.execute("DELETE FROM filter_profiles WHERE id = 103")
        db_connection.commit()


class TestAutomatedProfileResolution:
    """Tests for profile_id resolution in process-automated route."""

    def test_automated_profile_not_found(self, client, test_user, sample_excel_file):
        """Profile not found in automated route returns 404."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process-automated/{file_id}', json={
            'profile_id': 99999
        }, headers={'Authorization': f'Bearer {token}'})
        assert response.status_code == 404

    def test_automated_profile_access_denied(self, client, test_user, sample_excel_file, db_connection):
        """Cannot use another user's profile in automated route."""
        login = client.post('/api/login', json={
            'email': test_user['email'], 'password': test_user['password']
        })
        token = login.get_json()['access_token']

        db_connection.execute(
            """INSERT INTO filter_profiles
               (id, user_id, name, filter_rules_json, is_system_template)
               VALUES (104, 99999, 'OtherUser', '[{"column":"F","value":"0"}]', 0)"""
        )
        db_connection.commit()

        with open(sample_excel_file, 'rb') as f:
            upload = client.post('/api/upload',
                                  data={'file': (f, 'test_file.xlsx')},
                                  headers={'Authorization': f'Bearer {token}'},
                                  content_type='multipart/form-data')
        assert upload.status_code in [200, 201]
        file_id = upload.get_json()['file_id']

        response = client.post(f'/api/process-automated/{file_id}', json={
            'profile_id': 104
        }, headers={'Authorization': f'Bearer {token}'})
        assert response.status_code == 403

        # Cleanup
        db_connection.execute("DELETE FROM filter_profiles WHERE id = 104")
        db_connection.commit()
