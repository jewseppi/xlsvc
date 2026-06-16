"""
Module for generating Excel deletion reports.
Shows which rows were deleted during processing.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_deletion_report(deleted_rows_data, output_path, columns_removed=None, sheets_removed=None):
    """
    Generate an Excel workbook showing what was removed during processing:
    deleted rows (one sheet each) plus a Summary of removed columns/sheets.

    Args:
        deleted_rows_data: Dict of {sheet_name: [{'row_number': int, 'data': [cell values]}]}
        output_path: Where to save the report file
        columns_removed: optional list of column letters removed entirely
        sheets_removed: optional list of sheet names/indices removed entirely

    Returns:
        output_path if successful, None if there was nothing to report
    """
    columns_removed = columns_removed or []
    sheets_removed = sheets_removed or []
    deleted_rows_data = deleted_rows_data or {}
    has_rows = any(rows for rows in deleted_rows_data.values())

    if not has_rows and not columns_removed and not sheets_removed:
        print("No deletion data to generate report")
        return None

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Summary sheet listing entire columns / sheets removed (if any).
    if columns_removed or sheets_removed:
        summary = wb.create_sheet(title="Summary")
        summary.append(["What was removed"])
        summary["A1"].font = Font(bold=True, size=12)
        summary.append([])
        summary.append(["Columns removed (entire column):",
                        ", ".join(columns_removed) if columns_removed else "None"])
        summary.append(["Sheets removed (entire tab):",
                        ", ".join(str(s) for s in sheets_removed) if sheets_removed else "None"])
        summary.cell(row=3, column=1).font = Font(bold=True)
        summary.cell(row=4, column=1).font = Font(bold=True)
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 60

    sheets_added = 0
    for sheet_name, rows in deleted_rows_data.items():
        if not rows or len(rows) == 0:
            continue
            
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
        sheets_added += 1
        
        # Header row
        max_cols = len(rows[0]['data']) if rows else 0
        headers = ['Original Row #'] + [f'Col {chr(65+i)}' if i < 26 else f'Col {i+1}' for i in range(max_cols)]
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_info in rows:
            ws.append([row_info['row_number']] + row_info['data'])
        
        # Auto-size columns (approximate)
        for idx, column in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:  # pragma: no cover
                    pass  # pragma: no cover
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    total_sheets = len(wb.sheetnames)
    wb.save(output_path)
    wb.close()

    print(f"Deletion report saved to {output_path} with {total_sheets} sheet(s)")
    return output_path

def capture_row_data(sheet, row_num, max_cols=50):
    """
    Capture all cell values from a row (for openpyxl sheets).
    
    Args:
        sheet: openpyxl worksheet
        row_num: Row number (1-based)
        max_cols: Maximum columns to capture
    
    Returns:
        List of cell values
    """
    row_data = []
    for col in range(1, max_cols + 1):
        try:
            cell_value = sheet.cell(row=row_num, column=col).value
            row_data.append(cell_value if cell_value is not None else '')
        except:
            break
    
    # Trim trailing empty cells
    while row_data and row_data[-1] == '':
        row_data.pop()
    
    return row_data